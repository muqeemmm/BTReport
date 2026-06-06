"""
Build the Patient Characteristics Table.

Inputs (all expected in the same folder as this script):
    - aku_annotations_with_duplicates.xlsx
        Source of the final cohort. The unique values of the "Subject ID"
        column give the set of patients to include.
    - master_file.xlsx
        Source of demographic / molecular / imaging metadata for each
        patient. Patients are joined using the "Center ID" column of the
        AKU file, which matches either "Local ID" or "BraTS2021" in the
        master file.
    - subjects_to_discard.txt
        One Subject ID per line. Any Subject ID listed here is excluded
        from the final cohort.

Output:
    - patient_characteristics_table.xlsx
        A formatted Patient Characteristics Table, with datasets as
        columns and each characteristic (Age, Sex, IDH, 1p/19q, Molecular
        Subtype, Tumor Grade, Multi-class Segmentation) as row groups.
    - cohort_merged.xlsx
        The merged per-patient data used to compute the table (handy for
        auditing / downstream use).

Usage:
    python build_patient_characteristics_table.py
"""

from __future__ import annotations

import os
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent

# The current cohort source. The new AKU export lives in the JSONs BTReport
# folder and splits annotations across two sheets (Single Annotations +
# Duplicates) which together yield 646 unique Subject IDs. We try several
# plausible locations because the same project folder may be reached via
# different mount points (e.g., the user's filesystem vs. a sandbox).
_NEW_FILENAME = "AKU_Data_by_Observer_CenterID.xlsx"
AKU_FILE_NEW_CANDIDATES = (
    BASE_DIR / _NEW_FILENAME,
    BASE_DIR.parent / "Dataset_AKU_WHO" / "JSONs BTReport" / _NEW_FILENAME,
    BASE_DIR.parent / "JSONs BTReport" / _NEW_FILENAME,
)
AKU_FILE_LEGACY = BASE_DIR / "aku_annotations_with_duplicates.xlsx"
AKU_SHEETS = ("Single Annotations", "Duplicates")

MASTER_FILE = BASE_DIR / "master_file.xlsx"
DISCARD_FILE = BASE_DIR / "subjects_to_discard.txt"

OUT_TABLE = BASE_DIR / "patient_characteristics_table.xlsx"
OUT_TABLE_SUPP = BASE_DIR / "patient_characteristics_table_supplementary.xlsx"
OUT_MERGED = BASE_DIR / "cohort_merged.xlsx"

# Column order in the final table (left to right).
DATASET_ORDER = [
    "UCSF-PDGM",
    "UPENN-GBM",
    "TCGA-LGG",
    "TCGA-GBM",
    "IvyGAP",
    "EGD",
]


# ---------------------------------------------------------------------------
# Data loading / filtering / merging
# ---------------------------------------------------------------------------
def load_discard_list(path: Path) -> set[str]:
    ids: set[str] = set()
    if not path.exists():
        return ids
    for line in path.read_text().splitlines():
        line = line.strip()
        if line:
            ids.add(line)
    return ids


def _load_aku_subjects() -> pd.DataFrame:
    """Load the AKU cohort.

    Prefers the new export (AKU_Data_by_Observer_CenterID.xlsx) which splits
    annotations into a 'Single Annotations' sheet plus a 'Duplicates' sheet
    (subjects annotated by more than one observer). The two sheets are
    concatenated, then deduplicated on Subject ID. Falls back to the legacy
    single-sheet file if the new one is not available.
    """
    new_path = next((p for p in AKU_FILE_NEW_CANDIDATES if p.exists()), None)
    if new_path is not None:
        print(f"[cohort] reading AKU file: {new_path}")
        single = pd.read_excel(new_path, sheet_name="Single Annotations")
        # The Duplicates sheet has a banner row ("DUPLICATES") and no header;
        # reuse the column names from the Single Annotations sheet.
        dup_raw = pd.read_excel(
            new_path, sheet_name="Duplicates", header=None, skiprows=1
        )
        if not dup_raw.empty:
            dup_raw.columns = single.columns[: dup_raw.shape[1]]
        else:
            dup_raw = pd.DataFrame(columns=single.columns)
        combined = pd.concat([single, dup_raw], ignore_index=True)
        print(
            f"[cohort] sheets: Single Annotations={len(single)} rows, "
            f"Duplicates={len(dup_raw)} rows"
        )
        return combined

    print(f"[cohort] reading legacy AKU file: {AKU_FILE_LEGACY.name}")
    return pd.read_excel(AKU_FILE_LEGACY)


def build_cohort() -> pd.DataFrame:
    """Load inputs, apply the Subject ID filter, and join to the master file."""
    aku = _load_aku_subjects()
    master = pd.read_excel(MASTER_FILE)

    # Keep one row per unique Subject ID (first occurrence) and record its Center ID.
    aku_unique = (
        aku.dropna(subset=["Subject ID"])
        .drop_duplicates(subset=["Subject ID"], keep="first")
        .loc[:, ["Subject ID", "Center ID"]]
        .copy()
    )

    # Drop permanently-discarded subjects.
    discard = load_discard_list(DISCARD_FILE)
    before = len(aku_unique)
    aku_unique = aku_unique[~aku_unique["Subject ID"].isin(discard)].copy()
    print(
        f"[cohort] unique Subject IDs: {before} -> {len(aku_unique)} "
        f"after dropping {before - len(aku_unique)} discarded subject(s)."
    )

    # Build lookup tables from the master file. A Center ID can match either
    # a Local ID or a BraTS2021 ID, so try both.
    master_by_local = (
        master.dropna(subset=["Local ID"])
        .drop_duplicates(subset=["Local ID"], keep="first")
        .set_index("Local ID")
    )
    master_by_brats = (
        master.dropna(subset=["BraTS2021"])
        .drop_duplicates(subset=["BraTS2021"], keep="first")
        .set_index("BraTS2021")
    )

    merged_rows = []
    unmatched = []
    for _, row in aku_unique.iterrows():
        cid = row["Center ID"]
        mrow = None
        if pd.notna(cid) and cid in master_by_local.index:
            mrow = master_by_local.loc[cid].to_dict()
        elif pd.notna(cid) and cid in master_by_brats.index:
            mrow = master_by_brats.loc[cid].to_dict()
        if mrow is None:
            unmatched.append((row["Subject ID"], cid))
            continue
        mrow["Subject ID"] = row["Subject ID"]
        mrow["Center ID"] = cid
        merged_rows.append(mrow)

    if unmatched:
        print(f"[cohort] WARNING: {len(unmatched)} subjects could not be matched to master_file.")
        for s, c in unmatched[:10]:
            print(f"    {s} (Center ID: {c})")

    merged = pd.DataFrame(merged_rows)
    print(f"[cohort] merged rows: {len(merged)} / master columns: {merged.shape[1]}")
    return merged


# ---------------------------------------------------------------------------
# Value normalisation helpers
# ---------------------------------------------------------------------------
def _is_missing(x) -> bool:
    """Missing values in master_file are NaN or the literal string 'x'."""
    if x is None:
        return True
    try:
        if pd.isna(x):
            return True
    except (TypeError, ValueError):
        pass
    if isinstance(x, str) and x.strip().lower() in {"", "x", "nan", "none", "unknown"}:
        return True
    return False


def _coalesce(row: pd.Series, primary: str, fallback: str):
    """Return primary value if present, else fallback, else None."""
    v = row.get(primary)
    if not _is_missing(v):
        return v
    v = row.get(fallback)
    if not _is_missing(v):
        return v
    return None


def normalise_sex(v):
    if _is_missing(v):
        return "Unknown"
    s = str(v).strip().lower()
    if s.startswith("m"):
        return "Male"
    if s.startswith("f"):
        return "Female"
    return "Unknown"


def normalise_idh(v):
    if _is_missing(v):
        return "Unknown"
    s = str(v).strip().lower()
    if "mut" in s:
        return "Mutated"
    if "wild" in s:
        return "Wildtype"
    return "Unknown"


def normalise_1p19q(v):
    if _is_missing(v):
        return "Unknown"
    s = str(v).strip().lower()
    if "co" in s and "del" in s:  # co-deleted / rel. co-deleted
        return "Co-deleted"
    if "intact" in s or "non" in s or "not" in s:
        return "Intact"
    return "Unknown"


def normalise_subtype(v, grade=None):
    """Map WHO-2021 diagnosis string to a molecular subtype.

    Grade 4 IDH-mutant astrocytomas are reported as their own subtype,
    distinct from WHO grade 2/3 IDH-mutant astrocytomas.
    """
    if _is_missing(v):
        return "Unknown"
    s = str(v).strip().lower()

    # Resolve an integer grade if possible (used to split astrocytomas).
    g = None
    if not _is_missing(grade):
        try:
            g = int(float(grade))
        except (TypeError, ValueError):
            g = None

    if "oligodendroglioma" in s:
        return "Oligodendroglioma"
    if "astrocytoma" in s and ("mutant" in s or "idh-mut" in s.replace(" ", "-")):
        return "Grade 4 IDH-mutant astrocytoma" if g == 4 else "IDH-mutant astrocytoma"
    if "glioblastoma" in s:
        return "IDH-wildtype glioblastoma"
    if "astrocytoma" in s:
        return "Grade 4 IDH-mutant astrocytoma" if g == 4 else "IDH-mutant astrocytoma"
    return "Unknown"


def normalise_grade(v):
    if _is_missing(v):
        return "Unknown"
    try:
        g = int(float(v))
    except (TypeError, ValueError):
        return "Unknown"
    if g in (2, 3, 4):
        return f"WHO grade {g}"
    return "Unknown"


def normalise_segmentation(v):
    """Multiclass Tumor Segmentation: 1 = Manual, 0 = Automatic."""
    if _is_missing(v):
        return "Unknown"
    try:
        k = int(float(v))
    except (TypeError, ValueError):
        return "Unknown"
    if k == 1:
        return "Manual"
    if k == 0:
        return "Automatic"
    return "Unknown"


# ---------------------------------------------------------------------------
# Table construction
# ---------------------------------------------------------------------------
def prepare_table_dataframe(merged: pd.DataFrame) -> pd.DataFrame:
    """Build the per-row classification columns used by the table.

    Reads the post-transform cohort (cohort_merged.xlsx, 12 columns) when
    that file exists, so molecular subtype is taken straight from WHO2021
    (with the Grade-4 IDH-mutant astrocytoma split applied here). Falls
    back to deriving values from raw master columns if only the legacy
    schema is available.
    """
    df = merged.copy()

    # ----- molecular subtype -----
    if "WHO2021" in df.columns:
        df["_Subtype"] = df.apply(
            lambda r: normalise_subtype(r.get("WHO2021"), r.get("Grade")),
            axis=1,
        )
    else:
        df["Subtype_raw"] = df.apply(
            lambda r: _coalesce(r, "WHO 2021 (original)", "WHO 2021 (generated)"),
            axis=1,
        )
        df["_Subtype"] = df.apply(
            lambda r: normalise_subtype(r["Subtype_raw"], r.get("Grade")),
            axis=1,
        )

    # ----- the rest (column names differ between cleaned vs. raw schema) -----
    sex_col = "Gender" if "Gender" in df.columns else "Gender"
    idh_col = "IDH" if "IDH" in df.columns else "IDH (original)"
    p19_col = "1p/19q" if "1p/19q" in df.columns else "1p/19q (original)"

    df["_Sex"] = df[sex_col].map(normalise_sex)
    df["_IDH"] = df[idh_col].map(normalise_idh)
    df["_1p19q"] = df[p19_col].map(normalise_1p19q)
    df["_Grade"] = df["Grade"].map(normalise_grade)
    df["_Age"] = pd.to_numeric(df["Age (years)"], errors="coerce")

    # Segmentation lives in master_file (it was dropped from the cleaned merged).
    if "Multiclass Tumor Segmentation" in df.columns:
        df["_Seg"] = df["Multiclass Tumor Segmentation"].map(normalise_segmentation)
    else:
        df["_Seg"] = "Unknown"

    # Per-row WHO 2021 / WHO 2016 labels for the supplementary table.
    df["_WHO2021"] = (
        df["WHO2021"] if "WHO2021" in df.columns else pd.Series([None] * len(df))
    )
    df["_WHO2021"] = df["_WHO2021"].where(df["_WHO2021"].notna(), "Unknown")

    if "WHO2016" in df.columns:
        df["_WHO2016"] = df["WHO2016"].where(df["WHO2016"].notna(), "Unknown")
    else:
        df["_WHO2016"] = "Unknown"
    return df


def _fmt_n_pct(n: int, total: int) -> str:
    if total <= 0:
        return "0 (0%)"
    pct = 100.0 * n / total
    return f"{n} ({pct:.0f}%)"


def _fmt_mean_sd(series: pd.Series) -> str:
    series = pd.to_numeric(series, errors="coerce").dropna()
    if series.empty:
        return "--"
    mean = series.mean()
    sd = series.std(ddof=1) if len(series) > 1 else 0.0
    return f"{mean:.1f} ± {sd:.1f}"


def _counts_by_dataset(df: pd.DataFrame, col: str, datasets: list[str]) -> dict:
    """Return {dataset: Series(value_counts)} including a TOTAL key."""
    out = {}
    for ds in datasets:
        sub = df[df["Dataset"] == ds]
        out[ds] = sub[col].value_counts(dropna=False)
    out["TOTAL"] = df[col].value_counts(dropna=False)
    return out


def _cell(counts_by_ds: dict, ds: str, value: str, totals: dict) -> str:
    n = int(counts_by_ds[ds].get(value, 0))
    return _fmt_n_pct(n, totals[ds])


def build_table_rows(df: pd.DataFrame, datasets: list[str]) -> tuple[list[list], dict]:
    """Return a list of rows ready to be written into the worksheet.

    Sub-rows labelled ``Unknown`` whose counts are zero in every column
    (including TOTAL) are omitted to keep the table tight.
    """
    # Dataset-level totals (including TOTAL column).
    totals = {ds: int((df["Dataset"] == ds).sum()) for ds in datasets}
    totals["TOTAL"] = int(len(df))

    all_ds = datasets + ["TOTAL"]
    header = ["Characteristic"] + [f"{ds} (n={totals[ds]})" for ds in all_ds]

    rows: list[list] = [header]

    def _count_row(label: str, counts: dict) -> tuple[list, int]:
        """Build one sub-row and return (row, total-count-across-columns)."""
        row = [f"  {label}"]
        total_n = 0
        for ds in all_ds:
            n = int(counts[ds].get(label, 0)) if ds != "TOTAL" else int(counts["TOTAL"].get(label, 0))
            total_n += n
            row.append(_fmt_n_pct(n, totals[ds]))
        return row, total_n

    def _append(label: str, counts: dict, row_list: list) -> None:
        """Append a sub-row, skipping Unknown rows that are zero everywhere."""
        row, total_n = _count_row(label, counts)
        if label == "Unknown" and total_n == 0:
            return
        row_list.append(row)

    # ---- Age (years) ----
    rows.append(["Age (years)"] + [""] * len(all_ds))

    known_row = ["  Known (mean ± SD)"]
    for ds in all_ds:
        sub = df if ds == "TOTAL" else df[df["Dataset"] == ds]
        known_row.append(_fmt_mean_sd(sub["_Age"]))
    rows.append(known_row)

    # Age Unknown – only keep if any dataset has missing ages.
    unknown_row = ["  Unknown"]
    total_unknown = 0
    for ds in all_ds:
        sub = df if ds == "TOTAL" else df[df["Dataset"] == ds]
        n_unknown = int(sub["_Age"].isna().sum())
        total_unknown += n_unknown
        unknown_row.append(_fmt_n_pct(n_unknown, totals[ds]))
    if total_unknown > 0:
        rows.append(unknown_row)

    # ---- Sex ----
    rows.append(["Sex"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_Sex", datasets)
    for label in ["Female", "Male", "Unknown"]:
        _append(label, counts, rows)

    # ---- IDH ----
    rows.append(["IDH"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_IDH", datasets)
    for label in ["Mutated", "Wildtype", "Unknown"]:
        _append(label, counts, rows)

    # ---- 1p/19q ----
    rows.append(["1p/19q"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_1p19q", datasets)
    for label in ["Co-deleted", "Intact", "Unknown"]:
        _append(label, counts, rows)

    # ---- Molecular Subtype ----
    rows.append(["Molecular Subtype"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_Subtype", datasets)
    for label in [
        "Oligodendroglioma",
        "IDH-mutant astrocytoma",
        "Grade 4 IDH-mutant astrocytoma",
        "IDH-wildtype glioblastoma",
        "Unknown",
    ]:
        _append(label, counts, rows)

    # ---- Tumor Grade ----
    rows.append(["Tumor Grade"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_Grade", datasets)
    for label in ["WHO grade 2", "WHO grade 3", "WHO grade 4", "Unknown"]:
        _append(label, counts, rows)

    # ---- Multi-class Segmentation ----
    rows.append(["Multi-class Segmentation"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_Seg", datasets)
    for label in ["Manual", "Automatic", "Unknown"]:
        _append(label, counts, rows)

    return rows, totals


# Labels that appear under each WHO heading in the supplementary table.
WHO2021_LABELS = [
    "Oligodendroglioma, IDH-mutant, 1p/19q-codeleted",
    "Astrocytoma, IDH-mutant",
    "Glioblastoma, IDH-wildtype",
    "Unknown",
]

WHO2016_LABELS = [
    "Oligodendroglioma, IDH-mutant, 1p/19q-codeleted",
    "Oligodendroglioma",
    "Oligoastrocytoma",
    "Astrocytoma, IDH-mutant",
    "Glioblastoma, IDH-mutant",
    "Glioblastoma, IDH-wildtype",
    "Unknown",
]


def build_supplementary_table_rows(
    df: pd.DataFrame, datasets: list[str]
) -> tuple[list[list], dict]:
    """Supplementary version: same patient characteristics but with WHO 2021
    and WHO 2016 classifications as separate row groups (Molecular Subtype
    is dropped because it is replaced by these two groups).
    """
    totals = {ds: int((df["Dataset"] == ds).sum()) for ds in datasets}
    totals["TOTAL"] = int(len(df))

    all_ds = datasets + ["TOTAL"]
    header = ["Characteristic"] + [f"{ds} (n={totals[ds]})" for ds in all_ds]
    rows: list[list] = [header]

    def _count_row(label: str, counts: dict) -> tuple[list, int]:
        row = [f"  {label}"]
        total_n = 0
        for ds in all_ds:
            n = (
                int(counts["TOTAL"].get(label, 0))
                if ds == "TOTAL"
                else int(counts[ds].get(label, 0))
            )
            total_n += n
            row.append(_fmt_n_pct(n, totals[ds]))
        return row, total_n

    def _append(label: str, counts: dict, row_list: list) -> None:
        row, total_n = _count_row(label, counts)
        if label == "Unknown" and total_n == 0:
            return
        row_list.append(row)

    # ---- Age ----
    rows.append(["Age (years)"] + [""] * len(all_ds))
    known_row = ["  Known (mean ± SD)"]
    for ds in all_ds:
        sub = df if ds == "TOTAL" else df[df["Dataset"] == ds]
        known_row.append(_fmt_mean_sd(sub["_Age"]))
    rows.append(known_row)

    unknown_row = ["  Unknown"]
    total_unknown = 0
    for ds in all_ds:
        sub = df if ds == "TOTAL" else df[df["Dataset"] == ds]
        n_unknown = int(sub["_Age"].isna().sum())
        total_unknown += n_unknown
        unknown_row.append(_fmt_n_pct(n_unknown, totals[ds]))
    if total_unknown > 0:
        rows.append(unknown_row)

    # ---- Sex ----
    rows.append(["Sex"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_Sex", datasets)
    for label in ["Female", "Male", "Unknown"]:
        _append(label, counts, rows)

    # ---- IDH ----
    rows.append(["IDH"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_IDH", datasets)
    for label in ["Mutated", "Wildtype", "Unknown"]:
        _append(label, counts, rows)

    # ---- 1p/19q ----
    rows.append(["1p/19q"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_1p19q", datasets)
    for label in ["Co-deleted", "Intact", "Unknown"]:
        _append(label, counts, rows)

    # ---- WHO 2021 Classification ----
    rows.append(["WHO 2021 Classification"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_WHO2021", datasets)
    for label in WHO2021_LABELS:
        _append(label, counts, rows)

    # ---- WHO 2016 Classification ----
    rows.append(["WHO 2016 Classification"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_WHO2016", datasets)
    for label in WHO2016_LABELS:
        _append(label, counts, rows)

    # ---- Tumor Grade ----
    rows.append(["Tumor Grade"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_Grade", datasets)
    for label in ["WHO grade 2", "WHO grade 3", "WHO grade 4", "Unknown"]:
        _append(label, counts, rows)

    # ---- Multi-class Segmentation ----
    rows.append(["Multi-class Segmentation"] + [""] * len(all_ds))
    counts = _counts_by_dataset(df, "_Seg", datasets)
    for label in ["Manual", "Automatic", "Unknown"]:
        _append(label, counts, rows)

    return rows, totals


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
def write_table_xlsx(
    rows: list[list],
    totals: dict,
    datasets: list[str],
    out_path: Path,
    title: str | None = None,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Characteristics"

    # Title row.
    n_cols = len(rows[0])
    if title is None:
        title = f"Table 1. Patient Characteristics Table (N = {totals['TOTAL']})"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(name="Times New Roman", size=13, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Header + body.
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", start_color="D9E1F2")
    group_fill = PatternFill("solid", start_color="F2F2F2")

    group_labels = {
        "Age (years)",
        "Sex",
        "IDH",
        "1p/19q",
        "Molecular Subtype",
        "WHO 2021 Classification",
        "WHO 2016 Classification",
        "Tumor Grade",
        "Multi-class Segmentation",
    }

    start_row = 2  # header row in worksheet
    for r_idx, row in enumerate(rows):
        ws_row = start_row + r_idx
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=ws_row, column=c_idx, value=value)
            cell.border = border
            cell.font = Font(name="Times New Roman", size=11)
            if c_idx == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            if r_idx == 0:
                cell.fill = header_fill
                cell.font = Font(name="Times New Roman", size=11, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif row[0] in group_labels and c_idx == 1:
                cell.fill = group_fill
                cell.font = Font(name="Times New Roman", size=11, bold=True)
            elif row[0] in group_labels:
                cell.fill = group_fill

    # Column widths.
    ws.column_dimensions["A"].width = 34
    for c in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18

    ws.row_dimensions[1].height = 22

    # Footnote.
    foot_row = start_row + len(rows) + 1
    footnote = (
        "Values are N (%), except Age which is reported as mean ± SD. "
        "Cohort defined by the unique Subject IDs in AKU_Data_by_Observer_CenterID.xlsx "
        "(Single Annotations + Duplicates sheets); subjects listed in "
        "subjects_to_discard.txt are excluded. Metadata is sourced from "
        "master_file.xlsx, joined via the AKU Center ID."
    )
    ws.cell(row=foot_row, column=1, value=footnote)
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=n_cols)
    foot_cell = ws.cell(row=foot_row, column=1)
    foot_cell.font = Font(name="Times New Roman", size=9, italic=True)
    foot_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[foot_row].height = 40

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    merged_raw = build_cohort()

    # Apply the WHO2016 / WHO2021 derivation + column cleanup. The transform
    # script defines the rules in one place; we reuse them here so the table
    # is computed from the same cleaned columns the user audits in
    # cohort_merged.xlsx.
    import importlib.util

    transform_path = BASE_DIR / "transform_cohort_merged.py"
    spec = importlib.util.spec_from_file_location("transform_cohort_merged", transform_path)
    transform_mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(transform_mod)

    augmented = merged_raw.copy()
    augmented["WHO2021"] = augmented.apply(transform_mod.fill_who2021, axis=1)
    augmented["WHO2016"] = augmented.apply(transform_mod.derive_who2016, axis=1)

    # Build a tidy 12-column "cleaned" cohort matching the spec, and write it
    # to disk; the table is computed from `augmented` which still has the
    # extra columns (e.g., Multiclass Tumor Segmentation) needed for the
    # remaining rows.
    cleaned = augmented.copy()
    cleaned = cleaned.drop(columns=[c for c in cleaned.columns if c.endswith("(generated)")])
    rename_map = {c: c.rsplit("(original)", 1)[0].strip() for c in cleaned.columns if c.endswith("(original)")}
    cleaned = cleaned.rename(columns=rename_map)
    cleaned = cleaned.drop(columns=[c for c in ("Tumor Subtype",) if c in cleaned.columns])
    final_cols = [
        "Center ID", "Subject ID", "Dataset", "Hospital", "BraTS2021",
        "Gender", "Age (years)", "IDH", "1p/19q", "Grade", "WHO2016", "WHO2021",
    ]
    cleaned = cleaned.loc[:, final_cols]
    cleaned.to_excel(OUT_MERGED, index=False)
    print(f"[output] wrote cleaned cohort to {OUT_MERGED.name} ({cleaned.shape})")

    df = prepare_table_dataframe(augmented)

    # Keep only datasets that actually appear in the cohort, in the
    # preferred order, then append any stragglers alphabetically.
    present = [ds for ds in DATASET_ORDER if (df["Dataset"] == ds).any()]
    extras = sorted(set(df["Dataset"].dropna().unique()) - set(present))
    datasets = present + extras

    rows, totals = build_table_rows(df, datasets)
    write_table_xlsx(rows, totals, datasets, OUT_TABLE)
    print(f"[output] wrote Patient Characteristics Table to {OUT_TABLE.name}")

    supp_rows, _ = build_supplementary_table_rows(df, datasets)
    write_table_xlsx(
        supp_rows,
        totals,
        datasets,
        OUT_TABLE_SUPP,
        title=(
            f"Supplementary Table. Patient Characteristics with WHO 2021 and "
            f"WHO 2016 Classifications (N = {totals['TOTAL']})"
        ),
    )
    print(f"[output] wrote Supplementary Table to {OUT_TABLE_SUPP.name}")
    print(f"[output] total patients in table: {totals['TOTAL']}")
    for ds in datasets:
        print(f"    {ds}: {totals[ds]}")


if __name__ == "__main__":
    main()
