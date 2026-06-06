"""
Glioma classification via Groq (openai/gpt-oss-120b).

For each subject metadata JSON in the JSONs folder (btreport or btreport_pp
variant, selected by --json-type), predicts IDH status, 1p/19q co-deletion,
and CNS WHO grade following WHO CNS5 (2021).

Output is constrained to `glioma_classification_schema.json` (same folder)
and the prompt mirrors `glioma_classification_prompt_base.pdf`.

After inference, computes Accuracy, Sensitivity, Specificity, and F1 per task
against ground-truth labels in cohort_merged.xlsx — reported for the whole
cohort and per dataset.

Confidence elicitation follows the Chain-of-Thought Confidence Elicitation
(CoT CE) strategy of Tian et al. (2023), as benchmarked in Ren et al.,
"Towards Reliable Medical LLMs" (arXiv:2601.15645): the model must produce
its clinical reasoning BEFORE committing to a prediction and a confidence
score, so the chain of thought informs the score rather than rationalising
a number already chosen. The schema field order (reasoning -> features ->
prediction -> confidence) enforces this generation order under strict
structured output.
"""

import argparse
import csv
import json
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from colorama import Fore, Style, init as colorama_init
from dotenv import dotenv_values
from groq import APIStatusError, BadRequestError, Groq, RateLimitError
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score
from tqdm import tqdm


# --- Prompt (transcribed from glioma_classification_prompt_base.pdf,
# ---          adapted for Chain-of-Thought Confidence Elicitation) -----------

SYSTEM_PROMPT = (
    "You are an expert neuroradiologist performing pre-operative molecular "
    "subtyping of adult-type diffuse glioma from structured MRI metadata, "
    "following the WHO CNS5 (2021) classification. For every task you reason "
    "step by step BEFORE stating any prediction or confidence score: first lay "
    "out the clinically grounded explanation, then commit to the prediction, "
    "then assign the confidence. Output MUST be a single valid JSON object that "
    "conforms exactly to the GliomaClassificationOutput schema. No markdown, no "
    "prose outside the JSON, no code fences -- JSON only."
)

USER_PROMPT_TEMPLATE = """# ROLE
You are an expert neuroradiologist.

# TASK
You are given structured pre-operative brain MRI metadata for a single patient
with a suspected **adult-type diffuse glioma**, defined according to the WHO
CNS5 molecular classification. Based on this metadata, predict the following:

  (a) IDH mutation status        -> {{"mutant", "wildtype"}}
  (b) 1p/19q co-deletion status  -> {{"codeleted", "non-codeleted"}}
  (c) CNS WHO grade              -> {{2, 3, 4}}

# CHAIN-OF-THOUGHT CONFIDENCE ELICITATION
For each of the three tasks you MUST work in the following order. Do not skip,
reorder, or shortcut these steps:

  1. **Reasoning first.** Write a concise, clinically grounded explanation that
     weighs the available metadata. Explicitly contrast the evidence that
     supports each candidate answer against the evidence that argues against
     it. Reach the prediction as the *conclusion* of this reasoning.
  2. **Supporting metadata fields.** List the exact field names (verbatim) that
     support the prediction you reasoned toward.
  3. **Contradicting features.** List the fields that argue against it.
  4. **Prediction.** State the prediction that follows from steps 1-3.
  5. **Confidence score** (between 0.0 and 1.0). Derive the score *from* the
     reasoning above -- it must reflect the balance of supporting vs.
     contradicting evidence you already articulated, NOT a number chosen before
     reasoning. Assign high confidence only when key features are well
     supported and few contradict; assign low confidence when evidence is
     thin, conflicting, or absent.

The JSON schema lists `reasoning`, `supporting_features` and
`contradicting_features` before `prediction` and `confidence` for exactly this
reason: generate them in that order.

# INPUT CONSTRAINTS
- Input consists **only of structured MRI-derived metadata** (no images).
- Do not assume access to data beyond what is provided.
- Treat null values as **"not assessed"** (do not infer).

# AVAILABLE MRI SEQUENCES
T1w, T2w, T2-FLAIR, T1-Gd.
**Not available:** DWI/ADC, perfusion (rCBV/DSC), MRS (2HG), SWI/GRE.

# RULES
1. Use ONLY the metadata provided; no hallucinated features or sequences.
2. Every claim in `reasoning` must be traceable to a field named verbatim in
   `supporting_features`.
3. Do NOT invoke diffusion, perfusion, spectroscopy, calcifications, or SWI.
4. If evidence is insufficient: give a best estimate, set confidence in
   0.50-0.60, and state the uncertainty in reasoning.
5. Confidence must be >= 0.50 for binary tasks (otherwise flip the prediction).
6. The confidence score must be the conclusion of the reasoning, not its
   premise -- never write the reasoning to justify a pre-chosen score.
7. Output must be **valid JSON only** (no markdown, no commentary).

# INPUT
Patient subject_id: {subject_id}
MRI metadata:
{metadata_json}
"""


# --- Config ------------------------------------------------------------------

MODEL_ID    = "openai/gpt-oss-120b"
HERE        = Path(__file__).parent
SCHEMA_PATH = HERE / "glioma_classification_schema.json"
ENV_PATH    = Path(".env")
JSONS_DIR   = Path("data/Dataset_AKU_WHO/JSONs")
COHORT_XLSX = Path("data/dataset_table/cohort_merged.xlsx")


# --- ID helpers --------------------------------------------------------------

def load_id_map(xlsx_path: Path) -> dict:
    """
    Returns a dict keyed by any known folder name (Center ID or BraTS2021 ID)
    mapping to {"center_id": str, "subject_id": str} where:
      - center_id  = column 0 ("Center ID")   → used for output filenames
      - subject_id = column 1 ("Subject ID")  → passed to the API (redacted form)
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    id_map: dict = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        center_id  = str(row[0]).strip() if row[0] else None
        subject_id = str(row[1]).strip() if row[1] else None
        brats_id   = str(row[4]).strip() if (row[4] and row[4] != "x") else None
        entry = {"center_id": center_id, "subject_id": subject_id}
        if center_id:
            id_map[center_id] = entry
        if brats_id:
            id_map[brats_id] = entry
    return id_map


# --- Ground-truth helpers ----------------------------------------------------

def load_ground_truth(xlsx_path: Path) -> dict:
    """
    Returns a dict keyed by subject_id with:
      {"idh": str|None, "codeletion": str|None, "grade": int|None, "dataset": str}

    GT label normalisation:
      IDH:    "mutated"               -> "mutant"
              "wild type"             -> "wildtype"
      1p/19q: "co-deleted" /
              "rel. co-deleted"       -> "codeleted"
              "intact"                -> "non-codeleted"
              "x"                     -> None  (not assessed)
      Grade:  int as-is
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # headers: Center ID(0), Subject ID(1), Dataset(2), Hospital(3),
    #          BraTS2021(4), Gender(5), Age(6), IDH(7), 1p/19q(8), Grade(9)

    def _idh(v):
        if v == "mutated":   return "mutant"
        if v == "wild type": return "wildtype"
        return None

    def _codeletion(v):
        if v in ("co-deleted", "rel. co-deleted"): return "codeleted"
        if v == "intact":                          return "non-codeleted"
        return None  # "x" or unknown

    gt: dict = {}
    for row in rows[1:]:
        center_id = row[0]
        brats_id  = row[4] if (row[4] and row[4] != "x") else None
        entry = {
            "idh":        _idh(row[7]),
            "codeletion": _codeletion(row[8]),
            "grade":      row[9] if isinstance(row[9], int) else None,
            "dataset":    row[2],
        }
        if center_id:
            gt[center_id] = entry
        if brats_id:
            gt[brats_id] = entry

    return gt


# --- Metrics -----------------------------------------------------------------

def _binary_metrics(y_true, y_pred, pos_label) -> dict:
    """Accuracy, Sensitivity, Specificity, F1 for a binary task."""
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None, "weighted_f1": None}
    all_labels = sorted(set(y_true) | set(y_pred))
    neg_candidates = [l for l in all_labels if l != pos_label]
    # Fallback: derive neg_label from the known label space when only one class present
    neg_label = neg_candidates[0] if neg_candidates else (
        "wildtype" if pos_label == "mutant" else "non-codeleted"
    )
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, pos_label=pos_label, average="binary",
                   zero_division=0)
    wf1 = f1_score(y_true, y_pred, average="weighted", zero_division=0)
    cm  = confusion_matrix(y_true, y_pred, labels=[pos_label, neg_label])
    TP, FN = cm[0, 0], cm[0, 1]
    FP, TN = cm[1, 0], cm[1, 1]
    sens = TP / (TP + FN) if (TP + FN) > 0 else None
    spec = TN / (TN + FP) if (TN + FP) > 0 else None
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sens, 4) if sens is not None else None,
        "specificity": round(spec, 4) if spec is not None else None,
        "f1":          round(f1, 4),
        "weighted_f1": round(wf1, 4),
    }


def _grade_metrics(y_true, y_pred) -> dict:
    """Macro-averaged Accuracy, Sensitivity, Specificity, F1 for grade (2/3/4)."""
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None, "weighted_f1": None}
    classes = [2, 3, 4]
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, labels=classes, average="macro",
                   zero_division=0)
    wf1 = f1_score(y_true, y_pred, labels=classes, average="weighted",
                   zero_division=0)
    cm = confusion_matrix(y_true, y_pred, labels=classes)
    sens_list, spec_list = [], []
    for i in range(len(classes)):
        TP = cm[i, i]
        FN = cm[i, :].sum() - TP
        FP = cm[:, i].sum() - TP
        TN = cm.sum() - TP - FN - FP
        sens_list.append(TP / (TP + FN) if (TP + FN) > 0 else 0.0)
        spec_list.append(TN / (TN + FP) if (TN + FP) > 0 else 0.0)
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sum(sens_list) / len(sens_list), 4),
        "specificity": round(sum(spec_list) / len(spec_list), 4),
        "f1":          round(f1, 4),
        "weighted_f1": round(wf1, 4),
    }


def compute_metrics(predictions: dict, gt: dict) -> dict:
    """
    Returns metrics per task broken down by whole cohort and per dataset.
      {
        "whole_cohort": {"idh": {...}, "codeletion": {...}, "grade": {...}},
        "<dataset>":    {"idh": {...}, ...},
        ...
      }
    """
    data = defaultdict(lambda: {
        "idh_t": [], "idh_p": [],
        "cod_t": [], "cod_p": [],
        "grd_t": [], "grd_p": [],
    })

    for sid, pred in predictions.items():
        row = gt.get(sid)
        if row is None:
            continue
        ds = row["dataset"] or "Unknown"

        for bucket in ("whole_cohort", ds):
            d = data[bucket]
            if row["idh"] is not None and pred.get("idh") is not None:
                d["idh_t"].append(row["idh"])
                d["idh_p"].append(pred["idh"])
            if row["codeletion"] is not None and pred.get("codeletion") is not None:
                d["cod_t"].append(row["codeletion"])
                d["cod_p"].append(pred["codeletion"])
            if row["grade"] is not None and pred.get("grade") is not None:
                d["grd_t"].append(row["grade"])
                d["grd_p"].append(pred["grade"])

    results = {}
    for bucket, d in data.items():
        results[bucket] = {
            "idh":        _binary_metrics(d["idh_t"], d["idh_p"], "mutant"),
            "codeletion": _binary_metrics(d["cod_t"], d["cod_p"], "codeleted"),
            "grade":      _grade_metrics(d["grd_t"], d["grd_p"]),
        }
    return results


def print_metrics(metrics: dict) -> None:
    col_w = 22
    header = (f"{'Subset':<{col_w}} {'Task':<12} {'N':>5}  "
              f"{'Acc':>6}  {'Sens':>6}  {'Spec':>6}  {'F1':>6}  {'WF1':>6}")
    sep = "=" * len(header)
    print(f"\n{sep}\nCLASSIFICATION METRICS\n{sep}")
    print(header)
    print("-" * len(header))

    task_labels = {"idh": "IDH", "codeletion": "1p/19q", "grade": "Grade"}
    buckets = ["whole_cohort"] + sorted(k for k in metrics if k != "whole_cohort")

    for bucket in buckets:
        for task_key, label in task_labels.items():
            m = metrics[bucket][task_key]
            def _fmt(v):
                return f"{v:.4f}" if v is not None else "  N/A"
            print(f"{bucket:<{col_w}} {label:<12} {m['n']:>5}  "
                  f"{_fmt(m['accuracy']):>6}  {_fmt(m['sensitivity']):>6}  "
                  f"{_fmt(m['specificity']):>6}  {_fmt(m['f1']):>6}  "
                  f"{_fmt(m['weighted_f1']):>6}")
        print()


# --- Run ---------------------------------------------------------------------

def compute_and_save_metrics(output_dir) -> None:
    """Score already-saved classification outputs (no API calls)."""
    C, G, D = Fore.CYAN, Fore.GREEN, Style.DIM
    gt = load_ground_truth(COHORT_XLSX)
    print(f"\n{D}" + "─" * 72 + Style.RESET_ALL)
    print(f"{C}Collecting predictions for metric computation...{Style.RESET_ALL}")
    predictions: dict = {}
    for out_file in sorted(output_dir.glob("*_classification_gpt_oss_120b.json")):
        result = json.load(open(out_file))
        sid = result.get("subject_id",
                          out_file.stem.replace("_classification_gpt_oss_120b", ""))
        predictions[sid] = {
            "idh":        result.get("idh", {}).get("prediction"),
            "codeletion": result.get("one_p_nineteen_q", {}).get("prediction"),
            "grade":      result.get("who_grade", {}).get("prediction"),
        }

    metrics = compute_metrics(predictions, gt)
    print_metrics(metrics)

    metrics_path = output_dir / "metrics.json"
    with open(metrics_path, "w") as f:
        json.dump(metrics, f, indent=2)
    print(f"{G}Metrics saved{Style.RESET_ALL} → {metrics_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Classify gliomas via Groq on btreport or btreport_pp JSONs."
    )
    parser.add_argument(
        "--json-type",
        choices=["btreport", "btreport_pp"],
        required=True,
        help="Which metadata JSON variant to use (btreport or btreport_pp).",
    )
    parser.add_argument(
        "--metrics-only",
        action="store_true",
        help="Skip inference (no API calls) and only recompute metrics from "
             "already-saved outputs.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    colorama_init(autoreset=True)

    # Colour aliases (autoreset means no explicit RESET needed after each use)
    C = Fore.CYAN
    G = Fore.GREEN
    Y = Fore.YELLOW
    R = Fore.RED
    M = Fore.MAGENTA
    W = Style.BRIGHT       # bold/bright white for headers
    D = Style.DIM

    args      = parse_args()
    json_type = args.json_type  # "btreport" or "btreport_pp"

    OUTPUT_DIR = HERE / f"groq_outputs_{json_type}"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Metrics-only mode: no clients, no inference, no API calls.
    if args.metrics_only:
        compute_and_save_metrics(OUTPUT_DIR)
        raise SystemExit(0)

    TIMINGS_CSV = OUTPUT_DIR / "timings.csv"
    _csv_is_new = not TIMINGS_CSV.exists()
    _timings_fh = TIMINGS_CSV.open("a", newline="")
    _timings_writer = csv.DictWriter(
        _timings_fh,
        fieldnames=["folder_name", "elapsed_s", "timestamp"],
    )
    if _csv_is_new:
        _timings_writer.writeheader()
        _timings_fh.flush()

    schema = json.load(open(SCHEMA_PATH))
    config = dotenv_values(ENV_PATH)

    # Build a pool of clients from all GROQ_API_KEY_* entries in .env.
    # Keys are tried in order; on RateLimitError the next key is used.
    api_keys = [
        v for k, v in sorted(config.items())
        if k.startswith("GROQ_API_KEY_") and v
    ]
    if not api_keys:
        raise RuntimeError("No GROQ_API_KEY_* entries found in .env")
    clients = [Groq(api_key=k) for k in api_keys]
    key_idx = 0  # index of the currently active client
    print(f"{C}Loaded {W}{len(clients)}{C} Groq API key(s).  "
          f"Model: {W}{MODEL_ID}{C}  |  Output: {W}{OUTPUT_DIR.name}")

    id_map = load_id_map(COHORT_XLSX)

    subjects = sorted(d.name for d in JSONS_DIR.iterdir() if d.is_dir())

    pbar = tqdm(
        subjects,
        desc=f"{C}Classifying{Style.RESET_ALL}",
        unit="subj",
        ncols=90,
        bar_format=(
            "{desc}: {percentage:3.0f}%|{bar:35}| "
            "{n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]"
        ),
        colour="cyan",
    )

    DIVIDER = f"{D}" + "─" * 72 + Style.RESET_ALL

    for i, folder_name in enumerate(pbar, 1):
        pbar.set_postfix_str(f"{folder_name[:28]}", refresh=True)

        # Blank line + divider before each subject block for readability
        tqdm.write("")
        tqdm.write(DIVIDER)
        tqdm.write(
            f"{W}[{i}/{len(subjects)}]{Style.RESET_ALL}  "
            f"{C}{folder_name}{Style.RESET_ALL}"
        )

        # Resolve Center ID (for filename) and Subject ID (for API prompt).
        ids = id_map.get(folder_name)
        if ids:
            center_id      = ids["center_id"]
            api_subject_id = ids["subject_id"]
        else:
            tqdm.write(
                f"  {Y}WARN{Style.RESET_ALL}  no xlsx entry for {folder_name!r} "
                f"— using folder name as fallback"
            )
            center_id      = folder_name
            api_subject_id = folder_name

        metadata_path = JSONS_DIR / folder_name / f"{folder_name}_metadata_{json_type}.json"
        if not metadata_path.exists():
            tqdm.write(f"  {Y}SKIP{Style.RESET_ALL}  metadata not found: {metadata_path.name}")
            continue

        out_path = OUTPUT_DIR / f"{folder_name}_classification_gpt_oss_120b.json"
        if out_path.exists():
            tqdm.write(f"  {Y}SKIP{Style.RESET_ALL}  output already exists: {out_path.name}")
            continue

        tqdm.write(
            f"  {D}center_id={Style.RESET_ALL}{center_id}   "
            f"{D}api_id={Style.RESET_ALL}{api_subject_id}"
        )

        metadata    = json.load(open(metadata_path))
        user_prompt = USER_PROMPT_TEMPLATE.format(
            subject_id=api_subject_id,
            metadata_json=json.dumps(metadata, indent=2, default=str),
        )
        user_prompt_compact = USER_PROMPT_TEMPLATE.format(
            subject_id=api_subject_id,
            metadata_json=json.dumps(metadata, default=str),
        )

        # Two-level retry:
        #   outer loop — rotate API key on RateLimitError / too-large (413)
        #   inner loop — retry same key on json_validate_failed (BadRequestError)
        MAX_JSON_RETRIES = 3
        response = None
        skip_subject = False
        use_compact = False
        t0 = time.perf_counter()

        for _ in range(len(clients)):
            current_prompt = user_prompt_compact if use_compact else user_prompt
            for json_try in range(1, MAX_JSON_RETRIES + 1):
                try:
                    response = clients[key_idx].chat.completions.create(
                        model=MODEL_ID,
                        messages=[
                            {"role": "system", "content": SYSTEM_PROMPT},
                            {"role": "user",   "content": current_prompt},
                        ],
                        temperature=0.0,
                        seed=42,
                        max_completion_tokens=4096,
                        reasoning_effort="medium",
                        response_format={
                            "type": "json_schema",
                            "json_schema": {
                                "name": "GliomaClassificationOutput",
                                "schema": schema,
                                "strict": True,
                            },
                        },
                    )
                    break  # success → exit inner loop
                except RateLimitError as e:
                    next_idx = (key_idx + 1) % len(clients)
                    tqdm.write(
                        f"  {M}LIMIT{Style.RESET_ALL}  key {key_idx + 1}/{len(clients)} "
                        f"exhausted → switching to key {next_idx + 1}  "
                        f"{D}({e}){Style.RESET_ALL}"
                    )
                    key_idx = next_idx
                    time.sleep(2)
                    break  # rotate key → exit inner loop, continue outer
                except BadRequestError as e:
                    if "json_validate_failed" not in str(e):
                        raise
                    tqdm.write(
                        f"  {Y}RETRY{Style.RESET_ALL}  JSON schema validation failed "
                        f"(attempt {json_try}/{MAX_JSON_RETRIES}) — retrying same key"
                    )
                    time.sleep(1)
                    # continue inner loop for next json_try
                except APIStatusError as e:
                    if e.status_code != 413:
                        raise
                    next_idx = (key_idx + 1) % len(clients)
                    tqdm.write(
                        f"  {M}LIMIT{Style.RESET_ALL}  key {key_idx + 1}/{len(clients)} "
                        f"request too large (413) → compact JSON + switching to key "
                        f"{next_idx + 1}  {D}({e}){Style.RESET_ALL}"
                    )
                    key_idx = next_idx
                    use_compact = True
                    current_prompt = user_prompt_compact
                    time.sleep(1)
                    break  # rotate key → exit inner loop, continue outer
            else:
                # Inner for-else: exhausted all JSON retries without a break
                tqdm.write(
                    f"  {R}ERROR{Style.RESET_ALL}  JSON validation failed after "
                    f"{MAX_JSON_RETRIES} attempts — skipping {folder_name}"
                )
                skip_subject = True
                break  # no point rotating keys for a schema error

            if response is not None or skip_subject:
                break  # success or terminal failure → exit outer loop

        if response is None:
            if not skip_subject:
                tqdm.write(
                    f"  {R}ERROR{Style.RESET_ALL}  all {len(clients)} keys rate-limited "
                    f"— skipping {folder_name}"
                )
            continue

        elapsed = round(time.perf_counter() - t0, 2)

        result = json.loads(response.choices[0].message.content)
        # Store folder_name so the metrics stage can join against ground truth.
        result["subject_id"] = folder_name

        with open(out_path, "w") as f:
            json.dump(result, f, indent=2)

        _timings_writer.writerow({
            "folder_name": folder_name,
            "elapsed_s":   elapsed,
            "timestamp":   datetime.now().isoformat(timespec="seconds"),
        })
        _timings_fh.flush()

        tqdm.write(f"  {G}SAVED{Style.RESET_ALL}  {out_path.name}  {D}({elapsed}s){Style.RESET_ALL}")

    pbar.close()
    _timings_fh.close()
    print(f"{C}Timings saved{Style.RESET_ALL} → {TIMINGS_CSV}")

    # --- Metrics -----------------------------------------------------------------
    compute_and_save_metrics(OUTPUT_DIR)
