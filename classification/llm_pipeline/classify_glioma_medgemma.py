"""
Glioma classification via MedGemma-27B (Google) (Gemma-3 27B instruction-tuned for medical text)
served through an OpenAI-compatible chat-completions API.

This is the MedGemma-27B (Google) counterpart of `classify_glioma_deepseek.py`. It performs
the SAME task -- joint prediction of IDH status, 1p/19q co-deletion and CNS WHO grade
from structured pre-operative MRI metadata under WHO CNS5 (2021) -- using the SAME
Chain-of-Thought Confidence Elicitation (CoT CE) prompt, schema, metrics, and the SAME
structure-retry / content-based extraction machinery, so MedGemma-27B (Google) is directly
comparable to DeepSeek, Llama 4 and the other models in this folder.

ACCESS -- two OpenAI-compatible providers, picked with PROVIDER in .env
----------------------------------------------------------------------
1) featherless  -- flat-rate subscription serverless. Serves the HF repo directly
   under its repo id (no GPU rental). Caveat: these lower-traffic medical models are
   not kept "warm", so Featherless may cold-start them on demand and can return
   503 capacity_exhausted / 400 model_not_deployed until a slot frees up (the request
   loop waits these out). Set the repo id as MODEL_ID (defaults are pre-filled).

2) deepinfra    -- these four medical models are NOT in DeepInfra's serverless
   catalog, so you must first create a **Custom LLM deployment** (one-time):
     a. Go to https://deepinfra.com/dash/deployments -> Deploy a custom LLM.
     b. Point it at the Hugging Face repo `google/medgemma-27b-text-it`, pick a GPU
        (A100-80GB or H100-80GB) and save. DeepInfra gives the deployment an id of
        the form  <your-deepinfra-username>/<deployment-name>.
     c. Put that deployment id in MODEL_ID (REQUIRED for deepinfra; the HF repo id
        alone will NOT resolve).
   NOTE ON BILLING: a Custom LLM is a DEDICATED, autoscaling GPU deployment billed by
   runtime (it can scale to zero when idle), i.e. it is effectively pay-per-GPU-use,
   not per-token. This trades the no-GPU-rental property for reliable availability.

    PROVIDER             = featherless | deepinfra        # default featherless
    FEATHERLESS_API_KEY  = fw-...                          # if PROVIDER=featherless
    DEEPINFRA_API_KEY    = ...                             # if PROVIDER=deepinfra
    MODEL_ID             = <repo-id | deployment-id>       # required for deepinfra

See `llm_selection_verified_jun2026_with_medical.xlsx` (sheet "Medical Specialist
(API)") for the access-mode / pricing rationale.

MedGemma-27B is a standard (non-reasoning) instruct model: it answers
directly without a <think> block, so MAX_OUTPUT_TOKENS is modest.
IMPORTANT -- Gemma chat template has NO system role. SUPPORTS_SYSTEM_ROLE is set
to False below, so SYSTEM_PROMPT is folded into the first user message; sending a
separate system message would raise a template error on most Gemma hosts.
"""

import argparse
import csv
import json
import os
import re
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from colorama import Fore, Style, init as colorama_init
from dotenv import dotenv_values
from openai import APIStatusError, BadRequestError, OpenAI, RateLimitError
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score
from tqdm import tqdm


# --- Prompt (identical to classify_glioma_deepseek.py) -----------------------

SYSTEM_PROMPT = (
    "You are an expert neuroradiologist performing pre-operative molecular "
    "subtyping of adult-type diffuse glioma from structured MRI metadata, "
    "following the WHO CNS5 (2021) classification. For every task you reason "
    "step by step BEFORE stating any prediction or confidence score: first lay "
    "out the clinically grounded explanation, then commit to the prediction, "
    "then assign the confidence. Output MUST be a single valid JSON object that "
    "conforms exactly to the GliomaClassificationOutput schema. No markdown, no "
    "prose outside the JSON, no code fences -- JSON only."
)

USER_PROMPT_TEMPLATE = """# ROLE
You are an expert neuroradiologist.

# TASK
You are given structured pre-operative brain MRI metadata for a single patient
with a suspected **adult-type diffuse glioma**, defined according to the WHO
CNS5 molecular classification. Based on this metadata, predict the following:

  (a) IDH mutation status        -> {{"mutant", "wildtype"}}
  (b) 1p/19q co-deletion status  -> {{"codeleted", "non-codeleted"}}
  (c) CNS WHO grade              -> {{2, 3, 4}}

# CHAIN-OF-THOUGHT CONFIDENCE ELICITATION
For each of the three tasks you MUST work in the following order. Do not skip,
reorder, or shortcut these steps:

  1. **Reasoning first.** Write a concise, clinically grounded explanation that
     weighs the available metadata. Explicitly contrast the evidence that
     supports each candidate answer against the evidence that argues against
     it. Reach the prediction as the *conclusion* of this reasoning.
  2. **Supporting metadata fields.** List the exact field names (verbatim) that
     support the prediction you reasoned toward.
  3. **Contradicting features.** List the fields that argue against it.
  4. **Prediction.** State the prediction that follows from steps 1-3.
  5. **Confidence score** (between 0.0 and 1.0). Derive the score *from* the
     reasoning above -- it must reflect the balance of supporting vs.
     contradicting evidence you already articulated, NOT a number chosen before
     reasoning. Assign high confidence only when key features are well
     supported and few contradict; assign low confidence when evidence is
     thin, conflicting, or absent.

The JSON schema lists `reasoning`, `supporting_features` and
`contradicting_features` before `prediction` and `confidence` for exactly this
reason: generate them in that order.

# INPUT CONSTRAINTS
- Input consists **only of structured MRI-derived metadata** (no images).
- Do not assume access to data beyond what is provided.
- Treat null values as **"not assessed"** (do not infer).

# AVAILABLE MRI SEQUENCES
T1w, T2w, T2-FLAIR, T1-Gd.
**Not available:** DWI/ADC, perfusion (rCBV/DSC), MRS (2HG), SWI/GRE.

# RULES
1. Use ONLY the metadata provided; no hallucinated features or sequences.
2. Every claim in `reasoning` must be traceable to a field named verbatim in
   `supporting_features`.
3. Do NOT invoke diffusion, perfusion, spectroscopy, calcifications, or SWI.
4. If evidence is insufficient: give a best estimate, set confidence in
   0.50-0.60, and state the uncertainty in reasoning.
5. Confidence must be >= 0.50 for binary tasks (otherwise flip the prediction).
6. The confidence score must be the conclusion of the reasoning, not its
   premise -- never write the reasoning to justify a pre-chosen score.
7. Output must be **valid JSON only** (no markdown, no commentary).

# INPUT
Patient subject_id: {subject_id}
MRI metadata:
{metadata_json}
"""


# --- Config ------------------------------------------------------------------

MODEL_SLUG  = "medgemma_27b"            # used in output filenames / metric keys
HERE        = Path(__file__).parent
SCHEMA_PATH = HERE / "glioma_classification_schema.json"
ENV_PATH    = Path(".env")
JSONS_DIR   = Path("data/Dataset_AKU_WHO/JSONs")
COHORT_XLSX = Path("data/dataset_table/cohort_merged.xlsx")

# Two OpenAI-compatible hosts.
#   featherless: subscription serverless; "model_id" is the HF repo id served directly.
#   deepinfra:   this model is NOT in DeepInfra's serverless catalog, so it must be
#                deployed as a Custom LLM first. "model_id" here is only the HF repo to
#                DEPLOY; the value actually sent must be your deployment id
#                (<username>/<deployment-name>), supplied via MODEL_ID in .env.
PROVIDERS = {
    "featherless": {
        "base_url": "https://api.featherless.ai/v1",
        "key_env":  "FEATHERLESS_API_KEY",
        "model_id": "google/medgemma-27b-text-it",
    },
    "deepinfra": {
        "base_url":      "https://api.deepinfra.com/v1/openai",
        "key_env":       "DEEPINFRA_API_KEY",
        "model_id":      "google/medgemma-27b-text-it",  # HF repo to deploy (not callable as-is)
        "needs_deploy":  True,                      # requires a Custom-LLM deployment id
        "deploy_repo":   "google/medgemma-27b-text-it",
    },
}
DEFAULT_PROVIDER = "featherless"

# MedGemma-27B (Google) exposes no Llama-Guard-style per-request safety toggle, so
# nothing extra is sent in the request body.
ENABLE_LLAMA_GUARD = False

# This model is NOT a reasoning model (no <think> block). Reasoning tokens are emitted BEFORE the
# JSON answer, so max_tokens is sized to leave headroom and parse_model_json()
# strips any <think>...</think> block before parsing.
IS_REASONING_MODEL = False
MAX_OUTPUT_TOKENS  = 4096

# Without server-side guided decoding, some of these models (notably MedGemma on a
# DeepInfra custom deployment) fall into a whitespace-repetition loop: they finish the
# real JSON, then emit thousands of blank tokens until they hit max_tokens, so the
# object never closes (finish_reason=length) and won't parse. A small frequency_penalty
# penalises the repeated whitespace token and breaks the loop so the model closes the
# JSON within budget. Set to 0.0 to disable. (parse_model_json() also repairs truncated
# objects as a backstop.)
FREQUENCY_PENALTY = 0.3

# Some open chat templates (notably Gemma, which MedGemma is built on) do NOT
# accept a separate system role. When False, SYSTEM_PROMPT is folded into the
# first user message instead of being sent as a system message.
SUPPORTS_SYSTEM_ROLE = False

# Output constraint mode. Managed OpenAI-compatible hosts vary in structured-output
# support: DeepInfra enforces strict json_schema; Featherless may only honour
# json_object (or neither for some models). We default to STRICT json_schema like
# the gpt55/groq/deepseek scripts; if the host rejects it, the request loop
# auto-disables structured output for the rest of the run (the CoT-CE prompt and
# the <think>-stripping / schema-tolerant parser still apply).
# Modes: "json_schema" | "json_object" | "none".
RESPONSE_FORMAT_MODE = "json_schema"


def build_response_format():
    """OpenAI-style response_format for the configured mode (None if disabled)."""
    if RESPONSE_FORMAT_MODE == "json_schema":
        raw = json.load(open(SCHEMA_PATH))
        schema = {k: v for k, v in raw.items() if k != "$schema"}
        return {"type": "json_schema",
                "json_schema": {"name": "GliomaClassificationOutput",
                                "strict": True, "schema": schema}}
    if RESPONSE_FORMAT_MODE == "json_object":
        return {"type": "json_object"}
    return None


def build_messages(user_prompt: str, reinforce: str | None = None) -> list:
    """Assemble the chat messages, honouring SUPPORTS_SYSTEM_ROLE. When the model
    has no system role, SYSTEM_PROMPT is prepended to the user turn.

    The reinforcement nudge is FOLDED INTO the single user turn rather than added as a
    second user message. Strict-alternation chat templates (notably Gemma / MedGemma)
    reject two consecutive user messages with 'Conversation roles must alternate
    user/assistant/...'; merging keeps exactly one user turn and is harmless for the
    others."""
    user_content = user_prompt if not reinforce else (user_prompt + "\n\n" + reinforce)
    if SUPPORTS_SYSTEM_ROLE:
        return [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": user_content},
        ]
    return [
        {"role": "user", "content": SYSTEM_PROMPT + "\n\n" + user_content},
    ]


def _dump_raw(output_dir, subject, attempt, raw, finish):
    """When DEBUG_DUMP is set, save the model's raw text for a failed parse so the
    exact cause (unescaped quotes, trailing prose, truncation) can be inspected."""
    if not os.environ.get("DEBUG_DUMP"):
        return
    dbg = output_dir / "_debug"
    dbg.mkdir(parents=True, exist_ok=True)
    with open(dbg / f"{subject}_attempt{attempt}.txt", "w") as f:
        f.write(f"# finish_reason={finish}  chars={len(raw or '')}\n")
        f.write(raw or "")


# Open medical models sometimes collapse the three tasks into a single block (a
# scalar `prediction` answering only IDH). It isn't deterministic, so re-asking --
# with a reinforcement nudge and a little temperature -- usually fixes it. How many
# times to re-ask before giving up on a subject and skipping it (a later run retries).
STRUCTURE_RETRIES = 5

STRUCTURE_REINFORCE = (
    "Your previous response did not follow the required structure. Return a SINGLE "
    "JSON object with EXACTLY three top-level task objects named \"idh\", "
    "\"one_p_nineteen_q\", and \"who_grade\". Each task object must contain "
    "\"reasoning\", \"supporting_features\", \"contradicting_features\", "
    "\"prediction\", and \"confidence\". Do NOT merge the three tasks into one "
    "block, and do NOT put a single scalar string in \"prediction\". "
    "idh.prediction must be \"mutant\" or \"wildtype\"; "
    "one_p_nineteen_q.prediction must be \"codeleted\" or \"non-codeleted\"; "
    "who_grade.prediction must be the integer 2, 3, or 4. Output JSON only."
)


# --- Provider (OpenAI-compatible) client -------------------------------------

def load_client() -> tuple[OpenAI, str, str]:
    """Build an OpenAI client for the configured provider. Returns
    (client, model_id, provider_name). Reads .env, falling back to process env."""
    config = dotenv_values(ENV_PATH)

    def _get(key, default=None):
        return config.get(key) or os.environ.get(key) or default

    provider = (_get("PROVIDER", DEFAULT_PROVIDER) or DEFAULT_PROVIDER).lower()
    if provider not in PROVIDERS:
        raise RuntimeError(
            f"Unknown PROVIDER {provider!r}. Choose one of: {', '.join(PROVIDERS)}."
        )
    pconf = PROVIDERS[provider]
    api_key = _get(pconf["key_env"])
    if not api_key:
        raise RuntimeError(
            f"{pconf['key_env']} not set. Add it to .env (PROVIDER={provider}). "
            f"See the 'Medical Specialist (API)' sheet for how to obtain a key."
        )
    model_id_override = _get("MODEL_ID")
    model_id = model_id_override or pconf["model_id"]

    # DeepInfra has no serverless entry for these medical models -- they run as a
    # Custom LLM deployment, whose id is <username>/<deployment-name>. Refuse to start
    # with just the HF repo id (it would 400/404), and point the user at the deploy flow.
    if pconf.get("needs_deploy") and not model_id_override:
        raise RuntimeError(
            f"PROVIDER=deepinfra requires MODEL_ID set to your DeepInfra Custom-LLM "
            f"deployment id (format '<your-username>/<deployment-name>').\n"
            f"  These models are NOT in DeepInfra's serverless catalog. First deploy the "
            f"HF repo '{pconf['deploy_repo']}' as a Custom LLM at "
            f"https://deepinfra.com/dash/deployments (dedicated GPU, billed by runtime), "
            f"then set MODEL_ID to the deployment id it gives you."
        )

    client = OpenAI(base_url=pconf["base_url"], api_key=api_key)
    return client, model_id, provider


# --- ID helpers (identical to classify_glioma_deepseek.py) -------------------

def load_id_map(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    id_map: dict = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        center_id  = str(row[0]).strip() if row[0] else None
        subject_id = str(row[1]).strip() if row[1] else None
        brats_id   = str(row[4]).strip() if (row[4] and row[4] != "x") else None
        entry = {"center_id": center_id, "subject_id": subject_id}
        if center_id:
            id_map[center_id] = entry
        if brats_id:
            id_map[brats_id] = entry
    return id_map


# --- Ground-truth helpers (identical to classify_glioma_deepseek.py) ---------

def load_ground_truth(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def _idh(v):
        if v == "mutated":   return "mutant"
        if v == "wild type": return "wildtype"
        return None

    def _codeletion(v):
        if v in ("co-deleted", "rel. co-deleted"): return "codeleted"
        if v == "intact":                          return "non-codeleted"
        return None

    gt: dict = {}
    for row in rows[1:]:
        center_id = row[0]
        brats_id  = row[4] if (row[4] and row[4] != "x") else None
        entry = {
            "idh":        _idh(row[7]),
            "codeletion": _codeletion(row[8]),
            "grade":      row[9] if isinstance(row[9], int) else None,
            "dataset":    row[2],
        }
        if center_id:
            gt[center_id] = entry
        if brats_id:
            gt[brats_id] = entry
    return gt


# --- Metrics (identical to classify_glioma_deepseek.py) ----------------------

def _binary_metrics(y_true, y_pred, pos_label) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None, "weighted_f1": None}
    all_labels = sorted(set(y_true) | set(y_pred))
    neg_candidates = [l for l in all_labels if l != pos_label]
    neg_label = neg_candidates[0] if neg_candidates else (
        "wildtype" if pos_label == "mutant" else "non-codeleted"
    )
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, pos_label=pos_label, average="binary",
                   zero_division=0)
    wf1 = f1_score(y_true, y_pred, average="weighted", zero_division=0)
    cm  = confusion_matrix(y_true, y_pred, labels=[pos_label, neg_label])
    TP, FN = cm[0, 0], cm[0, 1]
    FP, TN = cm[1, 0], cm[1, 1]
    sens = TP / (TP + FN) if (TP + FN) > 0 else None
    spec = TN / (TN + FP) if (TN + FP) > 0 else None
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sens, 4) if sens is not None else None,
        "specificity": round(spec, 4) if spec is not None else None,
        "f1":          round(f1, 4),
        "weighted_f1": round(wf1, 4),
    }


def _grade_metrics(y_true, y_pred) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None, "weighted_f1": None}
    classes = [2, 3, 4]
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, labels=classes, average="macro",
                   zero_division=0)
    wf1 = f1_score(y_true, y_pred, labels=classes, average="weighted",
                   zero_division=0)
    cm = confusion_matrix(y_true, y_pred, labels=classes)
    sens_list, spec_list = [], []
    for i in range(len(classes)):
        TP = cm[i, i]
        FN = cm[i, :].sum() - TP
        FP = cm[:, i].sum() - TP
        TN = cm.sum() - TP - FN - FP
        sens_list.append(TP / (TP + FN) if (TP + FN) > 0 else 0.0)
        spec_list.append(TN / (TN + FP) if (TN + FP) > 0 else 0.0)
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sum(sens_list) / len(sens_list), 4),
        "specificity": round(sum(spec_list) / len(spec_list), 4),
        "f1":          round(f1, 4),
        "weighted_f1": round(wf1, 4),
    }


def compute_metrics(predictions: dict, gt: dict) -> dict:
    data = defaultdict(lambda: {
        "idh_t": [], "idh_p": [],
        "cod_t": [], "cod_p": [],
        "grd_t": [], "grd_p": [],
    })
    for sid, pred in predictions.items():
        row = gt.get(sid)
        if row is None:
            continue
        ds = row["dataset"] or "Unknown"
        for bucket in ("whole_cohort", ds):
            d = data[bucket]
            if row["idh"] is not None and pred.get("idh") is not None:
                d["idh_t"].append(row["idh"]); d["idh_p"].append(pred["idh"])
            if row["codeletion"] is not None and pred.get("codeletion") is not None:
                d["cod_t"].append(row["codeletion"]); d["cod_p"].append(pred["codeletion"])
            if row["grade"] is not None and pred.get("grade") is not None:
                d["grd_t"].append(row["grade"]); d["grd_p"].append(pred["grade"])

    results = {}
    for bucket, d in data.items():
        results[bucket] = {
            "idh":        _binary_metrics(d["idh_t"], d["idh_p"], "mutant"),
            "codeletion": _binary_metrics(d["cod_t"], d["cod_p"], "codeleted"),
            "grade":      _grade_metrics(d["grd_t"], d["grd_p"]),
        }
    return results


def print_metrics(metrics: dict) -> None:
    col_w = 22
    header = (f"{'Subset':<{col_w}} {'Task':<12} {'N':>5}  "
              f"{'Acc':>6}  {'Sens':>6}  {'Spec':>6}  {'F1':>6}  {'WF1':>6}")
    sep = "=" * len(header)
    print(f"\n{sep}\nCLASSIFICATION METRICS\n{sep}")
    print(header)
    print("-" * len(header))
    task_labels = {"idh": "IDH", "codeletion": "1p/19q", "grade": "Grade"}
    buckets = ["whole_cohort"] + sorted(k for k in metrics if k != "whole_cohort")
    for bucket in buckets:
        for task_key, label in task_labels.items():
            m = metrics[bucket][task_key]
            def _fmt(v):
                return f"{v:.4f}" if v is not None else "  N/A"
            print(f"{bucket:<{col_w}} {label:<12} {m['n']:>5}  "
                  f"{_fmt(m['accuracy']):>6}  {_fmt(m['sensitivity']):>6}  "
                  f"{_fmt(m['specificity']):>6}  {_fmt(m['f1']):>6}  "
                  f"{_fmt(m['weighted_f1']):>6}")
        print()


def compute_and_save_metrics(output_dir) -> None:
    """Score already-saved classification outputs (no API calls)."""
    C, G, D = Fore.CYAN, Fore.GREEN, Style.DIM
    gt = load_ground_truth(COHORT_XLSX)
    print(f"\n{D}" + "-" * 72 + Style.RESET_ALL)
    print(f"{C}Collecting predictions for metric computation...{Style.RESET_ALL}")
    predictions: dict = {}
    for out_file in sorted(output_dir.glob(f"*_classification_{MODEL_SLUG}.json")):
        result = json.load(open(out_file))
        sid = result.get("subject_id",
                          out_file.stem.replace(f"_classification_{MODEL_SLUG}", ""))
        predictions[sid] = extract_predictions(result)
    metrics = compute_metrics(predictions, gt)
    print_metrics(metrics)
    metrics_path = output_dir / "metrics.json"
    with open(metrics_path, "w") as f:
        json.dump(metrics, f, indent=2)
    print(f"{G}Metrics saved{Style.RESET_ALL} -> {metrics_path}")


# --- JSON extraction ---------------------------------------------------------

_FENCE_RE = re.compile(r"^```(?:json)?\s*|\s*```$", re.MULTILINE)
# Reasoning models (e.g. Baichuan-M2) may prepend a chain-of-thought block wrapped
# in <think>...</think> before the JSON answer. Strip it before parsing. Harmless
# for non-reasoning models that never emit it.
_THINK_RE = re.compile(r"<think>.*?</think>", re.DOTALL | re.IGNORECASE)


def _repair_truncated_json(text: str):
    """Best-effort recovery of a TRUNCATED JSON object.

    Without server-side guided decoding, these models sometimes emit a correct JSON
    prefix (all three task blocks) and then degenerate into a long whitespace run that
    eats the token budget, so the object is cut off before its closing braces. The real
    data is already there -- we just close the structure: trim trailing whitespace and
    any dangling partial token, drop a trailing comma/colon, then append the closing
    ] / } implied by the still-open brackets. Returns a dict or None.
    """
    start = text.find("{")
    if start == -1:
        return None
    s = text[start:].rstrip()

    # Walk once to find the last index that ends a *complete* element, tracking strings.
    in_str = esc = False
    last_safe = None
    for i, ch in enumerate(s):
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
                last_safe = i + 1
        else:
            if ch == '"':
                in_str = True
            elif ch in "}]":
                last_safe = i + 1
            elif ch in "0123456789":
                last_safe = i + 1
            elif ch in "el":          # tails of true / false / null
                last_safe = i + 1
    if last_safe is None:
        return None

    candidate = s[:last_safe].rstrip().rstrip(",").rstrip()

    # Recompute the open-bracket stack on the trimmed candidate, ignoring strings.
    stack = []
    in_str = esc = False
    for ch in candidate:
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch in "{[":
                stack.append(ch)
            elif ch in "}]":
                if stack:
                    stack.pop()
    closing = "".join("}" if c == "{" else "]" for c in reversed(stack))
    try:
        return json.loads(candidate + closing)
    except json.JSONDecodeError:
        return None


def parse_model_json(content: str) -> dict:
    """
    Parse the model's JSON. These open-weight hosts do not guarantee strict
    json_schema enforcement, and reasoning variants may wrap a chain of thought
    in <think>...</think>, so be defensive: drop the think block and any code
    fences, extract the outermost JSON object, and as a last resort repair a
    truncated object (common when the model degenerates into a whitespace loop and
    hits the token limit before closing the JSON).
    """
    text = _THINK_RE.sub("", content).strip()
    # If an unterminated <think> remains (truncated), keep everything after it.
    if "<think>" in text.lower():
        text = re.split(r"</?think>", text, flags=re.IGNORECASE)[-1].strip()
    text = _FENCE_RE.sub("", text).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start, end = text.find("{"), text.rfind("}")
        if start != -1 and end != -1 and end > start:
            try:
                return json.loads(text[start:end + 1])
            except json.JSONDecodeError:
                pass
        repaired = _repair_truncated_json(text)
        if repaired is not None:
            return repaired
        raise


# --- Prediction extraction (content-based, schema-tolerant) ------------------
# Open medical models don't reliably enforce the schema's field names, so across
# runs they emit the three task blocks under many shapes: task-named top-level keys
# (e.g. "idh_status", "IDH mutation status", "cns_who_grade"), or a nested
# "predictions" list/dict where each block carries a "type" field. We identify each
# task by the *content* of its key/type rather than a fixed alias list, so saved
# outputs score correctly without re-running inference. (The only unrecoverable
# shape is a single flat block that collapses all three tasks into one prediction.)

def _classify_key(name) -> str | None:
    s = str(name).lower()
    if "idh" in s:
        return "idh"
    if ("1p" in s or "19q" in s or "nineteen" in s or "one_p" in s
            or "codelet" in s or "co-delet" in s):
        return "codeletion"
    if "grade" in s:
        return "grade"
    return None


def _iter_task_blocks(result: dict):
    """Yield (task, block) for every identifiable task block in an output JSON,
    handling a nested `predictions` list/dict or task-named top-level keys."""
    preds = result.get("predictions")
    if isinstance(preds, list):
        for b in preds:
            if isinstance(b, dict):
                task = _classify_key(b.get("type", ""))
                if task:
                    yield task, b
        return
    if isinstance(preds, dict):
        for k, v in preds.items():
            if isinstance(v, dict):
                task = _classify_key(k)
                if task:
                    yield task, v
        return
    for k, v in result.items():
        if isinstance(v, dict):
            task = _classify_key(k)
            if task:
                yield task, v


def _norm_idh(v):
    if not isinstance(v, str):
        return v
    s = v.strip().lower()
    if "mut" in s:  return "mutant"
    if "wild" in s: return "wildtype"
    return v


def _norm_cod(v):
    if not isinstance(v, str):
        return v
    s = v.strip().lower()
    if "non" in s or "intact" in s: return "non-codeleted"
    if "codelet" in s or "co-delet" in s: return "codeleted"
    return v


def _coerce_grade(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def extract_predictions(result: dict) -> dict:
    """Return {"idh", "codeletion", "grade"} from one output JSON, identifying
    task blocks by content and normalising label/grade formats."""
    out = {"idh": None, "codeletion": None, "grade": None}
    for task, block in _iter_task_blocks(result):
        if out[task] is None:
            out[task] = block.get("prediction")
    return {
        "idh":        _norm_idh(out["idh"]),
        "codeletion": _norm_cod(out["codeletion"]),
        "grade":      _coerce_grade(out["grade"]),
    }


def result_is_complete(result: dict) -> bool:
    """True iff all three tasks are present with valid label/grade values."""
    p = extract_predictions(result)
    return (p["idh"] in ("mutant", "wildtype")
            and p["codeletion"] in ("codeleted", "non-codeleted")
            and p["grade"] in (2, 3, 4))


# --- CLI ---------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Classify gliomas via MedGemma-27B (Google) on an OpenAI-compatible "
                    "host (btreport or btreport_pp JSONs)."
    )
    parser.add_argument(
        "--json-type", choices=["btreport", "btreport_pp"], required=True,
        help="Which metadata JSON variant to use (btreport or btreport_pp).",
    )
    parser.add_argument(
        "--metrics-only", action="store_true",
        help="Skip inference (no API calls) and only recompute metrics from "
             "already-saved outputs.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    colorama_init(autoreset=True)
    C, G, Y, R, M, W, D = (Fore.CYAN, Fore.GREEN, Fore.YELLOW, Fore.RED,
                           Fore.MAGENTA, Style.BRIGHT, Style.DIM)

    args      = parse_args()
    json_type = args.json_type

    OUTPUT_DIR = HERE / f"medgemma_outputs_{json_type}"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if args.metrics_only:
        compute_and_save_metrics(OUTPUT_DIR)
        raise SystemExit(0)

    TIMINGS_CSV = OUTPUT_DIR / "timings.csv"
    _csv_is_new = not TIMINGS_CSV.exists()
    _timings_fh = TIMINGS_CSV.open("a", newline="")
    _timings_writer = csv.DictWriter(
        _timings_fh, fieldnames=["folder_name", "elapsed_s", "timestamp"])
    if _csv_is_new:
        _timings_writer.writeheader()
        _timings_fh.flush()

    schema = json.load(open(SCHEMA_PATH))

    client, MODEL_ID, provider = load_client()
    print(f"{C}OpenAI-compatible host  |  provider={W}{provider}{C}\n"
          f"Model: {W}{MODEL_ID}{C}  |  Output: {W}{OUTPUT_DIR.name}")

    id_map = load_id_map(COHORT_XLSX)
    subjects = sorted(d.name for d in JSONS_DIR.iterdir() if d.is_dir())

    pbar = tqdm(
        subjects, desc=f"{C}Classifying{Style.RESET_ALL}", unit="subj", ncols=90,
        bar_format=("{desc}: {percentage:3.0f}%|{bar:35}| "
                    "{n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]"),
        colour="cyan",
    )
    DIVIDER = f"{D}" + "-" * 72 + Style.RESET_ALL

    # Mutable so the request loop can disable structured output for the rest of the
    # run if the host rejects it (the prompt + parser still apply).
    rf_state = {"value": build_response_format()}

    for i, folder_name in enumerate(pbar, 1):
        pbar.set_postfix_str(f"{folder_name[:28]}", refresh=True)
        tqdm.write("")
        tqdm.write(DIVIDER)
        tqdm.write(f"{W}[{i}/{len(subjects)}]{Style.RESET_ALL}  "
                   f"{C}{folder_name}{Style.RESET_ALL}")

        ids = id_map.get(folder_name)
        if ids:
            center_id      = ids["center_id"]
            api_subject_id = ids["subject_id"]
        else:
            tqdm.write(f"  {Y}WARN{Style.RESET_ALL}  no xlsx entry for "
                       f"{folder_name!r} -- using folder name as fallback")
            center_id = api_subject_id = folder_name

        metadata_path = JSONS_DIR / folder_name / f"{folder_name}_metadata_{json_type}.json"
        if not metadata_path.exists():
            tqdm.write(f"  {Y}SKIP{Style.RESET_ALL}  metadata not found: {metadata_path.name}")
            continue

        out_path = OUTPUT_DIR / f"{folder_name}_classification_{MODEL_SLUG}.json"
        if out_path.exists():
            tqdm.write(f"  {Y}SKIP{Style.RESET_ALL}  output already exists: {out_path.name}")
            continue

        tqdm.write(f"  {D}center_id={Style.RESET_ALL}{center_id}   "
                   f"{D}api_id={Style.RESET_ALL}{api_subject_id}")

        metadata    = json.load(open(metadata_path))
        user_prompt = USER_PROMPT_TEMPLATE.format(
            subject_id=api_subject_id,
            metadata_json=json.dumps(metadata, indent=2, default=str),
        )

        # Generic transient errors (429 / 5xx other than capacity) get a short
        # budget. Serverless "capacity_exhausted" 503s are different -- the model
        # pool is oversubscribed and clears over MINUTES, not seconds -- so they get
        # their own much more patient budget with long, capped backoff. Set
        # CAPACITY_MAX_RETRIES=0 to fail fast on capacity instead of waiting.
        MAX_RETRIES          = 4     # network / 429 / generic 5xx per request
        CAPACITY_MAX_RETRIES = 8     # extra patience for 503 capacity_exhausted
        CAPACITY_WAIT_BASE   = 20    # seconds; grows linearly per capacity retry
        CAPACITY_WAIT_MAX    = 120   # cap on a single capacity backoff
        t0 = time.perf_counter()

        def _request(messages, temperature):
            """One chat call with rate-limit / server / capacity retries. Returns
            the response object, or None on a terminal failure."""
            gen_attempt = 0   # counts 429 + generic 5xx
            cap_attempt = 0   # counts 503 capacity_exhausted (separate budget)
            while True:
                try:
                    rf = rf_state["value"]
                    # Pass response_format via extra_body so the full {type, json_schema}
                    # object reaches the server verbatim. Older openai SDKs serialise the
                    # typed `response_format=` kwarg against their own model and silently
                    # DROP the `json_schema` field (sending only {"type":"json_schema"}),
                    # which vLLM hosts like Featherless then reject. extra_body bypasses
                    # that typed validation entirely.
                    return client.chat.completions.create(
                        model=MODEL_ID,
                        messages=messages,
                        temperature=temperature,
                        seed=42,
                        # Reasoning variants consume output tokens before the JSON,
                        # so allow headroom to avoid truncation.
                        max_tokens=MAX_OUTPUT_TOKENS,
                        # Break runaway whitespace/repetition loops (see FREQUENCY_PENALTY).
                        **({"frequency_penalty": FREQUENCY_PENALTY}
                           if FREQUENCY_PENALTY else {}),
                        # Strict structured output if the host accepts it (auto-
                        # disabled below on rejection).
                        **({"extra_body": {"response_format": rf}} if rf else {}),
                    )
                except RateLimitError as e:
                    gen_attempt += 1
                    if gen_attempt >= MAX_RETRIES:
                        tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  429 rate-limited "
                                   f"after {MAX_RETRIES} tries: {e}")
                        return None
                    wait = min(2 ** gen_attempt, 30)
                    tqdm.write(f"  {M}LIMIT{Style.RESET_ALL}  429 (attempt "
                               f"{gen_attempt}/{MAX_RETRIES}) -- backing off {wait}s  "
                               f"{D}({e}){Style.RESET_ALL}")
                    time.sleep(wait)
                # NOTE: BadRequestError is a subclass of APIStatusError, so it MUST be
                # caught first -- otherwise the structured-output auto-disable below
                # never runs and a 400 just aborts the subject.
                except BadRequestError as e:
                    msg = str(e).lower()
                    # Host may reject structured output -- disable it for the rest of
                    # the run and retry (prompt + schema-tolerant parser remain).
                    if rf_state["value"] is not None and (
                            "response_format" in msg or "json_schema" in msg
                            or "schema" in msg or "structured" in msg
                            or "grammar" in msg or "guided" in msg):
                        tqdm.write(f"  {Y}NOTE{Style.RESET_ALL}  host rejected "
                                   f"structured output -- disabling it for this run "
                                   f"and retrying.")
                        rf_state["value"] = None
                        continue
                    # Serverless models spin DOWN when idle: Featherless then returns a
                    # 400 model_not_deployed / "not available for inference" until it
                    # re-activates. This is transient -- wait (capacity budget) and retry,
                    # do NOT skip the subject.
                    if ("model_not_deployed" in msg or "not deployed" in msg
                            or "not available for inference" in msg):
                        cap_attempt += 1
                        if cap_attempt > CAPACITY_MAX_RETRIES:
                            tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  model still not "
                                       f"deployed after {CAPACITY_MAX_RETRIES} waits: {e}")
                            return None
                        wait = min(CAPACITY_WAIT_BASE * cap_attempt, CAPACITY_WAIT_MAX)
                        tqdm.write(f"  {M}DEPLOYING{Style.RESET_ALL}  400 model spun down "
                                   f"(wait {cap_attempt}/{CAPACITY_MAX_RETRIES}) -- "
                                   f"sleeping {wait}s for re-activation")
                        time.sleep(wait)
                        continue
                    tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  bad request: {e}")
                    return None
                except APIStatusError as e:
                    msg = str(e).lower()
                    # Serverless capacity exhaustion: wait it out patiently.
                    if e.status_code == 503 and (
                            "capacity" in msg or "at capacity" in msg
                            or "capacity_exhausted" in msg):
                        cap_attempt += 1
                        if cap_attempt > CAPACITY_MAX_RETRIES:
                            tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  still at capacity "
                                       f"after {CAPACITY_MAX_RETRIES} waits: {e}")
                            return None
                        wait = min(CAPACITY_WAIT_BASE * cap_attempt, CAPACITY_WAIT_MAX)
                        tqdm.write(f"  {M}CAPACITY{Style.RESET_ALL}  503 model at capacity "
                                   f"(wait {cap_attempt}/{CAPACITY_MAX_RETRIES}) -- "
                                   f"sleeping {wait}s")
                        time.sleep(wait)
                        continue
                    # Other 5xx: short generic retry budget.
                    if e.status_code >= 500:
                        gen_attempt += 1
                        if gen_attempt >= MAX_RETRIES:
                            tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  server "
                                       f"{e.status_code} after {MAX_RETRIES} tries: {e}")
                            return None
                        tqdm.write(f"  {Y}RETRY{Style.RESET_ALL}  server {e.status_code} "
                                   f"(attempt {gen_attempt}/{MAX_RETRIES})")
                        time.sleep(2 ** gen_attempt)
                        continue
                    tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  {e.status_code}: {e}")
                    return None

        # Structure-retry loop: re-ask when the model collapses the three tasks
        # into one block, until we get a complete object or exhaust the budget.
        result = None
        for s_attempt in range(1, STRUCTURE_RETRIES + 1):
            reinforce = STRUCTURE_REINFORCE if s_attempt > 1 else None
            messages = build_messages(user_prompt, reinforce)
            # A little temperature on retries breaks the degenerate single-block pattern.
            temperature = 0.0 if s_attempt == 1 else 0.4

            response = _request(messages, temperature)
            if response is None:
                break
            choice = response.choices[0]
            raw = choice.message.content or ""
            finish = getattr(choice, "finish_reason", None)
            # finish_reason is the key diagnostic: "length" => truncated (raise
            # MAX_OUTPUT_TOKENS); "stop" with invalid JSON => the host isn't enforcing
            # the schema and the model free-generated (run with DEBUG_DUMP=1 to inspect).
            try:
                cand = parse_model_json(raw)
            except json.JSONDecodeError as e:
                tqdm.write(f"  {Y}RETRY{Style.RESET_ALL}  unparseable JSON ({e}) "
                           f"[finish_reason={finish}, {len(raw)} chars] "
                           f"(structure attempt {s_attempt}/{STRUCTURE_RETRIES})")
                _dump_raw(OUTPUT_DIR, folder_name, s_attempt, raw, finish)
                continue
            result = cand
            if result_is_complete(cand):
                break
            tqdm.write(f"  {Y}RETRY{Style.RESET_ALL}  collapsed/incomplete output "
                       f"[finish_reason={finish}] "
                       f"(structure attempt {s_attempt}/{STRUCTURE_RETRIES})")
            _dump_raw(OUTPUT_DIR, folder_name, s_attempt, raw, finish)

        if result is None:
            tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  no usable response -- skipping {folder_name}")
            continue

        if not result_is_complete(result):
            # Don't save a collapsed result -- leave no output file so a later run
            # retries this subject (keeps the cohort's metrics honest).
            tqdm.write(f"  {R}UNRESOLVED{Style.RESET_ALL}  still collapsed after "
                       f"{STRUCTURE_RETRIES} tries -- skipping {folder_name} (re-run to retry)")
            continue

        elapsed = round(time.perf_counter() - t0, 2)
        result["subject_id"] = folder_name
        with open(out_path, "w") as f:
            json.dump(result, f, indent=2)

        _timings_writer.writerow({
            "folder_name": folder_name,
            "elapsed_s":   elapsed,
            "timestamp":   datetime.now().isoformat(timespec="seconds"),
        })
        _timings_fh.flush()
        tqdm.write(f"  {G}SAVED{Style.RESET_ALL}  {out_path.name}  {D}({elapsed}s){Style.RESET_ALL}")

    pbar.close()
    _timings_fh.close()
    print(f"{C}Timings saved{Style.RESET_ALL} -> {TIMINGS_CSV}")

    compute_and_save_metrics(OUTPUT_DIR)
