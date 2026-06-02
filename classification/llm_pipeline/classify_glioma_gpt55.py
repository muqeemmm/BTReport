"""
Glioma classification via the OpenAI Batch API (gpt-5.5 by default).

This is the OpenAI counterpart of `classify_glioma_gemini.py`. It performs the
SAME task — joint prediction of IDH status, 1p/19q co-deletion and CNS WHO grade
from structured pre-operative MRI metadata, under WHO CNS5 (2021) — using the
SAME Chain-of-Thought Confidence Elicitation (CoT CE) schema, but routes
inference through OpenAI's **asynchronous Batch API** to get the documented 50%
cost discount versus synchronous calls.

WHY THE STRUCTURE DIFFERS FROM THE GROQ SCRIPT
----------------------------------------------
Groq inference is synchronous: one request -> one response, looped per subject.
The OpenAI Batch API is asynchronous and has three phases that cannot be a
single loop:

    1. SUBMIT  - bundle every subject into batch job(s) and hand them to OpenAI.
    2. POLL    - wait (seconds to <=24h) until the job(s) reach a terminal state.
    3. FETCH   - download the results, write one JSON per subject, score them.

So this file is a small CLI with sub-commands (`submit`, `fetch`, `run`,
`metrics`, `status`) plus a persisted `batch_state.json` so a submitted job can
be picked up later from a different process.

OPENAI BATCH FORMAT (JSONL upload)
-----------------------------------
Unlike Gemini's inline requests, the OpenAI Batch API requires uploading a JSONL
file where each line is a request object with `custom_id`, `method`, `url`, and
`body`. The `custom_id` is used to match responses to requests in the output.

Steps:
    1. Build a JSONL string (one request per subject, using folder_name as
       custom_id).
    2. Upload via `client.files.create(file=..., purpose="batch")`.
    3. Create a batch with `client.batches.create(input_file_id=...,
       endpoint="/v1/chat/completions", completion_window="24h")`.
    4. Poll `client.batches.retrieve(batch_id)` until `status` is terminal.
    5. Retrieve output via `client.files.content(batch.output_file_id)`.
    6. Parse each JSONL output line: the JSON text is at
       `response.body.choices[0].message.content`.

Subjects are split into chunks of CHUNK_SIZE (40) so each JSONL file stays
manageable and a single failure doesn't block the whole cohort.

Schema, prompt and metrics are intentionally identical to the Gemini script so
the two models are directly comparable.
"""

import argparse
import csv
import io
import json
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import List, Literal

import openai
import openpyxl
from colorama import Fore, Style, init as colorama_init
from dotenv import dotenv_values
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score
from tqdm import tqdm


# --- Prompt (identical to classify_glioma_gemini.py) --------------------------

SYSTEM_PROMPT = (
    "You are an expert neuroradiologist performing pre-operative molecular "
    "subtyping of adult-type diffuse glioma from structured MRI metadata, "
    "following the WHO CNS5 (2021) classification. For every task you reason "
    "step by step BEFORE stating any prediction or confidence score: first lay "
    "out the clinically grounded explanation, then commit to the prediction, "
    "then assign the confidence. "
    "CRITICAL: Each classification task (IDH status, 1p/19q co-deletion, WHO "
    "grade) must be performed INDEPENDENTLY. Reason about each task separately "
    "using only the raw MRI metadata — do NOT let the prediction or confidence "
    "from one task influence the reasoning, prediction, or confidence of another. "
    "Output MUST be a single valid JSON object that "
    "conforms exactly to the GliomaClassificationOutput schema. No markdown, no "
    "prose outside the JSON, no code fences -- JSON only."
)

USER_PROMPT_TEMPLATE = """# ROLE
You are an expert neuroradiologist.

# TASK
You are given structured pre-operative brain MRI metadata for a single patient
with a suspected **adult-type diffuse glioma**, defined according to the WHO
CNS5 molecular classification. Based on this metadata, predict the following:

  (a) IDH mutation status        -> {{"mutant", "wildtype"}}
  (b) 1p/19q co-deletion status  -> {{"codeleted", "non-codeleted"}}
  (c) CNS WHO grade              -> {{2, 3, 4}}

# CHAIN-OF-THOUGHT CONFIDENCE ELICITATION
For each of the three tasks you MUST work in the following order. Do not skip,
reorder, or shortcut these steps:

  1. **Reasoning first.** Write a concise, clinically grounded explanation that
     weighs the available metadata. Explicitly contrast the evidence that
     supports each candidate answer against the evidence that argues against
     it. Reach the prediction as the *conclusion* of this reasoning.
  2. **Supporting metadata fields.** List the exact field names (verbatim) that
     support the prediction you reasoned toward.
  3. **Contradicting features.** List the fields that argue against it.
  4. **Prediction.** State the prediction that follows from steps 1-3.
  5. **Confidence score** (between 0.0 and 1.0). Derive the score *from* the
     reasoning above -- it must reflect the balance of supporting vs.
     contradicting evidence you already articulated, NOT a number chosen before
     reasoning. Assign high confidence only when key features are well
     supported and few contradict; assign low confidence when evidence is
     thin, conflicting, or absent.

The JSON schema lists `reasoning`, `supporting_features` and
`contradicting_features` before `prediction` and `confidence` for exactly this
reason: generate them in that order.

# INPUT CONSTRAINTS
- Input consists **only of structured MRI-derived metadata** (no images).
- Do not assume access to data beyond what is provided.
- Treat null values as **"not assessed"** (do not infer).

# AVAILABLE MRI SEQUENCES
T1w, T2w, T2-FLAIR, T1-Gd.
**Not available:** DWI/ADC, perfusion (rCBV/DSC), MRS (2HG), SWI/GRE.

# RULES
1. Use ONLY the metadata provided; no hallucinated features or sequences.
2. Every claim in `reasoning` must be traceable to a field named verbatim in
   `supporting_features`.
3. Do NOT invoke diffusion, perfusion, spectroscopy, calcifications, or SWI.
4. If evidence is insufficient: give a best estimate, set confidence in
   0.50-0.60, and state the uncertainty in reasoning.
5. Confidence must be >= 0.50 for binary tasks (otherwise flip the prediction).
6. The confidence score must be the conclusion of the reasoning, not its
   premise -- never write the reasoning to justify a pre-chosen score.
7. Output must be **valid JSON only** (no markdown, no commentary).

# INPUT
Patient subject_id: {subject_id}
MRI metadata:
{metadata_json}
"""


# --- Config ------------------------------------------------------------------

HERE        = Path(__file__).parent
ENV_PATH    = Path(".env")
JSONS_DIR   = Path("data/Dataset_AKU_WHO/JSONs")
COHORT_XLSX = Path("data/dataset_table/cohort_merged.xlsx")

DEFAULT_MODEL    = "gpt-5.5"
CHUNK_SIZE       = 40        # subjects per batch JSONL file (keep payload manageable)
POLL_INTERVAL_S  = 30        # seconds between status polls
SEED             = 42        # best-effort reproducibility (temperature is locked
                             # at 1 for gpt-5.x, so `seed` is the only determinism
                             # lever; valid only while system_fingerprint is stable)

TERMINAL_STATES = {"completed", "failed", "expired", "cancelled"}


def model_slug(model: str) -> str:
    """Returns model name as-is for filenames (gpt-5.5 -> gpt-5.5)."""
    return model


# --- Structured-output schema (loaded from glioma_classification_schema.json) -

_SCHEMA_PATH = HERE / "glioma_classification_schema.json"
with open(_SCHEMA_PATH) as _f:
    _GLIOMA_SCHEMA_RAW = json.load(_f)

# Strip JSON-schema meta-keys that OpenAI's API does not accept at the root.
_GLIOMA_SCHEMA = {k: v for k, v in _GLIOMA_SCHEMA_RAW.items()
                  if k not in ("$schema",)}


def build_response_format() -> dict:
    """
    OpenAI structured-output response_format dict (json_schema type).
    Mirrors the Groq script's pattern:
        response_format={"type": "json_schema", "json_schema": {...}}
    """
    return {
        "type": "json_schema",
        "json_schema": {
            "name": "GliomaClassificationOutput",
            "strict": True,
            "schema": _GLIOMA_SCHEMA,
        },
    }


# --- ID helpers (identical to Gemini script) ---------------------------------

def load_id_map(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    id_map: dict = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        center_id  = str(row[0]).strip() if row[0] else None
        subject_id = str(row[1]).strip() if row[1] else None
        brats_id   = str(row[4]).strip() if (row[4] and row[4] != "x") else None
        entry = {"center_id": center_id, "subject_id": subject_id}
        if center_id:
            id_map[center_id] = entry
        if brats_id:
            id_map[brats_id] = entry
    return id_map


# --- Ground-truth helpers (identical to Gemini script) -----------------------

def load_ground_truth(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def _idh(v):
        if v == "mutated":   return "mutant"
        if v == "wild type": return "wildtype"
        return None

    def _codeletion(v):
        if v in ("co-deleted", "rel. co-deleted"): return "codeleted"
        if v == "intact":                          return "non-codeleted"
        return None

    gt: dict = {}
    for row in rows[1:]:
        center_id = row[0]
        brats_id  = row[4] if (row[4] and row[4] != "x") else None
        entry = {
            "idh":        _idh(row[7]),
            "codeletion": _codeletion(row[8]),
            "grade":      row[9] if isinstance(row[9], int) else None,
            "dataset":    row[2],
        }
        if center_id:
            gt[center_id] = entry
        if brats_id:
            gt[brats_id] = entry
    return gt


# --- Metrics (identical to Gemini script) ------------------------------------

def _binary_metrics(y_true, y_pred, pos_label) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None}
    all_labels = sorted(set(y_true) | set(y_pred))
    neg_candidates = [l for l in all_labels if l != pos_label]
    neg_label = neg_candidates[0] if neg_candidates else (
        "wildtype" if pos_label == "mutant" else "non-codeleted"
    )
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, pos_label=pos_label, average="binary",
                   zero_division=0)
    cm  = confusion_matrix(y_true, y_pred, labels=[pos_label, neg_label])
    TP, FN = cm[0, 0], cm[0, 1]
    FP, TN = cm[1, 0], cm[1, 1]
    sens = TP / (TP + FN) if (TP + FN) > 0 else None
    spec = TN / (TN + FP) if (TN + FP) > 0 else None
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sens, 4) if sens is not None else None,
        "specificity": round(spec, 4) if spec is not None else None,
        "f1":          round(f1, 4),
    }


def _grade_metrics(y_true, y_pred) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None}
    classes = [2, 3, 4]
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, labels=classes, average="macro",
                   zero_division=0)
    cm = confusion_matrix(y_true, y_pred, labels=classes)
    sens_list, spec_list = [], []
    for i in range(len(classes)):
        TP = cm[i, i]
        FN = cm[i, :].sum() - TP
        FP = cm[:, i].sum() - TP
        TN = cm.sum() - TP - FN - FP
        sens_list.append(TP / (TP + FN) if (TP + FN) > 0 else 0.0)
        spec_list.append(TN / (TN + FP) if (TN + FP) > 0 else 0.0)
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sum(sens_list) / len(sens_list), 4),
        "specificity": round(sum(spec_list) / len(spec_list), 4),
        "f1":          round(f1, 4),
    }


def compute_metrics(predictions: dict, gt: dict) -> dict:
    data = defaultdict(lambda: {
        "idh_t": [], "idh_p": [],
        "cod_t": [], "cod_p": [],
        "grd_t": [], "grd_p": [],
    })
    for sid, pred in predictions.items():
        row = gt.get(sid)
        if row is None:
            continue
        ds = row["dataset"] or "Unknown"
        for bucket in ("whole_cohort", ds):
            d = data[bucket]
            if row["idh"] is not None and pred.get("idh") is not None:
                d["idh_t"].append(row["idh"]); d["idh_p"].append(pred["idh"])
            if row["codeletion"] is not None and pred.get("codeletion") is not None:
                d["cod_t"].append(row["codeletion"]); d["cod_p"].append(pred["codeletion"])
            if row["grade"] is not None and pred.get("grade") is not None:
                d["grd_t"].append(row["grade"]); d["grd_p"].append(pred["grade"])

    results = {}
    for bucket, d in data.items():
        results[bucket] = {
            "idh":        _binary_metrics(d["idh_t"], d["idh_p"], "mutant"),
            "codeletion": _binary_metrics(d["cod_t"], d["cod_p"], "codeleted"),
            "grade":      _grade_metrics(d["grd_t"], d["grd_p"]),
        }
    return results


def print_metrics(metrics: dict) -> None:
    col_w = 22
    header = (f"{'Subset':<{col_w}} {'Task':<12} {'N':>5}  "
              f"{'Acc':>6}  {'Sens':>6}  {'Spec':>6}  {'F1':>6}")
    sep = "=" * len(header)
    print(f"\n{sep}\nCLASSIFICATION METRICS\n{sep}")
    print(header)
    print("-" * len(header))
    task_labels = {"idh": "IDH", "codeletion": "1p/19q", "grade": "Grade"}
    buckets = ["whole_cohort"] + sorted(k for k in metrics if k != "whole_cohort")
    for bucket in buckets:
        for task_key, label in task_labels.items():
            m = metrics[bucket][task_key]
            def _fmt(v):
                return f"{v:.4f}" if v is not None else "  N/A"
            print(f"{bucket:<{col_w}} {label:<12} {m['n']:>5}  "
                  f"{_fmt(m['accuracy']):>6}  {_fmt(m['sensitivity']):>6}  "
                  f"{_fmt(m['specificity']):>6}  {_fmt(m['f1']):>6}")
        print()


# --- OpenAI client -----------------------------------------------------------

def make_client() -> openai.OpenAI:
    """
    Build an OpenAI client. Reads OPENAI_API_KEY from .env, falling back to
    the process environment if not present in .env.
    """
    config = dotenv_values(ENV_PATH)
    api_key = config.get("OPENAI_API_KEY")
    if api_key:
        return openai.OpenAI(api_key=api_key)
    # openai.OpenAI() will read OPENAI_API_KEY from the environment.
    return openai.OpenAI()


# --- State persistence -------------------------------------------------------

def state_path(output_dir: Path) -> Path:
    return output_dir / "batch_state.json"


def load_state(output_dir: Path) -> dict:
    p = state_path(output_dir)
    if p.exists():
        return json.load(open(p))
    return {"jobs": []}


def save_state(output_dir: Path, state: dict) -> None:
    with open(state_path(output_dir), "w") as f:
        json.dump(state, f, indent=2)


# --- Subject collection ------------------------------------------------------

def collect_subjects(json_type: str, output_dir: Path, model: str,
                     id_map: dict, force: bool) -> list:
    """
    Returns a list of dicts {folder_name, api_subject_id, prompt} for every
    subject still needing a prediction (skips ones whose output already exists,
    unless --force).
    """
    slug = model_slug(model)
    subjects = sorted(d.name for d in JSONS_DIR.iterdir() if d.is_dir())
    pending = []
    for folder_name in subjects:
        out_path = output_dir / f"{folder_name}_classification_{slug}.json"
        if out_path.exists() and not force:
            continue
        metadata_path = JSONS_DIR / folder_name / f"{folder_name}_metadata_{json_type}.json"
        if not metadata_path.exists():
            continue
        ids = id_map.get(folder_name)
        api_subject_id = ids["subject_id"] if ids else folder_name
        metadata = json.load(open(metadata_path))
        prompt = USER_PROMPT_TEMPLATE.format(
            subject_id=api_subject_id,
            metadata_json=json.dumps(metadata, indent=2, default=str),
        )
        pending.append({
            "folder_name": folder_name,
            "api_subject_id": api_subject_id,
            "prompt": prompt,
        })
    return pending


def chunked(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


# --- Phase: submit -----------------------------------------------------------

def phase_submit(client, output_dir, json_type, model, id_map, force, C, G, Y, W, D):
    pending = collect_subjects(json_type, output_dir, model, id_map, force)
    if not pending:
        print(f"{G}Nothing to submit{Style.RESET_ALL} — all subjects already have outputs.")
        return

    print(f"{C}Submitting {W}{len(pending)}{C} subjects in chunks of {CHUNK_SIZE} "
          f"using {W}{model}{C} (batch = 50% off).{Style.RESET_ALL}")

    state = load_state(output_dir)
    response_format = build_response_format()

    for ci, chunk in enumerate(chunked(pending, CHUNK_SIZE), 1):
        # Build JSONL content — one request object per line.
        lines = []
        keys = []
        for s in chunk:
            request_obj = {
                "custom_id": s["folder_name"],
                "method": "POST",
                "url": "/v1/chat/completions",
                "body": {
                    "model": model,
                    "messages": [
                        {"role": "system", "content": SYSTEM_PROMPT},
                        {"role": "user",   "content": s["prompt"]},
                    ],
                    "response_format": response_format,
                    # gpt-5.x are reasoning models: they reject `max_tokens`
                    # (use `max_completion_tokens`) and only accept the default
                    # temperature (1), so `temperature` is omitted entirely.
                    # The limit must be generous enough to cover *reasoning*
                    # tokens + the CoT JSON, or the model returns empty content.
                    "max_completion_tokens": 16000,
                    # Best-effort determinism: identical (seed, request) pairs
                    # are sampled the same way *as long as* system_fingerprint
                    # is unchanged. Captured back into each output for auditing.
                    "seed": SEED,
                },
            }
            lines.append(json.dumps(request_obj))
            keys.append(s["folder_name"])

        jsonl_bytes = "\n".join(lines).encode("utf-8")

        # Step 1: Upload the JSONL file.
        uploaded = client.files.create(
            file=(
                f"glioma_gpt55_{json_type}_chunk{ci}.jsonl",
                io.BytesIO(jsonl_bytes),
                "application/jsonl",
            ),
            purpose="batch",
        )

        # Step 2: Create the batch job referencing the uploaded file.
        batch = client.batches.create(
            input_file_id=uploaded.id,
            endpoint="/v1/chat/completions",
            completion_window="24h",
            metadata={"description": f"glioma-gpt55-{json_type}-chunk{ci}"},
        )

        state["jobs"].append({
            "batch_id":      batch.id,
            "input_file_id": uploaded.id,
            "model":         model,
            "json_type":     json_type,
            "chunk":         ci,
            "keys":          keys,   # ordered by submission; used as fallback
            "fetched":       False,
            "submitted_at":  datetime.now().isoformat(timespec="seconds"),
        })
        save_state(output_dir, state)
        print(f"  {G}SUBMITTED{Style.RESET_ALL} chunk {ci} "
              f"({len(keys)} subjects) -> {W}{batch.id}{Style.RESET_ALL}")

    print(f"{C}Saved job state -> {state_path(output_dir)}{Style.RESET_ALL}")
    print(f"{D}Now run with the same args using `fetch` (or `run`) to poll & "
          f"collect results.{Style.RESET_ALL}")


# --- Phase: poll + fetch -----------------------------------------------------

def _extract_json_text(response_line: dict):
    """
    Return (content_str, fingerprint, None) from an OpenAI batch output JSONL
    line, or (None, None, error_str) on failure.

    `fingerprint` is the response `system_fingerprint` (may be None): seed-based
    reproducibility only holds while this value is stable across runs.

    Output line format:
        {
          "custom_id": "...",
          "response": {
            "status_code": 200,
            "body": {"choices": [{"message": {"content": "..."}}],
                     "system_fingerprint": "fp_...", ...}
          },
          "error": null
        }
    """
    error = response_line.get("error")
    if error:
        return None, None, str(error)
    resp = response_line.get("response", {})
    if resp.get("status_code") != 200:
        return None, None, (f"status_code={resp.get('status_code')} "
                            f"body={resp.get('body')}")
    try:
        body = resp["body"]
        content = body["choices"][0]["message"]["content"]
        fingerprint = body.get("system_fingerprint")
        return content, fingerprint, None
    except (KeyError, IndexError) as e:
        return None, None, f"could not extract content ({e})"


def phase_fetch(client, output_dir, model, C, G, Y, R, W, D):
    slug = model_slug(model)
    state = load_state(output_dir)
    jobs = [j for j in state["jobs"] if not j.get("fetched")]
    if not jobs:
        print(f"{G}No outstanding jobs to fetch.{Style.RESET_ALL}")
        return

    TIMINGS_CSV = output_dir / "timings.csv"
    _csv_is_new = not TIMINGS_CSV.exists()
    _fh = TIMINGS_CSV.open("a", newline="")
    _writer = csv.DictWriter(_fh, fieldnames=["folder_name", "job_name", "timestamp"])
    if _csv_is_new:
        _writer.writeheader(); _fh.flush()

    for job_entry in jobs:
        batch_id = job_entry["batch_id"]
        keys     = job_entry["keys"]
        print(f"\n{C}Polling{Style.RESET_ALL} {W}{batch_id}{Style.RESET_ALL} "
              f"({len(keys)} subjects)")

        batch = client.batches.retrieve(batch_id)
        while batch.status not in TERMINAL_STATES:
            print(f"  {D}status={batch.status} — waiting {POLL_INTERVAL_S}s{Style.RESET_ALL}")
            time.sleep(POLL_INTERVAL_S)
            batch = client.batches.retrieve(batch_id)

        if batch.status != "completed":
            # Terminal failure/expiry — mark as done so it is never re-polled.
            print(f"  {R}{batch.status}{Style.RESET_ALL} — skipping (no results to save)")
            job_entry["fetched"] = True
            save_state(output_dir, state)
            continue

        output_file_id = batch.output_file_id
        if not output_file_id:
            # Batch reported "completed" but produced no output file — all
            # individual requests may have failed.  Check error_file_id for
            # diagnostics, then mark as done to prevent infinite re-polling.
            error_file_id = getattr(batch, "error_file_id", None)
            if error_file_id:
                print(f"  {R}ERROR{Style.RESET_ALL} batch completed with no output; "
                      f"errors in file {W}{error_file_id}{Style.RESET_ALL}")
            else:
                print(f"  {R}ERROR{Style.RESET_ALL} batch completed with neither "
                      f"output_file_id nor error_file_id — batch {batch_id}")
            job_entry["fetched"] = True
            save_state(output_dir, state)
            continue

        # Download and parse output JSONL.
        raw = client.files.content(output_file_id).read()
        output_lines = [
            json.loads(ln)
            for ln in raw.decode("utf-8").splitlines()
            if ln.strip()
        ]

        # Build lookup by custom_id (more robust than positional indexing).
        by_custom_id = {ln["custom_id"]: ln for ln in output_lines}
        if len(by_custom_id) != len(output_lines):
            print(f"  {Y}WARN{Style.RESET_ALL} duplicate custom_ids in output — "
                  f"last entry per id will be used.")

        saved = 0
        for folder_name in keys:
            response_line = by_custom_id.get(folder_name)
            if response_line is None:
                print(f"  {Y}WARN{Style.RESET_ALL} {folder_name}: not in output")
                continue
            text, fingerprint, err = _extract_json_text(response_line)
            if text is None:
                print(f"  {R}ERROR{Style.RESET_ALL} {folder_name}: {err}")
                continue
            try:
                result = json.loads(text)
            except json.JSONDecodeError as e:
                print(f"  {R}ERROR{Style.RESET_ALL} {folder_name}: bad JSON ({e})")
                continue

            result["subject_id"] = folder_name   # for the metrics join
            # Provenance for reproducibility auditing: same seed + same
            # fingerprint => same sampling; a fingerprint change explains drift.
            result["_seed"] = SEED
            result["_system_fingerprint"] = fingerprint
            out_path = output_dir / f"{folder_name}_classification_{slug}.json"
            with open(out_path, "w") as f:
                json.dump(result, f, indent=2)
            _writer.writerow({
                "folder_name": folder_name,
                "job_name":    batch_id,
                "timestamp":   datetime.now().isoformat(timespec="seconds"),
            })
            _fh.flush()
            saved += 1

        job_entry["fetched"] = True
        save_state(output_dir, state)
        print(f"  {G}SAVED{Style.RESET_ALL} {saved}/{len(keys)} outputs")

    _fh.close()
    print(f"{C}Timings -> {TIMINGS_CSV}{Style.RESET_ALL}")


# --- Phase: metrics ----------------------------------------------------------

def phase_metrics(output_dir, model, gt, C, G, D):
    slug = model_slug(model)
    print(f"\n{C}Collecting predictions for metric computation...{Style.RESET_ALL}")
    predictions: dict = {}
    for out_file in sorted(output_dir.glob(f"*_classification_{slug}.json")):
        result = json.load(open(out_file))
        sid = result.get("subject_id",
                          out_file.stem.replace(f"_classification_{slug}", ""))
        predictions[sid] = {
            "idh":        result.get("idh", {}).get("prediction"),
            "codeletion": result.get("one_p_nineteen_q", {}).get("prediction"),
            "grade":      result.get("who_grade", {}).get("prediction"),
        }
    if not predictions:
        print(f"{D}No output files found — run `submit` then `fetch` first.{Style.RESET_ALL}")
        return
    metrics = compute_metrics(predictions, gt)
    print_metrics(metrics)
    metrics_path = output_dir / "metrics.json"
    with open(metrics_path, "w") as f:
        json.dump(metrics, f, indent=2)
    print(f"{G}Metrics saved{Style.RESET_ALL} -> {metrics_path}")


# --- Phase: status -----------------------------------------------------------

def phase_status(client, output_dir, C, W, D):
    state = load_state(output_dir)
    if not state["jobs"]:
        print(f"{D}No jobs recorded in {state_path(output_dir)}{Style.RESET_ALL}")
        return
    for j in state["jobs"]:
        try:
            live = client.batches.retrieve(j["batch_id"]).status
        except Exception as e:                  # noqa: BLE001
            live = f"<lookup failed: {e}>"
        flag = "fetched" if j.get("fetched") else "pending"
        print(f"{C}{j['batch_id']}{Style.RESET_ALL}  chunk={j['chunk']}  "
              f"n={len(j['keys'])}  status={W}{live}{Style.RESET_ALL}  [{flag}]")


# --- CLI ---------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Classify gliomas via the OpenAI Batch API (50% cheaper).")
    p.add_argument("phase", choices=["submit", "fetch", "run", "metrics", "status"],
                   help="submit: create batch job(s); fetch: poll+download+save; "
                        "run: submit then fetch then metrics; metrics: score saved "
                        "outputs; status: print live job states.")
    p.add_argument("--json-type", choices=["btreport", "btreport_pp"],
                   help="Which metadata JSON variant to use (required except for `status`).")
    p.add_argument("--model", default=DEFAULT_MODEL,
                   help=f"OpenAI model id (default {DEFAULT_MODEL}).")
    p.add_argument("--force", action="store_true",
                   help="Re-submit subjects even if an output file already exists.")
    return p.parse_args()


def main() -> None:
    colorama_init(autoreset=True)
    C, G, Y, R, M, W, D = (Fore.CYAN, Fore.GREEN, Fore.YELLOW, Fore.RED,
                           Fore.MAGENTA, Style.BRIGHT, Style.DIM)

    args = parse_args()
    if args.phase != "status" and not args.json_type:
        raise SystemExit("--json-type is required for this phase.")

    output_dir = HERE / f"gpt55_outputs_{args.json_type}" if args.json_type \
        else HERE / "gpt55_outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    client = make_client()
    id_map = load_id_map(COHORT_XLSX) if args.phase in ("submit", "run") else {}
    gt     = load_ground_truth(COHORT_XLSX) if args.phase in ("run", "metrics") else {}

    if args.phase == "submit":
        phase_submit(client, output_dir, args.json_type, args.model, id_map,
                     args.force, C, G, Y, W, D)
    elif args.phase == "fetch":
        phase_fetch(client, output_dir, args.model, C, G, Y, R, W, D)
    elif args.phase == "metrics":
        phase_metrics(output_dir, args.model, gt, C, G, D)
    elif args.phase == "status":
        phase_status(client, output_dir, C, W, D)
    elif args.phase == "run":
        phase_submit(client, output_dir, args.json_type, args.model, id_map,
                     args.force, C, G, Y, W, D)
        phase_fetch(client, output_dir, args.model, C, G, Y, R, W, D)
        phase_metrics(output_dir, args.model, gt, C, G, D)


if __name__ == "__main__":
    main()
