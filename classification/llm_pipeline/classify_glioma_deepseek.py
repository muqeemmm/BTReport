"""
Glioma classification via DeepSeek (R1-0528 by default) on Vertex AI Model-as-a-Service.

This is the DeepSeek counterpart of `classify_glioma_llama4.py`. It performs the
SAME task — joint prediction of IDH status, 1p/19q co-deletion and CNS WHO grade
from structured pre-operative MRI metadata under WHO CNS5 (2021) — using the SAME
Chain-of-Thought Confidence Elicitation (CoT CE) prompt, schema, metrics, and the
SAME structure-retry / content-based extraction machinery, so DeepSeek is directly
comparable to Llama 4 and the other models in this folder.

WHY IT'S NEARLY IDENTICAL TO THE LLAMA 4 SCRIPT
-----------------------------------------------
DeepSeek is served on Vertex AI MaaS through the same **OpenAI-compatible**
ChatCompletions endpoint as Llama 4, with the same ADC (OAuth access token) auth.
Only three things differ:
  1. MODEL_ID uses the `deepseek-ai/...` publisher prefix.
  2. The serving region differs (DeepSeek R1 is us-central1; V3.2 supports the
     `global` endpoint). Set VERTEX_LOCATION accordingly — and "global" is handled.
  3. DeepSeek V3.1/R1 are reasoning models that may emit a <think>...</think> block
     (or a separate `reasoning_content`) before the JSON; parse_model_json() strips
     it. Billing is 100% Google Cloud / Vertex AI.

See `docs/llama4_vertex_setup.md` for the one-time prerequisites (enable the
Vertex AI API, accept the model licence on the DeepSeek model card, `gcloud auth`).

CONFIG comes from .env (or the process env):
    VERTEX_PROJECT_ID = your-gcp-project-id
    VERTEX_LOCATION   = us-central1   # region serving the chosen DeepSeek model
                                      # (use "global" for models that support it)
"""

import argparse
import csv
import json
import os
import re
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import google.auth
import google.auth.transport.requests
import openpyxl
from colorama import Fore, Style, init as colorama_init
from dotenv import dotenv_values
from openai import APIStatusError, BadRequestError, OpenAI, RateLimitError
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score
from tqdm import tqdm


# --- Prompt (identical to classify_glioma_groq.py) ---------------------------

SYSTEM_PROMPT = (
    "You are an expert neuroradiologist performing pre-operative molecular "
    "subtyping of adult-type diffuse glioma from structured MRI metadata, "
    "following the WHO CNS5 (2021) classification. For every task you reason "
    "step by step BEFORE stating any prediction or confidence score: first lay "
    "out the clinically grounded explanation, then commit to the prediction, "
    "then assign the confidence. Output MUST be a single valid JSON object that "
    "conforms exactly to the GliomaClassificationOutput schema. No markdown, no "
    "prose outside the JSON, no code fences -- JSON only."
)

USER_PROMPT_TEMPLATE = """# ROLE
You are an expert neuroradiologist.

# TASK
You are given structured pre-operative brain MRI metadata for a single patient
with a suspected **adult-type diffuse glioma**, defined according to the WHO
CNS5 molecular classification. Based on this metadata, predict the following:

  (a) IDH mutation status        -> {{"mutant", "wildtype"}}
  (b) 1p/19q co-deletion status  -> {{"codeleted", "non-codeleted"}}
  (c) CNS WHO grade              -> {{2, 3, 4}}

# CHAIN-OF-THOUGHT CONFIDENCE ELICITATION
For each of the three tasks you MUST work in the following order. Do not skip,
reorder, or shortcut these steps:

  1. **Reasoning first.** Write a concise, clinically grounded explanation that
     weighs the available metadata. Explicitly contrast the evidence that
     supports each candidate answer against the evidence that argues against
     it. Reach the prediction as the *conclusion* of this reasoning.
  2. **Supporting metadata fields.** List the exact field names (verbatim) that
     support the prediction you reasoned toward.
  3. **Contradicting features.** List the fields that argue against it.
  4. **Prediction.** State the prediction that follows from steps 1-3.
  5. **Confidence score** (between 0.0 and 1.0). Derive the score *from* the
     reasoning above -- it must reflect the balance of supporting vs.
     contradicting evidence you already articulated, NOT a number chosen before
     reasoning. Assign high confidence only when key features are well
     supported and few contradict; assign low confidence when evidence is
     thin, conflicting, or absent.

The JSON schema lists `reasoning`, `supporting_features` and
`contradicting_features` before `prediction` and `confidence` for exactly this
reason: generate them in that order.

# INPUT CONSTRAINTS
- Input consists **only of structured MRI-derived metadata** (no images).
- Do not assume access to data beyond what is provided.
- Treat null values as **"not assessed"** (do not infer).

# AVAILABLE MRI SEQUENCES
T1w, T2w, T2-FLAIR, T1-Gd.
**Not available:** DWI/ADC, perfusion (rCBV/DSC), MRS (2HG), SWI/GRE.

# RULES
1. Use ONLY the metadata provided; no hallucinated features or sequences.
2. Every claim in `reasoning` must be traceable to a field named verbatim in
   `supporting_features`.
3. Do NOT invoke diffusion, perfusion, spectroscopy, calcifications, or SWI.
4. If evidence is insufficient: give a best estimate, set confidence in
   0.50-0.60, and state the uncertainty in reasoning.
5. Confidence must be >= 0.50 for binary tasks (otherwise flip the prediction).
6. The confidence score must be the conclusion of the reasoning, not its
   premise -- never write the reasoning to justify a pre-chosen score.
7. Output must be **valid JSON only** (no markdown, no commentary).

# INPUT
Patient subject_id: {subject_id}
MRI metadata:
{metadata_json}
"""


# --- Config ------------------------------------------------------------------

# Vertex AI MaaS DeepSeek model id. The OpenAI-compatible endpoint expects the
# publisher-prefixed form "deepseek-ai/<model>". Swap to another DeepSeek MaaS id
# to benchmark a different model, e.g.:
#   deepseek-ai/deepseek-r1-0528-maas (default; reasoning, region us-central1)
#   deepseek-ai/deepseek-v3.1-maas    (hybrid thinking, region us-central1)
#   deepseek-ai/deepseek-v3.2-maas    (supports the "global" endpoint)
MODEL_ID    = "deepseek-ai/deepseek-r1-0528-maas"
MODEL_SLUG  = "deepseek_r1_0528"      # used in output filenames / metric keys
HERE        = Path(__file__).parent
SCHEMA_PATH = HERE / "glioma_classification_schema.json"
ENV_PATH    = Path(".env")
JSONS_DIR   = Path("data/Dataset_AKU_WHO/JSONs")
COHORT_XLSX = Path("data/dataset_table/cohort_merged.xlsx")

# Default region that serves the chosen DeepSeek model (overridable via .env).
# R1-0528 serves in us-central1. Use "global" only for models that support it (V3.2).
DEFAULT_LOCATION = "us-central1"

# DeepSeek MaaS has no Llama-Guard-style per-request safety toggle, so nothing
# extra is sent in the request body.
ENABLE_LLAMA_GUARD = False

# Output constraint mode. Vertex MaaS open models support STRICT structured output
# (https://docs.cloud.google.com/vertex-ai/generative-ai/docs/maas/capabilities/structured-output),
# so we default to constraining generation to glioma_classification_schema.json,
# exactly like the gpt55/groq scripts. R1 is a reasoning model, so it MIGHT reject
# `response_format`; if it does, the sync request loop auto-disables structured
# output for the rest of the run (and the prompt + <think>-stripping parser still
# apply). Modes: "json_schema" | "json_object" | "none".
RESPONSE_FORMAT_MODE = "json_schema"


def build_response_format():
    """OpenAI-style response_format for the configured mode (None if disabled)."""
    if RESPONSE_FORMAT_MODE == "json_schema":
        raw = json.load(open(SCHEMA_PATH))
        schema = {k: v for k, v in raw.items() if k != "$schema"}
        return {"type": "json_schema",
                "json_schema": {"name": "GliomaClassificationOutput",
                                "strict": True, "schema": schema}}
    if RESPONSE_FORMAT_MODE == "json_object":
        return {"type": "json_object"}
    return None

# Vertex MaaS models sometimes collapse the three tasks into a single block (a
# scalar `prediction` answering only IDH). It isn't deterministic, so re-asking —
# with a reinforcement nudge and a little temperature — usually fixes it. How many
# times to re-ask before giving up on a subject and skipping it (a later run retries).
STRUCTURE_RETRIES = 5

STRUCTURE_REINFORCE = (
    "Your previous response did not follow the required structure. Return a SINGLE "
    "JSON object with EXACTLY three top-level task objects named \"idh\", "
    "\"one_p_nineteen_q\", and \"who_grade\". Each task object must contain "
    "\"reasoning\", \"supporting_features\", \"contradicting_features\", "
    "\"prediction\", and \"confidence\". Do NOT merge the three tasks into one "
    "block, and do NOT put a single scalar string in \"prediction\". "
    "idh.prediction must be \"mutant\" or \"wildtype\"; "
    "one_p_nineteen_q.prediction must be \"codeleted\" or \"non-codeleted\"; "
    "who_grade.prediction must be the integer 2, 3, or 4. Output JSON only."
)


# --- Vertex AI (OpenAI-compatible) client ------------------------------------

class VertexTokenAuth:
    """
    Mints and refreshes a Google Cloud OAuth access token from Application
    Default Credentials, and rebuilds an OpenAI client pointed at the Vertex AI
    `endpoints/openapi` chat-completions endpoint. Access tokens are short-lived
    (~1 h); call `client()` before each request and it refreshes when near expiry.
    """

    def __init__(self, project_id: str, location: str):
        self.project_id = project_id
        self.location = location
        # The global endpoint has no region prefix on the host and uses
        # locations/global; regional endpoints prefix the host with the region.
        host = ("aiplatform.googleapis.com" if location == "global"
                else f"{location}-aiplatform.googleapis.com")
        self.base_url = (
            f"https://{host}/v1/"
            f"projects/{project_id}/locations/{location}/endpoints/openapi"
        )
        self._creds, _ = google.auth.default(
            scopes=["https://www.googleapis.com/auth/cloud-platform"]
        )
        self._request = google.auth.transport.requests.Request()
        self._client: OpenAI | None = None
        self._refresh()

    def _refresh(self) -> None:
        self._creds.refresh(self._request)
        self._client = OpenAI(base_url=self.base_url, api_key=self._creds.token)

    def client(self) -> OpenAI:
        # Refresh ~5 min before expiry (or if we have never minted a token).
        if (not self._creds.valid) or self._creds.expired:
            self._refresh()
        return self._client


def load_vertex_config() -> tuple[str, str]:
    """Read project id + location from .env, falling back to the process env."""
    config = dotenv_values(ENV_PATH)
    project = (config.get("VERTEX_PROJECT_ID") or os.environ.get("VERTEX_PROJECT_ID")
               or os.environ.get("GOOGLE_CLOUD_PROJECT"))
    location = (config.get("VERTEX_LOCATION") or os.environ.get("VERTEX_LOCATION")
                or DEFAULT_LOCATION)
    if not project:
        raise RuntimeError(
            "VERTEX_PROJECT_ID not set. Add it to .env (or export "
            "GOOGLE_CLOUD_PROJECT). See docs/llama4_vertex_setup.md."
        )
    return project, location


# --- ID helpers (identical to classify_glioma_groq.py) -----------------------

def load_id_map(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    id_map: dict = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        center_id  = str(row[0]).strip() if row[0] else None
        subject_id = str(row[1]).strip() if row[1] else None
        brats_id   = str(row[4]).strip() if (row[4] and row[4] != "x") else None
        entry = {"center_id": center_id, "subject_id": subject_id}
        if center_id:
            id_map[center_id] = entry
        if brats_id:
            id_map[brats_id] = entry
    return id_map


# --- Ground-truth helpers (identical to classify_glioma_groq.py) -------------

def load_ground_truth(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def _idh(v):
        if v == "mutated":   return "mutant"
        if v == "wild type": return "wildtype"
        return None

    def _codeletion(v):
        if v in ("co-deleted", "rel. co-deleted"): return "codeleted"
        if v == "intact":                          return "non-codeleted"
        return None

    gt: dict = {}
    for row in rows[1:]:
        center_id = row[0]
        brats_id  = row[4] if (row[4] and row[4] != "x") else None
        entry = {
            "idh":        _idh(row[7]),
            "codeletion": _codeletion(row[8]),
            "grade":      row[9] if isinstance(row[9], int) else None,
            "dataset":    row[2],
        }
        if center_id:
            gt[center_id] = entry
        if brats_id:
            gt[brats_id] = entry
    return gt


# --- Metrics (identical to classify_glioma_groq.py) --------------------------

def _binary_metrics(y_true, y_pred, pos_label) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None, "weighted_f1": None}
    all_labels = sorted(set(y_true) | set(y_pred))
    neg_candidates = [l for l in all_labels if l != pos_label]
    neg_label = neg_candidates[0] if neg_candidates else (
        "wildtype" if pos_label == "mutant" else "non-codeleted"
    )
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, pos_label=pos_label, average="binary",
                   zero_division=0)
    wf1 = f1_score(y_true, y_pred, average="weighted", zero_division=0)
    cm  = confusion_matrix(y_true, y_pred, labels=[pos_label, neg_label])
    TP, FN = cm[0, 0], cm[0, 1]
    FP, TN = cm[1, 0], cm[1, 1]
    sens = TP / (TP + FN) if (TP + FN) > 0 else None
    spec = TN / (TN + FP) if (TN + FP) > 0 else None
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sens, 4) if sens is not None else None,
        "specificity": round(spec, 4) if spec is not None else None,
        "f1":          round(f1, 4),
        "weighted_f1": round(wf1, 4),
    }


def _grade_metrics(y_true, y_pred) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None, "weighted_f1": None}
    classes = [2, 3, 4]
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, labels=classes, average="macro",
                   zero_division=0)
    wf1 = f1_score(y_true, y_pred, labels=classes, average="weighted",
                   zero_division=0)
    cm = confusion_matrix(y_true, y_pred, labels=classes)
    sens_list, spec_list = [], []
    for i in range(len(classes)):
        TP = cm[i, i]
        FN = cm[i, :].sum() - TP
        FP = cm[:, i].sum() - TP
        TN = cm.sum() - TP - FN - FP
        sens_list.append(TP / (TP + FN) if (TP + FN) > 0 else 0.0)
        spec_list.append(TN / (TN + FP) if (TN + FP) > 0 else 0.0)
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sum(sens_list) / len(sens_list), 4),
        "specificity": round(sum(spec_list) / len(spec_list), 4),
        "f1":          round(f1, 4),
        "weighted_f1": round(wf1, 4),
    }


def compute_metrics(predictions: dict, gt: dict) -> dict:
    data = defaultdict(lambda: {
        "idh_t": [], "idh_p": [],
        "cod_t": [], "cod_p": [],
        "grd_t": [], "grd_p": [],
    })
    for sid, pred in predictions.items():
        row = gt.get(sid)
        if row is None:
            continue
        ds = row["dataset"] or "Unknown"
        for bucket in ("whole_cohort", ds):
            d = data[bucket]
            if row["idh"] is not None and pred.get("idh") is not None:
                d["idh_t"].append(row["idh"]); d["idh_p"].append(pred["idh"])
            if row["codeletion"] is not None and pred.get("codeletion") is not None:
                d["cod_t"].append(row["codeletion"]); d["cod_p"].append(pred["codeletion"])
            if row["grade"] is not None and pred.get("grade") is not None:
                d["grd_t"].append(row["grade"]); d["grd_p"].append(pred["grade"])

    results = {}
    for bucket, d in data.items():
        results[bucket] = {
            "idh":        _binary_metrics(d["idh_t"], d["idh_p"], "mutant"),
            "codeletion": _binary_metrics(d["cod_t"], d["cod_p"], "codeleted"),
            "grade":      _grade_metrics(d["grd_t"], d["grd_p"]),
        }
    return results


def print_metrics(metrics: dict) -> None:
    col_w = 22
    header = (f"{'Subset':<{col_w}} {'Task':<12} {'N':>5}  "
              f"{'Acc':>6}  {'Sens':>6}  {'Spec':>6}  {'F1':>6}  {'WF1':>6}")
    sep = "=" * len(header)
    print(f"\n{sep}\nCLASSIFICATION METRICS\n{sep}")
    print(header)
    print("-" * len(header))
    task_labels = {"idh": "IDH", "codeletion": "1p/19q", "grade": "Grade"}
    buckets = ["whole_cohort"] + sorted(k for k in metrics if k != "whole_cohort")
    for bucket in buckets:
        for task_key, label in task_labels.items():
            m = metrics[bucket][task_key]
            def _fmt(v):
                return f"{v:.4f}" if v is not None else "  N/A"
            print(f"{bucket:<{col_w}} {label:<12} {m['n']:>5}  "
                  f"{_fmt(m['accuracy']):>6}  {_fmt(m['sensitivity']):>6}  "
                  f"{_fmt(m['specificity']):>6}  {_fmt(m['f1']):>6}  "
                  f"{_fmt(m['weighted_f1']):>6}")
        print()


def compute_and_save_metrics(output_dir) -> None:
    """Score already-saved classification outputs (no API calls)."""
    C, G, D = Fore.CYAN, Fore.GREEN, Style.DIM
    gt = load_ground_truth(COHORT_XLSX)
    print(f"\n{D}" + "─" * 72 + Style.RESET_ALL)
    print(f"{C}Collecting predictions for metric computation...{Style.RESET_ALL}")
    predictions: dict = {}
    for out_file in sorted(output_dir.glob(f"*_classification_{MODEL_SLUG}.json")):
        result = json.load(open(out_file))
        sid = result.get("subject_id",
                          out_file.stem.replace(f"_classification_{MODEL_SLUG}", ""))
        predictions[sid] = extract_predictions(result)
    metrics = compute_metrics(predictions, gt)
    print_metrics(metrics)
    metrics_path = output_dir / "metrics.json"
    with open(metrics_path, "w") as f:
        json.dump(metrics, f, indent=2)
    print(f"{G}Metrics saved{Style.RESET_ALL} → {metrics_path}")


# --- JSON extraction ---------------------------------------------------------

_FENCE_RE = re.compile(r"^```(?:json)?\s*|\s*```$", re.MULTILINE)
# DeepSeek V3.1/R1 are reasoning models and may prepend a chain-of-thought block
# wrapped in <think>...</think> before the JSON answer. Strip it before parsing.
_THINK_RE = re.compile(r"<think>.*?</think>", re.DOTALL | re.IGNORECASE)


def parse_model_json(content: str) -> dict:
    """
    Parse the model's JSON. DeepSeek on the Vertex OpenAI endpoint does not
    guarantee strict json_schema enforcement, and may wrap reasoning in a
    <think>...</think> block, so be defensive: drop the think block and any code
    fences, then extract the outermost JSON object if a direct parse fails.
    """
    text = _THINK_RE.sub("", content).strip()
    # If an unterminated <think> remains (truncated), keep everything after it.
    if "<think>" in text.lower():
        text = re.split(r"</?think>", text, flags=re.IGNORECASE)[-1].strip()
    text = _FENCE_RE.sub("", text).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start, end = text.find("{"), text.rfind("}")
        if start != -1 and end != -1 and end > start:
            return json.loads(text[start:end + 1])
        raise


# --- Prediction extraction (content-based, schema-tolerant) ------------------
# Llama on Vertex doesn't enforce the schema's field names, so across runs it
# emits the three task blocks under many different shapes: task-named top-level
# keys (e.g. "idh_status", "IDH mutation status", "cns_who_grade"), or a nested
# "predictions" list where each block carries a "type" field. We identify each
# task by the *content* of its key/type rather than a fixed alias list, so saved
# outputs score correctly without re-running inference. (The only unrecoverable
# shape is a single flat block that collapses all three tasks into one
# prediction — there is no way to tell which label belongs to which task.)

def _classify_key(name) -> str | None:
    s = str(name).lower()
    if "idh" in s:
        return "idh"
    # Covers "1p/19q", "1p19q", and the canonical schema key "one_p_nineteen_q".
    if ("1p" in s or "19q" in s or "nineteen" in s or "one_p" in s
            or "codelet" in s or "co-delet" in s):
        return "codeletion"
    if "grade" in s:
        return "grade"
    return None


def _iter_task_blocks(result: dict):
    """Yield (task, block) for every identifiable task block in an output JSON,
    handling a nested `predictions` list/dict or task-named top-level keys."""
    preds = result.get("predictions")
    if isinstance(preds, list):
        for b in preds:
            if isinstance(b, dict):
                task = _classify_key(b.get("type", ""))
                if task:
                    yield task, b
        return
    if isinstance(preds, dict):
        for k, v in preds.items():
            if isinstance(v, dict):
                task = _classify_key(k)
                if task:
                    yield task, v
        return
    for k, v in result.items():
        if isinstance(v, dict):
            task = _classify_key(k)
            if task:
                yield task, v


def _norm_idh(v):
    if not isinstance(v, str):
        return v
    s = v.strip().lower()
    if "mut" in s:  return "mutant"
    if "wild" in s: return "wildtype"
    return v


def _norm_cod(v):
    if not isinstance(v, str):
        return v
    s = v.strip().lower()
    if "non" in s or "intact" in s: return "non-codeleted"
    if "codelet" in s or "co-delet" in s: return "codeleted"
    return v


def _coerce_grade(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def extract_predictions(result: dict) -> dict:
    """Return {"idh", "codeletion", "grade"} from one output JSON, identifying
    task blocks by content and normalising label/grade formats."""
    out = {"idh": None, "codeletion": None, "grade": None}
    for task, block in _iter_task_blocks(result):
        if out[task] is None:
            out[task] = block.get("prediction")
    return {
        "idh":        _norm_idh(out["idh"]),
        "codeletion": _norm_cod(out["codeletion"]),
        "grade":      _coerce_grade(out["grade"]),
    }


def result_is_complete(result: dict) -> bool:
    """True iff all three tasks are present with valid label/grade values."""
    p = extract_predictions(result)
    return (p["idh"] in ("mutant", "wildtype")
            and p["codeletion"] in ("codeleted", "non-codeleted")
            and p["grade"] in (2, 3, 4))


# --- CLI ---------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Classify gliomas via DeepSeek on Vertex AI MaaS "
                    "(btreport or btreport_pp JSONs)."
    )
    parser.add_argument(
        "--json-type", choices=["btreport", "btreport_pp"], required=True,
        help="Which metadata JSON variant to use (btreport or btreport_pp).",
    )
    parser.add_argument(
        "--metrics-only", action="store_true",
        help="Skip inference (no API calls) and only recompute metrics from "
             "already-saved outputs.",
    )
    return parser.parse_args()


if __name__ == "__main__":
    colorama_init(autoreset=True)
    C, G, Y, R, M, W, D = (Fore.CYAN, Fore.GREEN, Fore.YELLOW, Fore.RED,
                           Fore.MAGENTA, Style.BRIGHT, Style.DIM)

    args      = parse_args()
    json_type = args.json_type

    OUTPUT_DIR = HERE / f"deepseek_outputs_{json_type}"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if args.metrics_only:
        compute_and_save_metrics(OUTPUT_DIR)
        raise SystemExit(0)

    TIMINGS_CSV = OUTPUT_DIR / "timings.csv"
    _csv_is_new = not TIMINGS_CSV.exists()
    _timings_fh = TIMINGS_CSV.open("a", newline="")
    _timings_writer = csv.DictWriter(
        _timings_fh, fieldnames=["folder_name", "elapsed_s", "timestamp"])
    if _csv_is_new:
        _timings_writer.writeheader()
        _timings_fh.flush()

    schema = json.load(open(SCHEMA_PATH))

    project, location = load_vertex_config()
    auth = VertexTokenAuth(project, location)
    print(f"{C}Vertex AI MaaS  |  project={W}{project}{C}  region={W}{location}{C}\n"
          f"Model: {W}{MODEL_ID}{C}  |  Output: {W}{OUTPUT_DIR.name}")

    id_map = load_id_map(COHORT_XLSX)
    subjects = sorted(d.name for d in JSONS_DIR.iterdir() if d.is_dir())

    pbar = tqdm(
        subjects, desc=f"{C}Classifying{Style.RESET_ALL}", unit="subj", ncols=90,
        bar_format=("{desc}: {percentage:3.0f}%|{bar:35}| "
                    "{n_fmt}/{total_fmt} [{elapsed}<{remaining}, {rate_fmt}]"),
        colour="cyan",
    )
    DIVIDER = f"{D}" + "─" * 72 + Style.RESET_ALL

    # DeepSeek MaaS takes no extra request-body knobs here.
    extra_body = None

    # Mutable so the request loop can disable structured output for the rest of the
    # run if R1 rejects it (the prompt + <think>-stripping parser still apply).
    rf_state = {"value": build_response_format()}

    for i, folder_name in enumerate(pbar, 1):
        pbar.set_postfix_str(f"{folder_name[:28]}", refresh=True)
        tqdm.write("")
        tqdm.write(DIVIDER)
        tqdm.write(f"{W}[{i}/{len(subjects)}]{Style.RESET_ALL}  "
                   f"{C}{folder_name}{Style.RESET_ALL}")

        ids = id_map.get(folder_name)
        if ids:
            center_id      = ids["center_id"]
            api_subject_id = ids["subject_id"]
        else:
            tqdm.write(f"  {Y}WARN{Style.RESET_ALL}  no xlsx entry for "
                       f"{folder_name!r} — using folder name as fallback")
            center_id = api_subject_id = folder_name

        metadata_path = JSONS_DIR / folder_name / f"{folder_name}_metadata_{json_type}.json"
        if not metadata_path.exists():
            tqdm.write(f"  {Y}SKIP{Style.RESET_ALL}  metadata not found: {metadata_path.name}")
            continue

        out_path = OUTPUT_DIR / f"{folder_name}_classification_{MODEL_SLUG}.json"
        if out_path.exists():
            tqdm.write(f"  {Y}SKIP{Style.RESET_ALL}  output already exists: {out_path.name}")
            continue

        tqdm.write(f"  {D}center_id={Style.RESET_ALL}{center_id}   "
                   f"{D}api_id={Style.RESET_ALL}{api_subject_id}")

        metadata    = json.load(open(metadata_path))
        user_prompt = USER_PROMPT_TEMPLATE.format(
            subject_id=api_subject_id,
            metadata_json=json.dumps(metadata, indent=2, default=str),
        )

        MAX_RETRIES = 4   # network/auth/rate-limit retries per request
        t0 = time.perf_counter()

        def _request(messages, temperature):
            """One chat call with network/auth/rate-limit retries. Returns the
            response object, or None on a terminal failure."""
            for attempt in range(1, MAX_RETRIES + 1):
                try:
                    rf = rf_state["value"]
                    return auth.client().chat.completions.create(
                        model=MODEL_ID,
                        messages=messages,
                        temperature=temperature,
                        seed=42,
                        # DeepSeek reasoning (<think>) consumes output tokens before
                        # the JSON, so allow plenty of headroom to avoid truncation.
                        max_tokens=8192,
                        # Strict structured output if the model accepts it (R1 may
                        # not — auto-disabled below on rejection).
                        **({"response_format": rf} if rf else {}),
                        **(extra_body or {}),
                    )
                except RateLimitError as e:
                    wait = min(2 ** attempt, 30)
                    tqdm.write(f"  {M}LIMIT{Style.RESET_ALL}  429 (attempt "
                               f"{attempt}/{MAX_RETRIES}) — backing off {wait}s  "
                               f"{D}({e}){Style.RESET_ALL}")
                    time.sleep(wait)
                except APIStatusError as e:
                    if e.status_code in (401, 403) and attempt < MAX_RETRIES:
                        tqdm.write(f"  {Y}AUTH{Style.RESET_ALL}  {e.status_code} — "
                                   f"refreshing token (attempt {attempt}/{MAX_RETRIES})")
                        auth._refresh()
                        time.sleep(1)
                        continue
                    if e.status_code >= 500 and attempt < MAX_RETRIES:
                        tqdm.write(f"  {Y}RETRY{Style.RESET_ALL}  server {e.status_code} "
                                   f"(attempt {attempt}/{MAX_RETRIES})")
                        time.sleep(2 ** attempt)
                        continue
                    tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  {e.status_code}: {e}")
                    return None
                except BadRequestError as e:
                    msg = str(e).lower()
                    # R1 may reject structured output — disable it for the rest of
                    # the run and retry (prompt + <think>-stripping parser remain).
                    if rf_state["value"] is not None and (
                            "response_format" in msg or "json_schema" in msg
                            or "schema" in msg or "structured" in msg):
                        tqdm.write(f"  {Y}NOTE{Style.RESET_ALL}  model rejected "
                                   f"structured output — disabling it for this run "
                                   f"and retrying.")
                        rf_state["value"] = None
                        continue
                    tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  bad request: {e}")
                    return None
            return None

        # Structure-retry loop: re-ask when the model collapses the three tasks
        # into one block, until we get a complete object or exhaust the budget.
        result = None
        for s_attempt in range(1, STRUCTURE_RETRIES + 1):
            messages = [
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": user_prompt},
            ]
            temperature = 0.0
            if s_attempt > 1:
                # Add a reinforcement nudge and a little temperature to break the
                # (otherwise repeated) degenerate single-block pattern.
                messages.append({"role": "user", "content": STRUCTURE_REINFORCE})
                temperature = 0.4

            response = _request(messages, temperature)
            if response is None:
                break
            try:
                cand = parse_model_json(response.choices[0].message.content)
            except json.JSONDecodeError as e:
                tqdm.write(f"  {Y}RETRY{Style.RESET_ALL}  unparseable JSON ({e}) "
                           f"(structure attempt {s_attempt}/{STRUCTURE_RETRIES})")
                continue
            result = cand
            if result_is_complete(cand):
                break
            tqdm.write(f"  {Y}RETRY{Style.RESET_ALL}  collapsed/incomplete output "
                       f"(structure attempt {s_attempt}/{STRUCTURE_RETRIES})")

        if result is None:
            tqdm.write(f"  {R}ERROR{Style.RESET_ALL}  no usable response — skipping {folder_name}")
            continue

        if not result_is_complete(result):
            # Don't save a collapsed result — leave no output file so a later run
            # retries this subject (keeps the cohort's metrics honest).
            tqdm.write(f"  {R}UNRESOLVED{Style.RESET_ALL}  still collapsed after "
                       f"{STRUCTURE_RETRIES} tries — skipping {folder_name} (re-run to retry)")
            continue

        elapsed = round(time.perf_counter() - t0, 2)
        result["subject_id"] = folder_name
        with open(out_path, "w") as f:
            json.dump(result, f, indent=2)

        _timings_writer.writerow({
            "folder_name": folder_name,
            "elapsed_s":   elapsed,
            "timestamp":   datetime.now().isoformat(timespec="seconds"),
        })
        _timings_fh.flush()
        tqdm.write(f"  {G}SAVED{Style.RESET_ALL}  {out_path.name}  {D}({elapsed}s){Style.RESET_ALL}")

    pbar.close()
    _timings_fh.close()
    print(f"{C}Timings saved{Style.RESET_ALL} → {TIMINGS_CSV}")

    compute_and_save_metrics(OUTPUT_DIR)
