"""
Glioma classification via Meta Llama 4 Maverick on Vertex AI — BATCH prediction.

This is the batch counterpart of `classify_glioma_llama4.py` (synchronous) and
the Vertex sibling of `classify_glioma_gpt55.py` (OpenAI Batch API). It performs
the SAME task — joint IDH / 1p19q / WHO-grade prediction from structured MRI
metadata under WHO CNS5 (2021) — with the SAME prompt, schema and metrics, but
routes inference through a Vertex AI **batchPredictionJob** to get the documented
50% discount versus the synchronous endpoint.

WHY THIS LOOKS ALMOST EXACTLY LIKE THE gpt55 BATCH SCRIPT
--------------------------------------------------------
Vertex AI batch for Llama consumes the **OpenAI batch JSONL schema** — each line
is {"custom_id", "method", "url": "/v1/chat/completions", "body": {...}} — i.e.
identical to what `classify_glioma_gpt55.py` already builds. So the three-phase
submit -> poll -> fetch shape and the `custom_id`-based result matching are reused
verbatim. Only the transport changes:

    OpenAI Batch API                  Vertex AI batchPredictionJob
    ----------------------------      --------------------------------------------
    client.files.create(...)     ->   upload JSONL to a Cloud Storage bucket (gs://)
    client.batches.create(...)   ->   POST .../batchPredictionJobs (REST, ADC auth)
    client.batches.retrieve(...) ->   GET  .../batchPredictionJobs/<id>
    client.files.content(...)    ->   download the output JSONL from gs://

Billing is 100% Google Cloud / Vertex AI. See docs/llama4_vertex_setup.md for the
one-time prerequisites; the batch path additionally needs a Cloud Storage bucket
and `roles/storage.objectAdmin` on it.

CONFIG (from .env or the process env):
    VERTEX_PROJECT_ID = your-gcp-project-id
    VERTEX_LOCATION   = us-east5                     # region serving Llama 4 MaaS
    VERTEX_GCS_BUCKET = your-bucket-name             # for batch I/O staging

CLI sub-commands (same as the gpt55 script):
    submit | fetch | run | metrics | status
plus a persisted `batch_state.json` so a submitted job can be picked up later.
"""

import argparse
import csv
import json
import os
import re
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import google.auth
import requests
from google.auth.transport.requests import AuthorizedSession
from google.cloud import storage
import openpyxl
from colorama import Fore, Style, init as colorama_init
from dotenv import dotenv_values
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score


# --- Prompt (identical to classify_glioma_groq.py) ---------------------------

SYSTEM_PROMPT = (
    "You are an expert neuroradiologist performing pre-operative molecular "
    "subtyping of adult-type diffuse glioma from structured MRI metadata, "
    "following the WHO CNS5 (2021) classification. For every task you reason "
    "step by step BEFORE stating any prediction or confidence score: first lay "
    "out the clinically grounded explanation, then commit to the prediction, "
    "then assign the confidence. Output MUST be a single valid JSON object that "
    "conforms exactly to the GliomaClassificationOutput schema. No markdown, no "
    "prose outside the JSON, no code fences -- JSON only."
)

USER_PROMPT_TEMPLATE = """# ROLE
You are an expert neuroradiologist.

# TASK
You are given structured pre-operative brain MRI metadata for a single patient
with a suspected **adult-type diffuse glioma**, defined according to the WHO
CNS5 molecular classification. Based on this metadata, predict the following:

  (a) IDH mutation status        -> {{"mutant", "wildtype"}}
  (b) 1p/19q co-deletion status  -> {{"codeleted", "non-codeleted"}}
  (c) CNS WHO grade              -> {{2, 3, 4}}

# CHAIN-OF-THOUGHT CONFIDENCE ELICITATION
For each of the three tasks you MUST work in the following order. Do not skip,
reorder, or shortcut these steps:

  1. **Reasoning first.** Write a concise, clinically grounded explanation that
     weighs the available metadata. Explicitly contrast the evidence that
     supports each candidate answer against the evidence that argues against
     it. Reach the prediction as the *conclusion* of this reasoning.
  2. **Supporting metadata fields.** List the exact field names (verbatim) that
     support the prediction you reasoned toward.
  3. **Contradicting features.** List the fields that argue against it.
  4. **Prediction.** State the prediction that follows from steps 1-3.
  5. **Confidence score** (between 0.0 and 1.0). Derive the score *from* the
     reasoning above -- it must reflect the balance of supporting vs.
     contradicting evidence you already articulated, NOT a number chosen before
     reasoning. Assign high confidence only when key features are well
     supported and few contradict; assign low confidence when evidence is
     thin, conflicting, or absent.

The JSON schema lists `reasoning`, `supporting_features` and
`contradicting_features` before `prediction` and `confidence` for exactly this
reason: generate them in that order.

# INPUT CONSTRAINTS
- Input consists **only of structured MRI-derived metadata** (no images).
- Do not assume access to data beyond what is provided.
- Treat null values as **"not assessed"** (do not infer).

# AVAILABLE MRI SEQUENCES
T1w, T2w, T2-FLAIR, T1-Gd.
**Not available:** DWI/ADC, perfusion (rCBV/DSC), MRS (2HG), SWI/GRE.

# RULES
1. Use ONLY the metadata provided; no hallucinated features or sequences.
2. Every claim in `reasoning` must be traceable to a field named verbatim in
   `supporting_features`.
3. Do NOT invoke diffusion, perfusion, spectroscopy, calcifications, or SWI.
4. If evidence is insufficient: give a best estimate, set confidence in
   0.50-0.60, and state the uncertainty in reasoning.
5. Confidence must be >= 0.50 for binary tasks (otherwise flip the prediction).
6. The confidence score must be the conclusion of the reasoning, not its
   premise -- never write the reasoning to justify a pre-chosen score.
7. Output must be **valid JSON only** (no markdown, no commentary).

# INPUT
Patient subject_id: {subject_id}
MRI metadata:
{metadata_json}
"""


# --- Config ------------------------------------------------------------------

MODEL_ID    = "meta/llama-4-maverick-17b-128e-instruct-maas"   # per-line body model
PUBLISHER_MODEL = "publishers/meta/models/llama-4-maverick-17b-128e-instruct-maas"
MODEL_SLUG  = "llama_4_maverick"
HERE        = Path(__file__).parent
SCHEMA_PATH = HERE / "glioma_classification_schema.json"
ENV_PATH    = Path(".env")
JSONS_DIR   = Path("data/Dataset_AKU_WHO/JSONs")
COHORT_XLSX = Path("data/dataset_table/cohort_merged.xlsx")

DEFAULT_LOCATION = "us-east5"
CHUNK_SIZE       = 200       # subjects per batch job (one GCS JSONL per chunk)
POLL_INTERVAL_S  = 60        # seconds between job-state polls
SEED             = 42

# Output constraint mode. Vertex MaaS open models support STRICT structured output,
# so we constrain generation to glioma_classification_schema.json (like gpt55/groq),
# preventing the key-drift/collapsed-block failures that plain json_object allowed.
# Batch can't adaptively fall back mid-job: if a model rejects json_schema, validate
# with the sync script first or set this to "json_object"/"none".
# Modes: "json_schema" | "json_object" | "none".
RESPONSE_FORMAT_MODE = "json_schema"


def build_response_format():
    """OpenAI-style response_format for the configured mode (None if disabled)."""
    if RESPONSE_FORMAT_MODE == "json_schema":
        raw = json.load(open(SCHEMA_PATH))
        schema = {k: v for k, v in raw.items() if k != "$schema"}
        return {"type": "json_schema",
                "json_schema": {"name": "GliomaClassificationOutput",
                                "strict": True, "schema": schema}}
    if RESPONSE_FORMAT_MODE == "json_object":
        return {"type": "json_object"}
    return None


# Vertex batchPredictionJob terminal states.
TERMINAL_STATES = {
    "JOB_STATE_SUCCEEDED", "JOB_STATE_FAILED",
    "JOB_STATE_CANCELLED", "JOB_STATE_EXPIRED",
}


# --- Vertex config / auth ----------------------------------------------------

def load_vertex_config() -> tuple[str, str, str]:
    """Read project id, location and GCS bucket from .env (env fallback)."""
    config = dotenv_values(ENV_PATH)
    project = (config.get("VERTEX_PROJECT_ID") or os.environ.get("VERTEX_PROJECT_ID")
               or os.environ.get("GOOGLE_CLOUD_PROJECT"))
    location = (config.get("VERTEX_LOCATION") or os.environ.get("VERTEX_LOCATION")
                or DEFAULT_LOCATION)
    bucket = (config.get("VERTEX_GCS_BUCKET") or os.environ.get("VERTEX_GCS_BUCKET"))
    if not project:
        raise RuntimeError("VERTEX_PROJECT_ID not set (see docs/llama4_vertex_setup.md).")
    if not bucket:
        raise RuntimeError("VERTEX_GCS_BUCKET not set — batch needs a Cloud Storage "
                           "bucket for I/O staging (see docs/llama4_vertex_setup.md).")
    return project, location, bucket


def make_session() -> AuthorizedSession:
    """ADC-authenticated HTTP session (auto-refreshes the access token)."""
    creds, _ = google.auth.default(
        scopes=["https://www.googleapis.com/auth/cloud-platform"])
    return AuthorizedSession(creds)


def api_root(location: str, project: str) -> str:
    return (f"https://{location}-aiplatform.googleapis.com/v1/"
            f"projects/{project}/locations/{location}")


# (connect, read) timeouts in seconds for control-plane GET/POST calls.
HTTP_TIMEOUT = (10, 60)


def get_json(session, url, retries: int = 6) -> dict:
    """
    GET `url` and return parsed JSON, tolerating the transient network timeouts
    and 5xx responses that occur over a multi-hour batch poll. Retries with
    exponential backoff instead of letting a single hiccup crash the run.
    """
    last = None
    for attempt in range(1, retries + 1):
        try:
            r = session.get(url, timeout=HTTP_TIMEOUT)
            if r.status_code >= 500:
                last = f"{r.status_code}: {r.text[:200]}"
                raise requests.exceptions.HTTPError(last)
            r.raise_for_status()
            return r.json()
        except requests.exceptions.RequestException as e:
            last = str(e)
            wait = min(2 ** attempt, 60)
            print(f"  {Style.DIM}transient poll error (attempt {attempt}/{retries}): "
                  f"{e} — retrying in {wait}s{Style.RESET_ALL}")
            time.sleep(wait)
    raise RuntimeError(f"GET {url} failed after {retries} retries: {last}")


# --- Structured-output schema ------------------------------------------------

with open(SCHEMA_PATH) as _f:
    _GLIOMA_SCHEMA_RAW = json.load(_f)
_GLIOMA_SCHEMA = {k: v for k, v in _GLIOMA_SCHEMA_RAW.items() if k != "$schema"}


# --- ID helpers (identical to classify_glioma_groq.py) -----------------------

def load_id_map(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    id_map: dict = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        center_id  = str(row[0]).strip() if row[0] else None
        subject_id = str(row[1]).strip() if row[1] else None
        brats_id   = str(row[4]).strip() if (row[4] and row[4] != "x") else None
        entry = {"center_id": center_id, "subject_id": subject_id}
        if center_id:
            id_map[center_id] = entry
        if brats_id:
            id_map[brats_id] = entry
    return id_map


# --- Ground-truth helpers (identical to classify_glioma_groq.py) -------------

def load_ground_truth(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def _idh(v):
        if v == "mutated":   return "mutant"
        if v == "wild type": return "wildtype"
        return None

    def _codeletion(v):
        if v in ("co-deleted", "rel. co-deleted"): return "codeleted"
        if v == "intact":                          return "non-codeleted"
        return None

    gt: dict = {}
    for row in rows[1:]:
        center_id = row[0]
        brats_id  = row[4] if (row[4] and row[4] != "x") else None
        entry = {
            "idh":        _idh(row[7]),
            "codeletion": _codeletion(row[8]),
            "grade":      row[9] if isinstance(row[9], int) else None,
            "dataset":    row[2],
        }
        if center_id:
            gt[center_id] = entry
        if brats_id:
            gt[brats_id] = entry
    return gt


# --- Metrics (identical to classify_glioma_groq.py) --------------------------

def _binary_metrics(y_true, y_pred, pos_label) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None, "weighted_f1": None}
    all_labels = sorted(set(y_true) | set(y_pred))
    neg_candidates = [l for l in all_labels if l != pos_label]
    neg_label = neg_candidates[0] if neg_candidates else (
        "wildtype" if pos_label == "mutant" else "non-codeleted"
    )
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, pos_label=pos_label, average="binary",
                   zero_division=0)
    wf1 = f1_score(y_true, y_pred, average="weighted", zero_division=0)
    cm  = confusion_matrix(y_true, y_pred, labels=[pos_label, neg_label])
    TP, FN = cm[0, 0], cm[0, 1]
    FP, TN = cm[1, 0], cm[1, 1]
    sens = TP / (TP + FN) if (TP + FN) > 0 else None
    spec = TN / (TN + FP) if (TN + FP) > 0 else None
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sens, 4) if sens is not None else None,
        "specificity": round(spec, 4) if spec is not None else None,
        "f1":          round(f1, 4),
        "weighted_f1": round(wf1, 4),
    }


def _grade_metrics(y_true, y_pred) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None, "weighted_f1": None}
    classes = [2, 3, 4]
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, labels=classes, average="macro",
                   zero_division=0)
    wf1 = f1_score(y_true, y_pred, labels=classes, average="weighted",
                   zero_division=0)
    cm = confusion_matrix(y_true, y_pred, labels=classes)
    sens_list, spec_list = [], []
    for i in range(len(classes)):
        TP = cm[i, i]
        FN = cm[i, :].sum() - TP
        FP = cm[:, i].sum() - TP
        TN = cm.sum() - TP - FN - FP
        sens_list.append(TP / (TP + FN) if (TP + FN) > 0 else 0.0)
        spec_list.append(TN / (TN + FP) if (TN + FP) > 0 else 0.0)
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sum(sens_list) / len(sens_list), 4),
        "specificity": round(sum(spec_list) / len(spec_list), 4),
        "f1":          round(f1, 4),
        "weighted_f1": round(wf1, 4),
    }


def compute_metrics(predictions: dict, gt: dict) -> dict:
    data = defaultdict(lambda: {
        "idh_t": [], "idh_p": [],
        "cod_t": [], "cod_p": [],
        "grd_t": [], "grd_p": [],
    })
    for sid, pred in predictions.items():
        row = gt.get(sid)
        if row is None:
            continue
        ds = row["dataset"] or "Unknown"
        for bucket in ("whole_cohort", ds):
            d = data[bucket]
            if row["idh"] is not None and pred.get("idh") is not None:
                d["idh_t"].append(row["idh"]); d["idh_p"].append(pred["idh"])
            if row["codeletion"] is not None and pred.get("codeletion") is not None:
                d["cod_t"].append(row["codeletion"]); d["cod_p"].append(pred["codeletion"])
            if row["grade"] is not None and pred.get("grade") is not None:
                d["grd_t"].append(row["grade"]); d["grd_p"].append(pred["grade"])

    results = {}
    for bucket, d in data.items():
        results[bucket] = {
            "idh":        _binary_metrics(d["idh_t"], d["idh_p"], "mutant"),
            "codeletion": _binary_metrics(d["cod_t"], d["cod_p"], "codeleted"),
            "grade":      _grade_metrics(d["grd_t"], d["grd_p"]),
        }
    return results


def print_metrics(metrics: dict) -> None:
    col_w = 22
    header = (f"{'Subset':<{col_w}} {'Task':<12} {'N':>5}  "
              f"{'Acc':>6}  {'Sens':>6}  {'Spec':>6}  {'F1':>6}  {'WF1':>6}")
    sep = "=" * len(header)
    print(f"\n{sep}\nCLASSIFICATION METRICS\n{sep}")
    print(header)
    print("-" * len(header))
    task_labels = {"idh": "IDH", "codeletion": "1p/19q", "grade": "Grade"}
    buckets = ["whole_cohort"] + sorted(k for k in metrics if k != "whole_cohort")
    for bucket in buckets:
        for task_key, label in task_labels.items():
            m = metrics[bucket][task_key]
            def _fmt(v):
                return f"{v:.4f}" if v is not None else "  N/A"
            print(f"{bucket:<{col_w}} {label:<12} {m['n']:>5}  "
                  f"{_fmt(m['accuracy']):>6}  {_fmt(m['sensitivity']):>6}  "
                  f"{_fmt(m['specificity']):>6}  {_fmt(m['f1']):>6}  "
                  f"{_fmt(m['weighted_f1']):>6}")
        print()


# --- JSON extraction (defensive — Vertex Llama doesn't guarantee strict schema)

_FENCE_RE = re.compile(r"^```(?:json)?\s*|\s*```$", re.MULTILINE)


def parse_model_json(content: str) -> dict:
    text = _FENCE_RE.sub("", content.strip()).strip()
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        start, end = text.find("{"), text.rfind("}")
        if start != -1 and end != -1 and end > start:
            return json.loads(text[start:end + 1])
        raise


# --- Prediction extraction (content-based, schema-tolerant) ------------------
# Llama on Vertex doesn't enforce the schema's field names, so across runs it
# emits the three task blocks under many different shapes: task-named top-level
# keys (e.g. "idh_status", "IDH mutation status", "cns_who_grade"), or a nested
# "predictions" list where each block carries a "type" field. We identify each
# task by the *content* of its key/type rather than a fixed alias list. (The only
# unrecoverable shape is a single flat block that collapses all three tasks into
# one prediction.)

def _classify_key(name) -> str | None:
    s = str(name).lower()
    if "idh" in s:
        return "idh"
    # Covers "1p/19q", "1p19q", and the canonical schema key "one_p_nineteen_q".
    if ("1p" in s or "19q" in s or "nineteen" in s or "one_p" in s
            or "codelet" in s or "co-delet" in s):
        return "codeletion"
    if "grade" in s:
        return "grade"
    return None


def _iter_task_blocks(result: dict):
    """Yield (task, block) for every identifiable task block, handling a nested
    `predictions` list/dict or task-named top-level keys."""
    preds = result.get("predictions")
    if isinstance(preds, list):
        for b in preds:
            if isinstance(b, dict):
                task = _classify_key(b.get("type", ""))
                if task:
                    yield task, b
        return
    if isinstance(preds, dict):
        for k, v in preds.items():
            if isinstance(v, dict):
                task = _classify_key(k)
                if task:
                    yield task, v
        return
    for k, v in result.items():
        if isinstance(v, dict):
            task = _classify_key(k)
            if task:
                yield task, v


def _norm_idh(v):
    if not isinstance(v, str):
        return v
    s = v.strip().lower()
    if "mut" in s:  return "mutant"
    if "wild" in s: return "wildtype"
    return v


def _norm_cod(v):
    if not isinstance(v, str):
        return v
    s = v.strip().lower()
    if "non" in s or "intact" in s: return "non-codeleted"
    if "codelet" in s or "co-delet" in s: return "codeleted"
    return v


def _coerce_grade(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def extract_predictions(result: dict) -> dict:
    """Return {"idh", "codeletion", "grade"} from one output JSON, identifying
    task blocks by content and normalising label/grade formats."""
    out = {"idh": None, "codeletion": None, "grade": None}
    for task, block in _iter_task_blocks(result):
        if out[task] is None:
            out[task] = block.get("prediction")
    return {
        "idh":        _norm_idh(out["idh"]),
        "codeletion": _norm_cod(out["codeletion"]),
        "grade":      _coerce_grade(out["grade"]),
    }


# --- State persistence (identical pattern to gpt55 script) -------------------

def state_path(output_dir: Path) -> Path:
    return output_dir / "batch_state.json"


def load_state(output_dir: Path) -> dict:
    p = state_path(output_dir)
    return json.load(open(p)) if p.exists() else {"jobs": []}


def save_state(output_dir: Path, state: dict) -> None:
    with open(state_path(output_dir), "w") as f:
        json.dump(state, f, indent=2)


# --- Subject collection (identical to gpt55 script) --------------------------

def collect_subjects(json_type, output_dir, id_map, force, in_flight=None) -> list:
    """
    Subjects still needing a prediction. Skips any that already have an output
    file, and (unless --force) any that are already covered by an unfetched batch
    job (`in_flight`) so repeated `submit`/`run` calls don't create duplicate —
    and duplicately billed — jobs for the same subjects.
    """
    in_flight = in_flight or set()
    subjects = sorted(d.name for d in JSONS_DIR.iterdir() if d.is_dir())
    pending = []
    for folder_name in subjects:
        out_path = output_dir / f"{folder_name}_classification_{MODEL_SLUG}.json"
        if out_path.exists() and not force:
            continue
        if folder_name in in_flight and not force:
            continue
        metadata_path = JSONS_DIR / folder_name / f"{folder_name}_metadata_{json_type}.json"
        if not metadata_path.exists():
            continue
        ids = id_map.get(folder_name)
        api_subject_id = ids["subject_id"] if ids else folder_name
        metadata = json.load(open(metadata_path))
        prompt = USER_PROMPT_TEMPLATE.format(
            subject_id=api_subject_id,
            metadata_json=json.dumps(metadata, indent=2, default=str),
        )
        pending.append({"folder_name": folder_name, "prompt": prompt})
    return pending


def chunked(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def build_request_line(folder_name: str, prompt: str) -> dict:
    """One OpenAI-schema batch request line (Vertex consumes this format)."""
    body = {
        "model": MODEL_ID,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user",   "content": prompt},
        ],
        "temperature": 0.0,
        "seed": SEED,
        "max_tokens": 4096,
    }
    # Strict structured output (json_schema) constrains generation to the schema;
    # the defensive content-based extractor remains a safety net at read time.
    rf = build_response_format()
    if rf:
        body["response_format"] = rf
    return {
        "custom_id": folder_name,
        "method": "POST",
        "url": "/v1/chat/completions",
        "body": body,
    }


# --- Phase: submit -----------------------------------------------------------

def phase_submit(session, gcs, bucket, api, output_dir, json_type, id_map,
                 force, C, G, Y, W, D) -> list:
    """Submit batch job(s). Returns the resource names of the jobs created."""
    state = load_state(output_dir)

    # Subjects already covered by an unfetched job — don't resubmit them (avoids
    # duplicate, duplicately-billed jobs). Fetch those first, or use --force.
    in_flight = {k for j in state["jobs"] if not j.get("fetched")
                 for k in j.get("keys", [])}
    pending = collect_subjects(json_type, output_dir, id_map, force, in_flight)
    if not pending:
        if in_flight:
            print(f"{G}Nothing to submit{Style.RESET_ALL} — {len(in_flight)} subject(s) "
                  f"are already in unfetched jobs. Run {W}fetch{Style.RESET_ALL}{G} first "
                  f"(or use {W}--force{Style.RESET_ALL}{G} to resubmit anyway).")
        else:
            print(f"{G}Nothing to submit{Style.RESET_ALL} — all subjects already have outputs.")
        return []
    if in_flight and not force:
        print(f"{D}Skipping {len(in_flight)} subject(s) already in unfetched jobs.{Style.RESET_ALL}")

    print(f"{C}Submitting {W}{len(pending)}{C} subjects in chunks of {CHUNK_SIZE} "
          f"to Vertex batch (50% off).{Style.RESET_ALL}")

    submitted_names: list = []
    bucket_obj = gcs.bucket(bucket)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")

    for ci, chunk in enumerate(chunked(pending, CHUNK_SIZE), 1):
        keys = [s["folder_name"] for s in chunk]
        jsonl = "\n".join(json.dumps(build_request_line(s["folder_name"], s["prompt"]))
                          for s in chunk)

        base       = f"llama4_batch/{json_type}/{stamp}/chunk{ci}"
        input_blob = f"{base}/input.jsonl"
        output_pre = f"{base}/output/"

        # Step 1: upload JSONL input to Cloud Storage.
        bucket_obj.blob(input_blob).upload_from_string(
            jsonl, content_type="application/jsonl")
        input_uri  = f"gs://{bucket}/{input_blob}"
        output_uri = f"gs://{bucket}/{output_pre}"

        # Step 2: create the batchPredictionJob (REST).
        payload = {
            "displayName": f"glioma-llama4-{json_type}-chunk{ci}-{stamp}",
            "model": PUBLISHER_MODEL,
            "inputConfig": {
                "instancesFormat": "jsonl",
                "gcsSource": {"uris": input_uri},
            },
            "outputConfig": {
                "predictionsFormat": "jsonl",
                "gcsDestination": {"outputUriPrefix": output_uri},
            },
        }
        r = session.post(f"{api}/batchPredictionJobs", json=payload)
        if r.status_code >= 400:
            print(f"  {Y}ERROR{Style.RESET_ALL} chunk {ci} submit failed "
                  f"({r.status_code}): {r.text}")
            continue
        job = r.json()

        state["jobs"].append({
            "name":        job["name"],          # projects/.../batchPredictionJobs/<id>
            "json_type":   json_type,
            "chunk":       ci,
            "keys":        keys,
            "input_uri":   input_uri,
            "output_uri":  output_uri,
            "fetched":     False,
            "submitted_at": datetime.now().isoformat(timespec="seconds"),
        })
        save_state(output_dir, state)
        submitted_names.append(job["name"])
        print(f"  {G}SUBMITTED{Style.RESET_ALL} chunk {ci} ({len(keys)} subjects) "
              f"-> {W}{job['name'].split('/')[-1]}{Style.RESET_ALL}")

    print(f"{C}Saved job state -> {state_path(output_dir)}{Style.RESET_ALL}")
    print(f"{D}Run `fetch` (or `run`) with the same args to poll & collect.{Style.RESET_ALL}")
    return submitted_names


# --- Phase: poll + fetch -----------------------------------------------------

def _extract_from_line(line: dict):
    """
    Pull (custom_id, content_str, error) from one Vertex/OpenAI batch output line.
    Handles the common shapes defensively:
        {"custom_id","response":{"status_code","body":{"choices":[...]}}}
        {"custom_id","response":{"choices":[...]}}
        {"custom_id","prediction":{"choices":[...]}}
    """
    cid = line.get("custom_id") or line.get("id")
    if line.get("error"):
        return cid, None, str(line["error"])
    resp = line.get("response")
    candidate = None
    if isinstance(resp, dict):
        if resp.get("status_code") not in (None, 200):
            return cid, None, f"status_code={resp.get('status_code')} body={resp.get('body')}"
        candidate = resp.get("body", resp)
    candidate = candidate or line.get("prediction") or line
    try:
        return cid, candidate["choices"][0]["message"]["content"], None
    except (KeyError, IndexError, TypeError):
        return cid, None, f"no choices (keys={list(line.keys())})"


def _download_output_lines(gcs, bucket, output_uri):
    """List + read every .jsonl produced under the job's output prefix."""
    prefix = output_uri.replace(f"gs://{bucket}/", "")
    lines = []
    for blob in gcs.bucket(bucket).list_blobs(prefix=prefix):
        if not blob.name.endswith(".jsonl"):
            continue
        for ln in blob.download_as_text().splitlines():
            if ln.strip():
                lines.append(json.loads(ln))
    return lines


def phase_fetch(session, gcs, bucket, output_dir, C, G, Y, R, W, D, only_jobs=None):
    state = load_state(output_dir)
    jobs = [j for j in state["jobs"] if not j.get("fetched")]
    # In `run` mode, only poll the jobs submitted in this invocation — not stale
    # unfetched jobs left over from earlier runs (use the standalone `fetch`
    # command, which has no filter, to drain those).
    if only_jobs is not None:
        jobs = [j for j in jobs if j["name"] in only_jobs]
    if not jobs:
        print(f"{G}No outstanding jobs to fetch.{Style.RESET_ALL}")
        return

    TIMINGS_CSV = output_dir / "timings.csv"
    _csv_is_new = not TIMINGS_CSV.exists()
    _fh = TIMINGS_CSV.open("a", newline="")
    _writer = csv.DictWriter(_fh, fieldnames=["folder_name", "job_name", "timestamp"])
    if _csv_is_new:
        _writer.writeheader(); _fh.flush()

    for job_entry in jobs:
        name = job_entry["name"]
        keys = job_entry["keys"]
        job_id = name.split("/")[-1]
        print(f"\n{C}Polling{Style.RESET_ALL} {W}{job_id}{Style.RESET_ALL} "
              f"({len(keys)} subjects)")

        # Poll using the resource name directly against the regional endpoint.
        get_url = _job_url(job_entry)
        job = get_json(session, get_url)
        while job.get("state") not in TERMINAL_STATES:
            print(f"  {D}state={job.get('state')} — waiting {POLL_INTERVAL_S}s{Style.RESET_ALL}")
            time.sleep(POLL_INTERVAL_S)
            job = get_json(session, get_url)

        state_name = job.get("state")
        if state_name != "JOB_STATE_SUCCEEDED":
            # A failed/expired job (e.g. code 13 INTERNAL) may still have written
            # some predictions to GCS before failing. Try to salvage whatever
            # rows exist; subjects with no output keep no file and are picked up
            # again on the next `submit`.
            print(f"  {R}{state_name}{Style.RESET_ALL} — {job.get('error', '')} "
                  f"{D}(attempting partial salvage){Style.RESET_ALL}")

        try:
            output_lines = _download_output_lines(gcs, bucket, job_entry["output_uri"])
        except Exception as e:                          # noqa: BLE001
            output_lines = []
            print(f"  {Y}WARN{Style.RESET_ALL} could not read outputs: {e}")

        if not output_lines:
            print(f"  {R}no output rows{Style.RESET_ALL} — nothing to save "
                  f"(re-run `submit` to retry these subjects)")
            job_entry["fetched"] = True
            save_state(output_dir, state)
            continue

        by_cid = {cid: ln for ln in output_lines
                  for cid in [ln.get("custom_id") or ln.get("id")] if cid}

        saved = 0
        for folder_name in keys:
            line = by_cid.get(folder_name)
            if line is None:
                print(f"  {Y}WARN{Style.RESET_ALL} {folder_name}: not in output")
                continue
            _, text, err = _extract_from_line(line)
            if text is None:
                print(f"  {R}ERROR{Style.RESET_ALL} {folder_name}: {err}")
                continue
            try:
                result = parse_model_json(text)
            except json.JSONDecodeError as e:
                print(f"  {R}ERROR{Style.RESET_ALL} {folder_name}: bad JSON ({e})")
                continue

            result["subject_id"] = folder_name
            result["_seed"] = SEED
            out_path = output_dir / f"{folder_name}_classification_{MODEL_SLUG}.json"
            with open(out_path, "w") as f:
                json.dump(result, f, indent=2)
            _writer.writerow({
                "folder_name": folder_name, "job_name": job_id,
                "timestamp": datetime.now().isoformat(timespec="seconds"),
            })
            _fh.flush()
            saved += 1

        job_entry["fetched"] = True
        save_state(output_dir, state)
        print(f"  {G}SAVED{Style.RESET_ALL} {saved}/{len(keys)} outputs")

    _fh.close()
    print(f"{C}Timings -> {TIMINGS_CSV}{Style.RESET_ALL}")


def _job_url(job_entry: dict) -> str:
    """
    Build the GET URL for a batchPredictionJob from its stored resource name
    `projects/<p>/locations/<loc>/batchPredictionJobs/<id>`.
    """
    name = job_entry["name"]
    # name contains locations/<loc>; the regional host must match that location.
    loc = name.split("/locations/")[1].split("/")[0]
    return f"https://{loc}-aiplatform.googleapis.com/v1/{name}"


# --- Phase: metrics (identical to gpt55 script) ------------------------------

def phase_metrics(output_dir, gt, C, G, D):
    print(f"\n{C}Collecting predictions for metric computation...{Style.RESET_ALL}")
    predictions: dict = {}
    for out_file in sorted(output_dir.glob(f"*_classification_{MODEL_SLUG}.json")):
        result = json.load(open(out_file))
        sid = result.get("subject_id",
                          out_file.stem.replace(f"_classification_{MODEL_SLUG}", ""))
        predictions[sid] = extract_predictions(result)
    if not predictions:
        print(f"{D}No output files found — run `submit` then `fetch` first.{Style.RESET_ALL}")
        return
    metrics = compute_metrics(predictions, gt)
    print_metrics(metrics)
    metrics_path = output_dir / "metrics.json"
    with open(metrics_path, "w") as f:
        json.dump(metrics, f, indent=2)
    print(f"{G}Metrics saved{Style.RESET_ALL} -> {metrics_path}")


# --- Phase: status -----------------------------------------------------------

def phase_status(session, output_dir, C, W, D):
    state = load_state(output_dir)
    if not state["jobs"]:
        print(f"{D}No jobs recorded in {state_path(output_dir)}{Style.RESET_ALL}")
        return
    for j in state["jobs"]:
        try:
            live = get_json(session, _job_url(j)).get("state", "<no state>")
        except Exception as e:                  # noqa: BLE001
            live = f"<lookup failed: {e}>"
        flag = "fetched" if j.get("fetched") else "pending"
        print(f"{C}{j['name'].split('/')[-1]}{Style.RESET_ALL}  chunk={j['chunk']}  "
              f"n={len(j['keys'])}  state={W}{live}{Style.RESET_ALL}  [{flag}]")


# --- CLI ---------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Classify gliomas via Vertex AI batch prediction "
                    "(Llama 4 Maverick, 50% cheaper).")
    p.add_argument("phase", choices=["submit", "fetch", "run", "metrics", "status"],
                   help="submit: upload+create job(s); fetch: poll+download+save; "
                        "run: submit then fetch then metrics; metrics: score saved "
                        "outputs; status: print live job states.")
    p.add_argument("--json-type", choices=["btreport", "btreport_pp"],
                   help="Which metadata JSON variant to use (required except `status`/`metrics`).")
    p.add_argument("--force", action="store_true",
                   help="Re-submit subjects even if an output file already exists.")
    return p.parse_args()


def main() -> None:
    colorama_init(autoreset=True)
    C, G, Y, R, M, W, D = (Fore.CYAN, Fore.GREEN, Fore.YELLOW, Fore.RED,
                           Fore.MAGENTA, Style.BRIGHT, Style.DIM)

    args = parse_args()
    if args.phase in ("submit", "fetch", "run") and not args.json_type:
        raise SystemExit("--json-type is required for this phase.")

    output_dir = HERE / (f"llama4_batch_outputs_{args.json_type}"
                         if args.json_type else "llama4_batch_outputs")
    output_dir.mkdir(parents=True, exist_ok=True)

    project, location, bucket = load_vertex_config()
    session = make_session()
    gcs = storage.Client(project=project)
    api = api_root(location, project)
    print(f"{C}Vertex AI batch  |  project={W}{project}{C}  region={W}{location}{C}  "
          f"bucket={W}{bucket}{C}\nModel: {W}{MODEL_ID}{C}  |  Output: {W}{output_dir.name}")

    id_map = load_id_map(COHORT_XLSX) if args.phase in ("submit", "run") else {}
    gt     = load_ground_truth(COHORT_XLSX) if args.phase in ("run", "metrics") else {}

    if args.phase == "submit":
        phase_submit(session, gcs, bucket, api, output_dir, args.json_type,
                     id_map, args.force, C, G, Y, W, D)
    elif args.phase == "fetch":
        phase_fetch(session, gcs, bucket, output_dir, C, G, Y, R, W, D)
    elif args.phase == "metrics":
        phase_metrics(output_dir, gt, C, G, D)
    elif args.phase == "status":
        phase_status(session, output_dir, C, W, D)
    elif args.phase == "run":
        submitted = phase_submit(session, gcs, bucket, api, output_dir,
                                 args.json_type, id_map, args.force, C, G, Y, W, D)
        # Only poll the jobs we just submitted (ignore stale unfetched jobs from
        # earlier runs — drain those with the standalone `fetch` command).
        phase_fetch(session, gcs, bucket, output_dir, C, G, Y, R, W, D,
                    only_jobs=set(submitted))
        phase_metrics(output_dir, gt, C, G, D)


if __name__ == "__main__":
    main()
