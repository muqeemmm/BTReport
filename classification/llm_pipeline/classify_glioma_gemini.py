"""
Glioma classification via the Gemini Batch API (gemini-2.5-pro by default).

This is the Gemini counterpart of `classify_glioma_groq.py`. It performs the
SAME task — joint prediction of IDH status, 1p/19q co-deletion and CNS WHO grade
from structured pre-operative MRI metadata, under WHO CNS5 (2021) — using the
SAME Chain-of-Thought Confidence Elicitation (CoT CE) schema, but routes
inference through Gemini's **asynchronous Batch API** to get the documented 50%
cost discount versus synchronous calls.

WHY THE STRUCTURE DIFFERS FROM THE GROQ SCRIPT
----------------------------------------------
Groq inference is synchronous: one request -> one response, looped per subject.
The Gemini Batch API is asynchronous and has three phases that cannot be a
single loop:

    1. SUBMIT  - bundle every subject into batch job(s) and hand them to Google.
    2. POLL    - wait (seconds to <=24h) until the job(s) reach a terminal state.
    3. FETCH   - download the results, write one JSON per subject, score them.

So this file is a small CLI with sub-commands (`submit`, `fetch`, `run`,
`metrics`, `status`) plus a persisted `batch_state.json` so a submitted job can
be picked up later from a different process.

WHY INLINE REQUESTS + CHUNKING (not a JSONL input file)
-------------------------------------------------------
The Batch API accepts either inline requests or an uploaded JSONL file. We use
inline requests because the google-genai SDK lets us pass a **Pydantic class**
as `response_schema`, and the SDK serialises it to Gemini's schema format
correctly *including field order*. Field order is load-bearing here: CoT CE
requires `reasoning` / `supporting_features` / `contradicting_features` to be
generated BEFORE `prediction` and `confidence`. Pydantic preserves declaration
order, so the model is forced to reason first. (A hand-written JSON schema does
not guarantee key order, and `propertyOrdering` would have to be set by hand.)

Inline payloads are capped at ~20 MB per job, so we split the cohort into chunks
(default 40 subjects/job) and submit one batch job per chunk. The JSONL-file
alternative (single keyed job, 2 GB limit) is described in the accompanying
guideline if you outgrow this.

Schema, prompt and metrics are intentionally identical to the Groq script so the
two models are directly comparable.
"""

import argparse
import csv
import json
import time
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import List, Literal

import openpyxl
from colorama import Fore, Style, init as colorama_init
from dotenv import dotenv_values
from google import genai
from google.genai import types
from pydantic import BaseModel, Field
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score
from tqdm import tqdm


# --- Prompt (identical to classify_glioma_groq.py) ---------------------------

SYSTEM_PROMPT = (
    "You are an expert neuroradiologist performing pre-operative molecular "
    "subtyping of adult-type diffuse glioma from structured MRI metadata, "
    "following the WHO CNS5 (2021) classification. For every task you reason "
    "step by step BEFORE stating any prediction or confidence score: first lay "
    "out the clinically grounded explanation, then commit to the prediction, "
    "then assign the confidence. "
    "CRITICAL: Each classification task (IDH status, 1p/19q co-deletion, WHO "
    "grade) must be performed INDEPENDENTLY. Reason about each task separately "
    "using only the raw MRI metadata — do NOT let the prediction or confidence "
    "from one task influence the reasoning, prediction, or confidence of another. "
    "Output MUST be a single valid JSON object that "
    "conforms exactly to the GliomaClassificationOutput schema. No markdown, no "
    "prose outside the JSON, no code fences -- JSON only."
)

USER_PROMPT_TEMPLATE = """# ROLE
You are an expert neuroradiologist.

# TASK
You are given structured pre-operative brain MRI metadata for a single patient
with a suspected **adult-type diffuse glioma**, defined according to the WHO
CNS5 molecular classification. Based on this metadata, predict the following:

  (a) IDH mutation status        -> {{"mutant", "wildtype"}}
  (b) 1p/19q co-deletion status  -> {{"codeleted", "non-codeleted"}}
  (c) CNS WHO grade              -> {{2, 3, 4}}

# CHAIN-OF-THOUGHT CONFIDENCE ELICITATION
For each of the three tasks you MUST work in the following order. Do not skip,
reorder, or shortcut these steps:

  1. **Reasoning first.** Write a concise, clinically grounded explanation that
     weighs the available metadata. Explicitly contrast the evidence that
     supports each candidate answer against the evidence that argues against
     it. Reach the prediction as the *conclusion* of this reasoning.
  2. **Supporting metadata fields.** List the exact field names (verbatim) that
     support the prediction you reasoned toward.
  3. **Contradicting features.** List the fields that argue against it.
  4. **Prediction.** State the prediction that follows from steps 1-3.
  5. **Confidence score** (between 0.0 and 1.0). Derive the score *from* the
     reasoning above -- it must reflect the balance of supporting vs.
     contradicting evidence you already articulated, NOT a number chosen before
     reasoning. Assign high confidence only when key features are well
     supported and few contradict; assign low confidence when evidence is
     thin, conflicting, or absent.

The JSON schema lists `reasoning`, `supporting_features` and
`contradicting_features` before `prediction` and `confidence` for exactly this
reason: generate them in that order.

# INPUT CONSTRAINTS
- Input consists **only of structured MRI-derived metadata** (no images).
- Do not assume access to data beyond what is provided.
- Treat null values as **"not assessed"** (do not infer).

# AVAILABLE MRI SEQUENCES
T1w, T2w, T2-FLAIR, T1-Gd.
**Not available:** DWI/ADC, perfusion (rCBV/DSC), MRS (2HG), SWI/GRE.

# RULES
1. Use ONLY the metadata provided; no hallucinated features or sequences.
2. Every claim in `reasoning` must be traceable to a field named verbatim in
   `supporting_features`.
3. Do NOT invoke diffusion, perfusion, spectroscopy, calcifications, or SWI.
4. If evidence is insufficient: give a best estimate, set confidence in
   0.50-0.60, and state the uncertainty in reasoning.
5. Confidence must be >= 0.50 for binary tasks (otherwise flip the prediction).
6. The confidence score must be the conclusion of the reasoning, not its
   premise -- never write the reasoning to justify a pre-chosen score.
7. Output must be **valid JSON only** (no markdown, no commentary).

# INPUT
Patient subject_id: {subject_id}
MRI metadata:
{metadata_json}
"""


# --- Structured-output schema (Pydantic mirror of glioma_classification_schema.json)
# Field declaration order == generation order == CoT CE order. Do not reorder.

class TaskIDH(BaseModel):
    reasoning: str = Field(..., min_length=1,
                           description="Clinically grounded chain of thought, written BEFORE the prediction.")
    supporting_features: List[str] = Field(default_factory=list)
    contradicting_features: List[str] = Field(default_factory=list)
    prediction: Literal["mutant", "wildtype"]
    confidence: float = Field(..., ge=0.0, le=1.0)


class TaskCodeletion(BaseModel):
    reasoning: str = Field(..., min_length=1,
                           description="Clinically grounded chain of thought, written BEFORE the prediction.")
    supporting_features: List[str] = Field(default_factory=list)
    contradicting_features: List[str] = Field(default_factory=list)
    prediction: Literal["codeleted", "non-codeleted"]
    confidence: float = Field(..., ge=0.0, le=1.0)


class TaskGrade(BaseModel):
    reasoning: str = Field(..., min_length=1,
                           description="Clinically grounded chain of thought, written BEFORE the prediction.")
    supporting_features: List[str] = Field(default_factory=list)
    contradicting_features: List[str] = Field(default_factory=list)
    # Gemini enum support is for STRING types; keep grade a plain int (2/3/4)
    # constrained in the prompt and validated downstream, mirroring the GT ints.
    prediction: int = Field(..., description="CNS WHO grade: one of 2, 3, or 4.")
    confidence: float = Field(..., ge=0.0, le=1.0)


class GliomaClassificationOutput(BaseModel):
    subject_id: str
    idh: TaskIDH
    one_p_nineteen_q: TaskCodeletion
    who_grade: TaskGrade
    integrated_diagnosis: str = Field(
        ..., description="Full WHO-CNS5 integrated diagnosis string.")
    consistency_check_passed: bool
    overall_comment: str


# --- Config ------------------------------------------------------------------

HERE        = Path(__file__).parent
ENV_PATH    = Path(".env")
JSONS_DIR   = Path("data/Dataset_AKU_WHO/JSONs")
COHORT_XLSX = Path("data/dataset_table/cohort_merged.xlsx")

DEFAULT_MODEL    = "gemini-2.5-pro"
CHUNK_SIZE       = 40        # subjects per inline batch job (keep payload < 20MB)
POLL_INTERVAL_S  = 30        # seconds between status polls

TERMINAL_STATES = {
    "JOB_STATE_SUCCEEDED", "JOB_STATE_FAILED",
    "JOB_STATE_CANCELLED", "JOB_STATE_EXPIRED",
}


def model_slug(model: str) -> str:
    """gemini-2.5-pro -> gemini_2_5_pro (for filenames / keys)."""
    return model.replace("-", "_").replace(".", "_")


# --- ID helpers (identical to Groq script) -----------------------------------

def load_id_map(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    id_map: dict = {}
    for row in list(ws.iter_rows(values_only=True))[1:]:
        center_id  = str(row[0]).strip() if row[0] else None
        subject_id = str(row[1]).strip() if row[1] else None
        brats_id   = str(row[4]).strip() if (row[4] and row[4] != "x") else None
        entry = {"center_id": center_id, "subject_id": subject_id}
        if center_id:
            id_map[center_id] = entry
        if brats_id:
            id_map[brats_id] = entry
    return id_map


# --- Ground-truth helpers (identical to Groq script) --------------------------

def load_ground_truth(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def _idh(v):
        if v == "mutated":   return "mutant"
        if v == "wild type": return "wildtype"
        return None

    def _codeletion(v):
        if v in ("co-deleted", "rel. co-deleted"): return "codeleted"
        if v == "intact":                          return "non-codeleted"
        return None

    gt: dict = {}
    for row in rows[1:]:
        center_id = row[0]
        brats_id  = row[4] if (row[4] and row[4] != "x") else None
        entry = {
            "idh":        _idh(row[7]),
            "codeletion": _codeletion(row[8]),
            "grade":      row[9] if isinstance(row[9], int) else None,
            "dataset":    row[2],
        }
        if center_id:
            gt[center_id] = entry
        if brats_id:
            gt[brats_id] = entry
    return gt


# --- Metrics (identical to Groq script) --------------------------------------

def _binary_metrics(y_true, y_pred, pos_label) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None}
    all_labels = sorted(set(y_true) | set(y_pred))
    neg_candidates = [l for l in all_labels if l != pos_label]
    neg_label = neg_candidates[0] if neg_candidates else (
        "wildtype" if pos_label == "mutant" else "non-codeleted"
    )
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, pos_label=pos_label, average="binary",
                   zero_division=0)
    cm  = confusion_matrix(y_true, y_pred, labels=[pos_label, neg_label])
    TP, FN = cm[0, 0], cm[0, 1]
    FP, TN = cm[1, 0], cm[1, 1]
    sens = TP / (TP + FN) if (TP + FN) > 0 else None
    spec = TN / (TN + FP) if (TN + FP) > 0 else None
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sens, 4) if sens is not None else None,
        "specificity": round(spec, 4) if spec is not None else None,
        "f1":          round(f1, 4),
    }


def _grade_metrics(y_true, y_pred) -> dict:
    if not y_true:
        return {"n": 0, "accuracy": None, "sensitivity": None,
                "specificity": None, "f1": None}
    classes = [2, 3, 4]
    acc = accuracy_score(y_true, y_pred)
    f1  = f1_score(y_true, y_pred, labels=classes, average="macro",
                   zero_division=0)
    cm = confusion_matrix(y_true, y_pred, labels=classes)
    sens_list, spec_list = [], []
    for i in range(len(classes)):
        TP = cm[i, i]
        FN = cm[i, :].sum() - TP
        FP = cm[:, i].sum() - TP
        TN = cm.sum() - TP - FN - FP
        sens_list.append(TP / (TP + FN) if (TP + FN) > 0 else 0.0)
        spec_list.append(TN / (TN + FP) if (TN + FP) > 0 else 0.0)
    return {
        "n":           len(y_true),
        "accuracy":    round(acc, 4),
        "sensitivity": round(sum(sens_list) / len(sens_list), 4),
        "specificity": round(sum(spec_list) / len(spec_list), 4),
        "f1":          round(f1, 4),
    }


def compute_metrics(predictions: dict, gt: dict) -> dict:
    data = defaultdict(lambda: {
        "idh_t": [], "idh_p": [],
        "cod_t": [], "cod_p": [],
        "grd_t": [], "grd_p": [],
    })
    for sid, pred in predictions.items():
        row = gt.get(sid)
        if row is None:
            continue
        ds = row["dataset"] or "Unknown"
        for bucket in ("whole_cohort", ds):
            d = data[bucket]
            if row["idh"] is not None and pred.get("idh") is not None:
                d["idh_t"].append(row["idh"]); d["idh_p"].append(pred["idh"])
            if row["codeletion"] is not None and pred.get("codeletion") is not None:
                d["cod_t"].append(row["codeletion"]); d["cod_p"].append(pred["codeletion"])
            if row["grade"] is not None and pred.get("grade") is not None:
                d["grd_t"].append(row["grade"]); d["grd_p"].append(pred["grade"])

    results = {}
    for bucket, d in data.items():
        results[bucket] = {
            "idh":        _binary_metrics(d["idh_t"], d["idh_p"], "mutant"),
            "codeletion": _binary_metrics(d["cod_t"], d["cod_p"], "codeleted"),
            "grade":      _grade_metrics(d["grd_t"], d["grd_p"]),
        }
    return results


def print_metrics(metrics: dict) -> None:
    col_w = 22
    header = (f"{'Subset':<{col_w}} {'Task':<12} {'N':>5}  "
              f"{'Acc':>6}  {'Sens':>6}  {'Spec':>6}  {'F1':>6}")
    sep = "=" * len(header)
    print(f"\n{sep}\nCLASSIFICATION METRICS\n{sep}")
    print(header)
    print("-" * len(header))
    task_labels = {"idh": "IDH", "codeletion": "1p/19q", "grade": "Grade"}
    buckets = ["whole_cohort"] + sorted(k for k in metrics if k != "whole_cohort")
    for bucket in buckets:
        for task_key, label in task_labels.items():
            m = metrics[bucket][task_key]
            def _fmt(v):
                return f"{v:.4f}" if v is not None else "  N/A"
            print(f"{bucket:<{col_w}} {label:<12} {m['n']:>5}  "
                  f"{_fmt(m['accuracy']):>6}  {_fmt(m['sensitivity']):>6}  "
                  f"{_fmt(m['specificity']):>6}  {_fmt(m['f1']):>6}")
        print()


# --- Gemini client -----------------------------------------------------------

def make_client() -> genai.Client:
    """
    Build a Gemini client. Reads GEMINI_API_KEY (or GOOGLE_API_KEY) from .env,
    falling back to the process environment if not present in .env.
    """
    config = dotenv_values(ENV_PATH)
    api_key = (config.get("GEMINI_API_KEY") or config.get("GOOGLE_API_KEY"))
    if api_key:
        return genai.Client(api_key=api_key)
    # genai.Client() will read GEMINI_API_KEY/GOOGLE_API_KEY from the env.
    return genai.Client()


def build_request_config() -> dict:
    """Per-request config dict for an inline batch request (CoT-CE structured output)."""
    return {
        "system_instruction": SYSTEM_PROMPT,
        "temperature": 0.0,
        "seed": 42,
        "max_output_tokens": 8192,
        "response_mime_type": "application/json",
        "response_schema": GliomaClassificationOutput,
    }


# --- State persistence -------------------------------------------------------

def state_path(output_dir: Path) -> Path:
    return output_dir / "batch_state.json"


def load_state(output_dir: Path) -> dict:
    p = state_path(output_dir)
    if p.exists():
        return json.load(open(p))
    return {"jobs": []}


def save_state(output_dir: Path, state: dict) -> None:
    with open(state_path(output_dir), "w") as f:
        json.dump(state, f, indent=2)


# --- Subject collection ------------------------------------------------------

def collect_subjects(json_type: str, output_dir: Path, model: str,
                     id_map: dict, force: bool) -> list:
    """
    Returns a list of dicts {folder_name, api_subject_id, prompt} for every
    subject still needing a prediction (skips ones whose output already exists,
    unless --force).
    """
    slug = model_slug(model)
    subjects = sorted(d.name for d in JSONS_DIR.iterdir() if d.is_dir())
    pending = []
    for folder_name in subjects:
        out_path = output_dir / f"{folder_name}_classification_{slug}.json"
        if out_path.exists() and not force:
            continue
        metadata_path = JSONS_DIR / folder_name / f"{folder_name}_metadata_{json_type}.json"
        if not metadata_path.exists():
            continue
        ids = id_map.get(folder_name)
        api_subject_id = ids["subject_id"] if ids else folder_name
        metadata = json.load(open(metadata_path))
        prompt = USER_PROMPT_TEMPLATE.format(
            subject_id=api_subject_id,
            metadata_json=json.dumps(metadata, indent=2, default=str),
        )
        pending.append({
            "folder_name": folder_name,
            "api_subject_id": api_subject_id,
            "prompt": prompt,
        })
    return pending


def chunked(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


# --- Phase: submit -----------------------------------------------------------

def phase_submit(client, output_dir, json_type, model, id_map, force, C, G, Y, W, D):
    pending = collect_subjects(json_type, output_dir, model, id_map, force)
    if not pending:
        print(f"{G}Nothing to submit{Style.RESET_ALL} — all subjects already have outputs.")
        return

    print(f"{C}Submitting {W}{len(pending)}{C} subjects in chunks of {CHUNK_SIZE} "
          f"using {W}{model}{C} (batch = 50% off).{Style.RESET_ALL}")

    state = load_state(output_dir)
    req_cfg = build_request_config()

    for ci, chunk in enumerate(chunked(pending, CHUNK_SIZE), 1):
        inline_requests = []
        keys = []
        for s in chunk:
            inline_requests.append({
                "contents": [{"parts": [{"text": s["prompt"]}], "role": "user"}],
                "config": req_cfg,
            })
            keys.append(s["folder_name"])

        job = client.batches.create(
            model=model,
            src=inline_requests,
            config={"display_name": f"glioma-{model_slug(model)}-{json_type}-chunk{ci}"},
        )
        state["jobs"].append({
            "name": job.name,
            "model": model,
            "json_type": json_type,
            "chunk": ci,
            "keys": keys,            # ordered; inline responses come back in this order
            "fetched": False,
            "submitted_at": datetime.now().isoformat(timespec="seconds"),
        })
        save_state(output_dir, state)
        print(f"  {G}SUBMITTED{Style.RESET_ALL} chunk {ci} "
              f"({len(keys)} subjects) -> {W}{job.name}{Style.RESET_ALL}")

    print(f"{C}Saved job state -> {state_path(output_dir)}{Style.RESET_ALL}")
    print(f"{D}Now run with the same args using `fetch` (or `run`) to poll & "
          f"collect results.{Style.RESET_ALL}")


# --- Phase: poll + fetch -----------------------------------------------------

def _extract_json_text(inline_response):
    """Return the model's JSON string from an inline response, or None on error."""
    if getattr(inline_response, "error", None):
        return None, str(inline_response.error)
    resp = getattr(inline_response, "response", None)
    if resp is None:
        return None, "empty response"
    try:
        return resp.text, None
    except Exception as e:                      # noqa: BLE001
        return None, f"no text ({e})"


def phase_fetch(client, output_dir, model, C, G, Y, R, W, D):
    slug = model_slug(model)
    state = load_state(output_dir)
    jobs = [j for j in state["jobs"] if not j.get("fetched")]
    if not jobs:
        print(f"{G}No outstanding jobs to fetch.{Style.RESET_ALL}")
        return

    TIMINGS_CSV = output_dir / "timings.csv"
    _csv_is_new = not TIMINGS_CSV.exists()
    _fh = TIMINGS_CSV.open("a", newline="")
    _writer = csv.DictWriter(_fh, fieldnames=["folder_name", "job_name", "timestamp"])
    if _csv_is_new:
        _writer.writeheader(); _fh.flush()

    for job_entry in jobs:
        job_name = job_entry["name"]
        keys = job_entry["keys"]
        print(f"\n{C}Polling{Style.RESET_ALL} {W}{job_name}{Style.RESET_ALL} "
              f"({len(keys)} subjects)")

        batch_job = client.batches.get(name=job_name)
        while batch_job.state.name not in TERMINAL_STATES:
            print(f"  {D}state={batch_job.state.name} — waiting {POLL_INTERVAL_S}s{Style.RESET_ALL}")
            time.sleep(POLL_INTERVAL_S)
            batch_job = client.batches.get(name=job_name)

        if batch_job.state.name != "JOB_STATE_SUCCEEDED":
            print(f"  {R}{batch_job.state.name}{Style.RESET_ALL} — "
                  f"{getattr(batch_job, 'error', None)}")
            continue

        responses = batch_job.dest.inlined_responses or []
        if len(responses) != len(keys):
            print(f"  {Y}WARN{Style.RESET_ALL} response count {len(responses)} "
                  f"!= subject count {len(keys)} — mapping by position up to min().")

        saved = 0
        for folder_name, inline_response in zip(keys, responses):
            text, err = _extract_json_text(inline_response)
            if text is None:
                print(f"  {R}ERROR{Style.RESET_ALL} {folder_name}: {err}")
                continue
            try:
                result = json.loads(text)
            except json.JSONDecodeError as e:
                print(f"  {R}ERROR{Style.RESET_ALL} {folder_name}: bad JSON ({e})")
                continue

            result["subject_id"] = folder_name   # for the metrics join
            out_path = output_dir / f"{folder_name}_classification_{slug}.json"
            with open(out_path, "w") as f:
                json.dump(result, f, indent=2)
            _writer.writerow({
                "folder_name": folder_name,
                "job_name": job_name,
                "timestamp": datetime.now().isoformat(timespec="seconds"),
            })
            _fh.flush()
            saved += 1

        job_entry["fetched"] = True
        save_state(output_dir, state)
        print(f"  {G}SAVED{Style.RESET_ALL} {saved}/{len(keys)} outputs")

    _fh.close()
    print(f"{C}Timings -> {TIMINGS_CSV}{Style.RESET_ALL}")


# --- Phase: metrics ----------------------------------------------------------

def phase_metrics(output_dir, model, gt, C, G, D):
    slug = model_slug(model)
    print(f"\n{C}Collecting predictions for metric computation...{Style.RESET_ALL}")
    predictions: dict = {}
    for out_file in sorted(output_dir.glob(f"*_classification_{slug}.json")):
        result = json.load(open(out_file))
        sid = result.get("subject_id",
                          out_file.stem.replace(f"_classification_{slug}", ""))
        predictions[sid] = {
            "idh":        result.get("idh", {}).get("prediction"),
            "codeletion": result.get("one_p_nineteen_q", {}).get("prediction"),
            "grade":      result.get("who_grade", {}).get("prediction"),
        }
    if not predictions:
        print(f"{D}No output files found — run `submit` then `fetch` first.{Style.RESET_ALL}")
        return
    metrics = compute_metrics(predictions, gt)
    print_metrics(metrics)
    metrics_path = output_dir / "metrics.json"
    with open(metrics_path, "w") as f:
        json.dump(metrics, f, indent=2)
    print(f"{G}Metrics saved{Style.RESET_ALL} -> {metrics_path}")


# --- Phase: status -----------------------------------------------------------

def phase_status(client, output_dir, C, W, D):
    state = load_state(output_dir)
    if not state["jobs"]:
        print(f"{D}No jobs recorded in {state_path(output_dir)}{Style.RESET_ALL}")
        return
    for j in state["jobs"]:
        try:
            live = client.batches.get(name=j["name"]).state.name
        except Exception as e:                 # noqa: BLE001
            live = f"<lookup failed: {e}>"
        flag = "fetched" if j.get("fetched") else "pending"
        print(f"{C}{j['name']}{Style.RESET_ALL}  chunk={j['chunk']}  "
              f"n={len(j['keys'])}  state={W}{live}{Style.RESET_ALL}  [{flag}]")


# --- CLI ---------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Classify gliomas via the Gemini Batch API (50% cheaper).")
    p.add_argument("phase", choices=["submit", "fetch", "run", "metrics", "status"],
                   help="submit: create batch job(s); fetch: poll+download+save; "
                        "run: submit then fetch then metrics; metrics: score saved "
                        "outputs; status: print live job states.")
    p.add_argument("--json-type", choices=["btreport", "btreport_pp"],
                   help="Which metadata JSON variant to use (required except for `status`).")
    p.add_argument("--model", default=DEFAULT_MODEL,
                   help=f"Gemini model id (default {DEFAULT_MODEL}).")
    p.add_argument("--force", action="store_true",
                   help="Re-submit subjects even if an output file already exists.")
    return p.parse_args()


def main() -> None:
    colorama_init(autoreset=True)
    C, G, Y, R, M, W, D = (Fore.CYAN, Fore.GREEN, Fore.YELLOW, Fore.RED,
                           Fore.MAGENTA, Style.BRIGHT, Style.DIM)

    args = parse_args()
    if args.phase != "status" and not args.json_type:
        raise SystemExit("--json-type is required for this phase.")

    output_dir = HERE / f"gemini_outputs_{args.json_type}" if args.json_type \
        else HERE / "gemini_outputs"
    output_dir.mkdir(parents=True, exist_ok=True)

    client = make_client()
    id_map = load_id_map(COHORT_XLSX) if args.phase in ("submit", "run") else {}
    gt     = load_ground_truth(COHORT_XLSX) if args.phase in ("run", "metrics") else {}

    if args.phase == "submit":
        phase_submit(client, output_dir, args.json_type, args.model, id_map,
                     args.force, C, G, Y, W, D)
    elif args.phase == "fetch":
        phase_fetch(client, output_dir, args.model, C, G, Y, R, W, D)
    elif args.phase == "metrics":
        phase_metrics(output_dir, args.model, gt, C, G, D)
    elif args.phase == "status":
        phase_status(client, output_dir, C, W, D)
    elif args.phase == "run":
        phase_submit(client, output_dir, args.json_type, args.model, id_map,
                     args.force, C, G, Y, W, D)
        phase_fetch(client, output_dir, args.model, C, G, Y, R, W, D)
        phase_metrics(output_dir, args.model, gt, C, G, D)


if __name__ == "__main__":
    main()
